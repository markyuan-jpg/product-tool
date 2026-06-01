# -*- coding: utf-8 -*-

import os, sys, json, re, logging
import asyncio
import functools
import sqlite3
from contextlib import asynccontextmanager
from datetime import datetime

from pathlib import Path


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

logger = logging.getLogger(__name__)


# Load .env

try:

    from dotenv import load_dotenv

    load_dotenv()

except ImportError:

    pass


# Add product_tool/src to path

PROJECT_ROOT = Path(__file__).resolve().parent.parent

SRC_DIR = PROJECT_ROOT / "product_tool" / "src"

sys.path.insert(0, str(SRC_DIR))

sys.path.insert(0, str(PROJECT_ROOT / "product_tool"))


from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, Request, Depends, Header, Body
from sqlalchemy.ext.asyncio import AsyncSession

from fastapi.middleware.cors import CORSMiddleware

from fastapi.responses import FileResponse, JSONResponse

from fastapi.staticfiles import StaticFiles

import pandas as pd

from openpyxl import load_workbook


#  product_tool 模块级导入(缓存,避免每个请求重import

from src.company import load_company, save_company

from src.core.image import match_sku_folder

from src.output.quotation_excel import create_quotation

from src.output.pi_generator import generate_pi_xlsx

from src.output.pdf_generator import create_quote_pdf

from src.packing.generator import generate_packing_list, generate_commercial_invoice

from universal_parser import parse as universal_parse, sheet_to_markdown

from universal_parser import detect_header_row, parse_with_colmap, score_result

from ai_parser import load_cache, save_cache, ai_detect_columns, parse_text_to_products, get_cache_key as ai_cache_key
from score import score_dataframe
from src.rates import get_rate as _get_exchange_rate


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: initialize DB tables. Shutdown: no-op."""
    try:
        from database import init_db as _init
        await _init()
        logger.info("Database tables created/verified")
    except Exception as e:
        logger.error(f"Database init on startup failed: {e}")
    yield

app = FastAPI(title="报价整合工具 API", version="1.0.0", max_upload_size=100_000_000, lifespan=lifespan)


#  统一上传大小限制 

MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50MB

DEFAULT_PAYMENT_TERMS = "本合同签订七个工作日内支付定金30%,收到定金后30日内交货,发货前付清剩余70%余款 Terms of payment: The 30% deposit shall be paid within 7 working days after contract signed, with delivery to be completed within 60 days upon receipt of the deposit. The remaining 70% balance must be paid in full prior to shipment."


def compute_coverage(df) -> float:

    """计算 AI 解析覆盖关键列非空比"""

    if not isinstance(df, pd.DataFrame):
        return 0.0

    if df is None or df.empty:

        return 0.0

    key_cols = ['model', 'name_zh', 'price_rmb']

    non_null = sum(df[c].notna().sum() for c in key_cols if c in df.columns)

    total = sum(len(df) for c in key_cols if c in df.columns)

    return non_null / total if total > 0 else 0.0


async def limit_upload_size(file: UploadFile) -> bytes:

    """读取上传文件并检查不超过 MAX_UPLOAD_BYTES"""

    content = await file.read()

    if len(content) > MAX_UPLOAD_BYTES:

        raise HTTPException(413, f"文件超过 {MAX_UPLOAD_BYTES // 1024 // 1024}MB 限制,请压缩后重试")

    return content


#  统一公司配置加载 

_COMPANY_CACHE = None
_COMPANY_CACHE_TS = 0.0
_SKU_INDEX_CACHE = {}
_SKU_INDEX_CACHE_TS = 0.0

def _load_company_config() -> dict:

    """统一加载公司配置(内存缓存 30 秒，避免每个请求读磁盘)"""

    import time
    global _COMPANY_CACHE, _COMPANY_CACHE_TS
    now = time.time()
    if _COMPANY_CACHE is not None and now - _COMPANY_CACHE_TS < 30:
        return _COMPANY_CACHE

    try:

        cfg = load_company()
        _COMPANY_CACHE = cfg
        _COMPANY_CACHE_TS = now
        return cfg

    except Exception as e:

        logger.error(f"Failed to load company config: {e}")

        return {}


# Increase max request body size

import starlette.datastructures

starlette.datastructures.MAX_MEMORY_SIZE = 100_000_000  # 100MB for form fields


# Initialize DBs on startup (fail gracefully if PostgreSQL unavailable)

BASE_URL = os.getenv("BASE_URL", "")
CORS_ORIGINS = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://127.0.0.1:8000",
    "http://localhost:8000",
]
if BASE_URL:
    # BASE_URL should be the frontend domain, e.g. https://quoteflow.it.com
    CORS_ORIGINS.append(BASE_URL)
    WWW_BASE_URL = BASE_URL.replace("://", "://www.")
    if WWW_BASE_URL != BASE_URL:
        CORS_ORIGINS.append(WWW_BASE_URL)
else:
    # Fallback: allow both www and non-www for the production frontend
    CORS_ORIGINS.append("https://quoteflow.it.com")
    CORS_ORIGINS.append("https://www.quoteflow.it.com")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# CORS 异常处理:确保错误响应也包含跨域头
@app.exception_handler(Exception)
async def cors_error_handler(request, exc):
    from fastapi.responses import JSONResponse
    from starlette.middleware.cors import CORSMiddleware as _CORS
    # 手动添加CORS头（从请求Origin匹配或取生产域名）
    origin = request.headers.get("origin", "")
    allowed = origin if origin in CORS_ORIGINS else CORS_ORIGINS[-1]
    headers = {
        "Access-Control-Allow-Origin": allowed,
        "Access-Control-Allow-Credentials": "true",
        "Access-Control-Allow-Methods": "*",
        "Access-Control-Allow-Headers": "*",
    }
    status = 500
    detail = str(exc) if str(exc) else "Internal Server Error"
    if hasattr(exc, 'status_code'):
        status = exc.status_code
        detail = exc.detail if hasattr(exc, 'detail') else detail
    return JSONResponse(status_code=status, content={"detail": detail}, headers=headers)


#  安全头中间件 

@app.middleware("http")

async def add_security_headers(request: Request, call_next):

    response = await call_next(request)

    response.headers["X-Content-Type-Options"] = "nosniff"

    response.headers["X-Frame-Options"] = "DENY"

    response.headers["X-XSS-Protection"] = "1; mode=block"

    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"

    return response


UPLOAD_DIR = Path(__file__).parent / "uploads"

OUTPUT_DIR = Path(__file__).parent / "outputs"

UPLOAD_DIR.mkdir(exist_ok=True)

OUTPUT_DIR.mkdir(exist_ok=True)


#  Auth (must be before payment — used in Depends())

from auth import (
    hash_password, verify_password, create_access_token, create_refresh_token,
    get_current_user, get_current_user_optional, require_pro,
    get_user_by_username
)
from sqlalchemy.ext.asyncio import AsyncSession
from database import get_session, User
from sqlalchemy import select

#  Payment

from payment import create_checkout_session, verify_webhook, handle_webhook_event


@app.post("/api/payment/create-checkout")

async def payment_checkout(user: User = Depends(get_current_user)):
    """Create Creem checkout session for Pro upgrade."""
    result = create_checkout_session(
        user_id=user.id,
        email=user.email or '',
    )
    if not result:
        raise HTTPException(502, "支付服务暂不可用，请稍后再试")
    return {"url": result.get('checkout_url') or result.get('url', '')}


@app.post("/api/payment/webhook")

async def payment_webhook(request: Request, db: AsyncSession = Depends(get_session)):
    """Creem webhook handler — receives subscription lifecycle events."""
    payload = await request.body()
    signature = request.headers.get('creem-signature', '')

    if not verify_webhook(payload, signature):
        raise HTTPException(400, "Invalid signature")

    try:
        body = json.loads(payload)
    except Exception:
        raise HTTPException(400, "Invalid JSON")

    event = body.get('type', '') or body.get('event', '')
    data = body.get('data', {}) or body

    action = handle_webhook_event(event, data)
    if not action:
        return {"status": "ignored"}

    parts = action.split(':')
    action_type = parts[0]
    user_id = int(parts[1])

    user = await db.get(User, user_id)
    if not user:
        logger.warning(f"Webhook user not found: {user_id}")
        return {"status": "user_not_found"}

    if action_type == 'activate':
        user.tier = 'pro'
        logger.info(f"User {user_id} upgraded to pro via Creem")
    elif action_type == 'deactivate':
        user.tier = 'free'
        logger.info(f"User {user_id} downgraded to free")

    await db.commit()
    return {"status": "ok"}


#  Health 

@app.get("/api/health")

def health():

    return {"status": "ok", "time": datetime.utcnow().isoformat()}


@app.get("/api/exchange-rate")

async def exchange_rate(from_currency: str = Query("USD"), to_currency: str = Query("CNY")):
    """Get real-time exchange rate between two currencies."""
    try:
        rate = _get_exchange_rate(from_currency.upper(), to_currency.upper())
        return {"from": from_currency.upper(), "to": to_currency.upper(), "rate": rate, "source": "api.exchangerate.host"}
    except Exception as e:
        raise HTTPException(502, f"获取汇率失败: {str(e)}")


#  Auth endpoints

@app.post("/api/auth/register")

async def register(
    username: str = Form(...),
    password: str = Form(...),
    db: AsyncSession = Depends(get_session),
):
    if not re.match(r'^[a-zA-Z0-9_]+$', username):
        raise HTTPException(400, "用户名仅限英文,数字,下划线")
    if len(password) < 6:
        raise HTTPException(400, "密码至少 6 位")

    existing = await get_user_by_username(db, username)
    if existing:
        raise HTTPException(400, "用户名已存在")

    pw_hash = hash_password(password)
    new_user = User(username=username, password_hash=pw_hash, tier='free')
    db.add(new_user)
    await db.commit()
    await db.refresh(new_user)

    access_token = create_access_token(new_user.id)
    refresh_token = create_refresh_token(new_user.id)

    resp = JSONResponse({
        "token": access_token,
        "user": {"id": new_user.id, "username": new_user.username, "tier": new_user.tier}
    })
    resp.set_cookie(
        key="refresh_token", value=refresh_token,
        httponly=True, secure=True, samesite="strict",
        path="/api/auth", max_age=7*24*3600
    )
    return resp


@app.post("/api/auth/login")

async def login(
    username: str = Form(...),
    password: str = Form(...),
    db: AsyncSession = Depends(get_session),
):
    user = await get_user_by_username(db, username)
    if not user:
        raise HTTPException(400, "用户名或密码错误")
    if not verify_password(password, user.password_hash):
        raise HTTPException(400, "用户名或密码错误")

    access_token = create_access_token(user.id)
    refresh_token = create_refresh_token(user.id)

    resp = JSONResponse({
        "token": access_token,
        "user": {"id": user.id, "username": user.username, "tier": user.tier}
    })
    resp.set_cookie(
        key="refresh_token", value=refresh_token,
        httponly=True, secure=True, samesite="strict",
        path="/api/auth", max_age=7*24*3600
    )
    return resp


@app.get("/api/user/me")

async def user_me(user: User = Depends(get_current_user)):
    return {
        "id": user.id, "username": user.username, "tier": user.tier,
        "upload_count": user.upload_count, "email": user.email or ''
    }


@app.get("/api/user/usage")

async def user_usage(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_session)):
    from sqlalchemy import func, select
    from database import Product
    product_count = 0
    try:
        result = await db.execute(select(func.count(Product.id)).where(Product.user_id == user.id))
        product_count = result.scalar() or 0
    except Exception:
        pass
    return {
        "upload_count": user.upload_count, "limit": 20,
        "product_count": product_count, "product_limit": 200,
        "tier": user.tier
    }


@app.post("/api/auth/refresh")

async def refresh_token(request: Request, db: AsyncSession = Depends(get_session)):
    """Refresh access token using httpOnly refresh_token cookie."""
    refresh_token_str = request.cookies.get('refresh_token')
    if not refresh_token_str:
        raise HTTPException(401, "未登录")

    from auth import decode_token, get_user_by_id, create_access_token
    payload = decode_token(refresh_token_str)
    if payload is None or payload.get('type') != 'refresh':
        raise HTTPException(401, "登录已过期，请重新登录")

    user_id = int(payload['sub'])
    user = await get_user_by_id(db, user_id)
    if not user:
        raise HTTPException(401, "用户不存在")

    new_access = create_access_token(user.id)
    return {"token": new_access}


@app.put("/api/auth/change-password")

async def change_password(
    old_password: str = Form(...),
    new_password: str = Form(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_session),
):
    if len(new_password) < 6:
        raise HTTPException(400, "新密码至少6位")

    if not verify_password(old_password, user.password_hash):
        raise HTTPException(400, "旧密码错误")

    user.password_hash = hash_password(new_password)
    await db.commit()
    return {"status": "ok"}


#  Products API (数据库路径与 product_manage/db.py 统一) 

_DEFAULT_DB = Path.home() / ".product_tool" / "products.db"

PRODUCTS_DB_PATH = Path(os.environ.get("PRODUCT_TOOL_DB_PATH", str(_DEFAULT_DB)))


def _init_products_db():
    """初始products 表结构(启动时调用一次)"""

    import sqlite3

    PRODUCTS_DB_PATH.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(str(PRODUCTS_DB_PATH))

    conn.row_factory = sqlite3.Row

    conn.execute("""

        CREATE TABLE IF NOT EXISTS web_products (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            user_id TEXT NOT NULL,

            model TEXT,

            name_zh TEXT,

            spec_zh TEXT,

            price_rmb REAL,

            image_path TEXT,

            category TEXT,

            currency TEXT DEFAULT 'RMB',

            carton_size TEXT DEFAULT '',

            gross_weight REAL DEFAULT 0,

            net_weight REAL DEFAULT 0,

            cbm REAL DEFAULT 0,

            units_per_carton INTEGER DEFAULT 0,

            packing_type TEXT DEFAULT '',

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP

        )

    """)

    conn.execute("""

        CREATE TABLE IF NOT EXISTS web_quotations (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            user_id TEXT NOT NULL,

            product_ids TEXT,

            file_name TEXT,

            file_path TEXT DEFAULT '',

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP

        )

    """)

    conn.execute("CREATE INDEX IF NOT EXISTS idx_wp_user ON web_products(user_id)")

    conn.execute("CREATE INDEX IF NOT EXISTS idx_wp_time ON web_products(created_at)")

    conn.execute("CREATE INDEX IF NOT EXISTS idx_wq_user ON web_quotations(user_id)")

    # Migrations (safe to run repeatedly)

    for migration in [

        "ALTER TABLE web_quotations ADD COLUMN file_path TEXT DEFAULT ''",

        "ALTER TABLE web_products ADD COLUMN currency TEXT DEFAULT 'RMB'",

        "ALTER TABLE web_products ADD COLUMN carton_size TEXT DEFAULT ''",

        "ALTER TABLE web_products ADD COLUMN gross_weight REAL DEFAULT 0",

        "ALTER TABLE web_products ADD COLUMN net_weight REAL DEFAULT 0",

        "ALTER TABLE web_products ADD COLUMN cbm REAL DEFAULT 0",

        "ALTER TABLE web_products ADD COLUMN units_per_carton INTEGER DEFAULT 0",

        "ALTER TABLE web_products ADD COLUMN packing_type TEXT DEFAULT ''",
        "ALTER TABLE web_products ADD COLUMN price_cny REAL DEFAULT 0",

    ]:

        try:

            conn.execute(migration)

        except Exception:

            pass

    conn.commit()

    conn.close()


_init_products_db()


def _get_products_db():
    """Return SQLite connection for product/quote data.
    TODO: Migrate to PostgreSQL ORM (separate task)."""
    conn = sqlite3.connect(str(PRODUCTS_DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _extract_packaging_from_spec(spec_zh: str) -> dict:

    """spec_zh 中提取结构化包装信息"""

    result = {'carton_size': '', 'gross_weight': 0, 'net_weight': 0, 'cbm': 0, 'units_per_carton': 0, 'packing_type': ''}

    if not spec_zh:

        return result

    

    # 按常见分隔符拆分为键值对

    parts = re.split(r'[;\n]', spec_zh)

    for part in parts:

        part = part.strip()

        if ':' not in part:

            continue

        key, val = part.split(':', 1)

        key = key.strip().lower()

        val = val.strip()

        

        # 外箱尺寸 / Carton Size / Packing Size / Dimensions

        if any(k in key for k in ['carton size', '外箱尺寸', 'packing size', 'package size', '包装尺寸', 'carton', '测量']):

            result['carton_size'] = part.split(':', 1)[1].strip()

        

        # 毛重 / Gross Weight / GW

        if any(k in key for k in ['gross weight', 'g.w.', '毛重', 'gross']):

            nums = re.findall(r'[\d.]+', val.replace('kg', '').replace('KG', ''))

            if nums:

                try: result['gross_weight'] = float(nums[0])

                except Exception: pass

        

        # 净/ Net Weight / NW

        if any(k in key for k in ['net weight', 'n.w.', '净', 'net']):

            nums = re.findall(r'[\d.]+', val.replace('kg', '').replace('KG', ''))

            if nums:

                try: result['net_weight'] = float(nums[0])

                except Exception: pass

        

        # CBM / 体积

        if any(k in key for k in ['cbm', '体积', 'meas', 'measurement']):

            nums = re.findall(r'[\d.]+', val.replace('cbm', '').replace('CBM', '').replace('m', ''))

            if nums:

                try: result['cbm'] = float(nums[0])

                except Exception: pass

        

        # 每箱数量 / QTY/CTN

        if any(k in key for k in ['qty/ctn', 'pcs/ctn', '每箱数量', 'units per carton', 'qty per carton']):

            nums = re.findall(r'[\d]+', val)

            if nums:

                try: result['units_per_carton'] = int(nums[0])

                except Exception: pass

        

        # 包装类型

        if any(k in key for k in ['包装', 'packing type', 'package type', 'packing']):

            val_clean = val.strip().lower()

            if val_clean not in ('gw', 'nw', 'gross weight', 'net weight', 'cbm', ''):

                result['packing_type'] = val

    

    return result


@app.post("/api/products/save")

async def save_products(products: str = Form(...), user: dict = Depends(get_current_user)):

    items = json.loads(products)

    if not items or not isinstance(items, list):

        raise HTTPException(400, "产品列表不能为空")

    # 检查每个产品至少要model

    for item in items:

        if not item.get('model', '').strip():

            raise HTTPException(400, "产品型号不能为空")

    from product_repo import save_products, count_products
    uid = user.id

    # 免费版检查产品数量上限
    if user.tier != "pro":
        current_count = count_products(uid)
        if current_count + len(items) > 200:
            raise HTTPException(403, "免费版最多保存200个产品,升级专业版可解除限制")

    inserted = save_products(uid, items)
    return {"status": "ok", "inserted": inserted}


@app.get("/api/products")

async def get_products(user: dict = Depends(get_current_user)):

    from product_repo import get_products
    return get_products(user.id)


@app.delete("/api/products/{product_id}")

async def delete_product(product_id: int, user: dict = Depends(get_current_user)):

    from product_repo import delete_product as repo_delete
    repo_delete(product_id, user.id)
    return {"status": "deleted"}


@app.post("/api/products/batch-delete")

async def batch_delete_products(product_ids: str = Form(...), user: dict = Depends(get_current_user)):

    ids = json.loads(product_ids)
    from product_repo import batch_delete_products as repo_batch_delete
    count = repo_batch_delete([int(x) for x in ids], user.id)
    return {"status": "deleted", "count": count}


#  Quotations API 

@app.get("/api/quotations")

async def get_quotations(user: dict = Depends(get_current_user)):

    from product_repo import get_quotations
    rows = get_quotations(user.id)
    result = []

    for r in rows:

        try:

            pids = json.loads(r.get("product_ids", "[]"))

            if isinstance(pids, list):

                r["model_count"] = len(pids)

        except Exception:

            r["model_count"] = 0

        fn = (r.get("file_name") or "").lower()

        if "形式发票" in fn or ("pi" in fn and "pdf" not in fn):

            r["title"] = f"PI #{r['id']}"

        elif "装箱" in fn or "packing" in fn:

            r["title"] = f"装箱#{r['id']}"

        elif "发票" in fn or "invoice" in fn:

            r["title"] = f"商业发票 #{r['id']}"

        elif "pdf" in fn:

            r["title"] = f"PDF报价#{r['id']}"

        else:

            r["title"] = f"报价#{r['id']}"

        result.append(r)

    return {"quotations": result}


#  Serve product images 

@app.get("/api/images")

@app.get("/api/images/")

async def serve_image(path: str = Query(...)):

    # Security: resolve symlinks/.. and check against allowlist

    allowed_dirs = [

        Path(PROJECT_ROOT / "product_tool" / "temp_images").resolve(),

        Path(PROJECT_ROOT / "product_tool" / "data").resolve(),

        Path(UPLOAD_DIR / "images").resolve(),

    ]

    try:

        abs_path = Path(path).resolve()

    except Exception:

        raise HTTPException(400, "无效的文件路径")

    if not any(abs_path == d or str(abs_path).startswith(str(d) + os.sep) for d in allowed_dirs):

        raise HTTPException(403, "无权访问此文件")

    if not abs_path.is_file():

        raise HTTPException(404, "图片未找到")

    ext = abs_path.suffix.lower()

    media_types = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png', '.gif': 'image/gif'}

    media_type = media_types.get(ext, 'application/octet-stream')

    # Cache: images static, 24h cache
    import email.utils
    _st = abs_path.stat()
    _headers = {
        'Cache-Control': 'public, max-age=86400',
        'ETag': f'"{int(_st.st_mtime)}-{_st.st_size}"',
        'Last-Modified': email.utils.formatdate(_st.st_mtime, usegmt=True),
    }
    return FileResponse(str(abs_path), media_type=media_type, headers=_headers)


#  Template: upload file to extract company info 

@app.post("/api/template/upload")

async def upload_template(file: UploadFile = File(...)):

    ext = Path(file.filename).suffix.lower()

    if ext not in ('.xlsx', '.xls'):

        raise HTTPException(400, "仅支.xlsx 模板文件")

    

    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    save_path = UPLOAD_DIR / f"template_{ts}_{file.filename}"

    content = await limit_upload_size(file)

    with open(save_path, "wb") as f:

        f.write(content)

    

    try:

        from openpyxl import load_workbook

        wb = load_workbook(save_path, data_only=True)

        ws = wb.active

        

        company_info = {"company_name": "", "address": "", "address_en": "", "contact": "", "phone": "", "phone_en": ""}

        headers = []

        # Scan first 10 rows for company info
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):

            for cell in row:

                val = str(cell.value or '').strip()

                if not val:

                    continue

                val_lower = val.lower()

                if any(k in val_lower for k in ['公司', 'company', '企业']):

                    company_info['company_name'] = val

                elif any(k in val_lower for k in ['地址', 'address', 'add']):

                    company_info['address'] = val

                elif any(k in val_lower for k in ['联系', 'contact', '联络']):

                    company_info['contact'] = val

                elif any(k in val_lower for k in ['电话', 'phone', 'tel', '手机']):

                    company_info['phone'] = val
        # Detect header row (look for common product column names)





        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):

            for cell in row:

                if cell and str(cell).strip():

                    cval = str(cell).strip().lower()

                    if any(k in cval for k in ['型号', 'model', '产品', 'product', '名称', 'name', '规格', 'spec', '价格', 'price', '数量', 'qty', '数量']):

                        headers = [str(c.value or '').strip() for c in ws[1] if c.value]

                        break

            if headers:

                break

        

        wb.close()

        os.remove(save_path)

        

        return {"company": company_info, "columns": headers, "detected": bool(company_info.get('company_name'))}

    except Exception as e:

        if os.path.exists(save_path):

            os.remove(save_path)

        raise HTTPException(500, f"模板解析失败: {str(e)}")


#  Template: save config(写入统一 company.json

@app.post("/api/template/save")

async def save_template(config: str = Form(...), user: dict = Depends(get_current_user)):
    global _COMPANY_CACHE

    data = json.loads(config)

    try:

        from src.company import load_company, save_company

        _COMPANY_CACHE = None

        current = load_company()

        # Only update UI-visible fields (not bank info)
        for k in ['name', 'name_en', 'address', 'address_en', 'city', 'tel', 'email', 'website', 'contact_person', 'logo_path']:

            if k in data:

                current[k] = data[k]

        save_company(current)

    except Exception as e:

        raise HTTPException(500, f"保存失败: {str(e)}")

    return {"status": "ok"}


#  Template: get saved config(只返回 UI 可见字段

@app.get("/api/template")

async def get_template(user: dict = Depends(get_current_user)):

    cfg = _load_company_config()

    return {
        "name": cfg.get("name", ""),
        "name_en": cfg.get("name_en", ""),
        "address": cfg.get("address", ""),
        "address_en": cfg.get("address_en", ""),
        "city": cfg.get("city", ""),
        "tel": cfg.get("tel", ""),
        "email": cfg.get("email", ""),
        "website": cfg.get("website", ""),
        "contact_person": cfg.get("contact_person", ""),
        "logo_path": cfg.get("logo_path", ""),
    }



#  Logo 上传 

LOGO_DIR = Path.home() / ".product_tool"

LOGO_DIR.mkdir(exist_ok=True)


@app.post("/api/company/logo")

async def upload_logo(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    global _COMPANY_CACHE

    ext = Path(file.filename).suffix.lower()

    if ext not in ('.jpg', '.jpeg', '.png', '.gif'):

        raise HTTPException(400, "仅支jpg/png/gif 格式")

    logo_path = LOGO_DIR / f"logo{ext}"

    content = await limit_upload_size(file)

    with open(logo_path, "wb") as f:

        f.write(content)

    # 更新配置中的 logo_path

    try:

        from src.company import load_company, save_company

        cfg = load_company()

        cfg["logo_path"] = str(logo_path)

        save_company(cfg)
        _COMPANY_CACHE = None

    except Exception as e:

        logger.warning("Logo保存失败: %s", e)

    return {"path": str(logo_path)}


# Bank info storage (replaces localStorage)
BANK_INFO_FILE = os.path.join(os.path.dirname(__file__), "data", "bank_info.json")

@app.post("/api/bank/save")
async def save_bank_info(
    beneficiary: str = Form(""),
    bank_name: str = Form(""),
    bank_address: str = Form(""),
    account_no: str = Form(""),
    swift_code: str = Form(""),
    user: User = Depends(get_current_user)):
    os.makedirs(os.path.dirname(BANK_INFO_FILE), exist_ok=True)
    data = {
        "beneficiary": beneficiary,
        "bank_name": bank_name,
        "bank_address": bank_address,
        "account_no": account_no,
        "swift_code": swift_code,
    }
    with open(BANK_INFO_FILE, "w") as f:
        json.dump(data, f)
    return {"status": "ok"}

@app.get("/api/bank/load")
async def load_bank_info(user: User = Depends(get_current_user)):
    try:
        with open(BANK_INFO_FILE, "r") as f:
            return json.load(f)
    except Exception:
        return {}


#  Parse file (2-step pipeline) 

@app.post("/api/parse")

async def parse_file(

    file: UploadFile = File(...),

    user: User = Depends(get_current_user_optional),
    db: AsyncSession = Depends(get_session)):

    ext = Path(file.filename).suffix.lower()

    if ext not in ('.xlsx', '.xls', '.pdf', '.docx'):

        raise HTTPException(400, f"不支持的文件格式:{ext},仅支持 Excel/PDF/Word")


    # Free user upload limit check (only for logged-in users)
    if user:
        from auth import check_upload_limit
        if not await check_upload_limit(user, db):
            raise HTTPException(403, "Free users limited to 20 uploads/month. Upgrade Pro to remove limit.")


    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    save_path = UPLOAD_DIR / f"{ts}_{file.filename}"

    content = await limit_upload_size(file)

    with open(save_path, "wb") as f:

        f.write(content)


    try:

        df = None

        parse_source = ''

        cache_key = ''


        #  Step 1 & 2: Universal + Specialized parsers

        df = None
        df2 = None
        parse_source = ''

        if ext in ('.xlsx', '.xls'):

            # 并行执行通用解析器和专用解析器
            from run import parse_file as run_parse_file
            loop = asyncio.get_event_loop()

            async def _run_both():
                uni_fut = loop.run_in_executor(None, universal_parse, str(save_path))
                spec_fut = loop.run_in_executor(None, run_parse_file, str(save_path))
                uni_result, spec_result = await asyncio.gather(uni_fut, spec_fut)
                return uni_result, spec_result

            uni_result, spec_result = await _run_both()
            df, ptype, count, cache_key = uni_result
            df2 = spec_result
            if count > 0:
                parse_source = f'universal_{ptype}'

        elif ext == '.pdf':

            # Check for scanned PDF (image-based, no extractable text)
            try:
                from src.core.pdf_parser import is_likely_scanned_pdf
                if is_likely_scanned_pdf(str(save_path)):
                    raise HTTPException(400, "该PDF文件为扫描件，暂不支持。请上传文字型PDF文件（非图片/扫描件）")
            except HTTPException:
                raise
            except Exception:
                pass  # If detection fails, proceed normally

            from pdf_handler import extract_products_from_pdf_v2

            df2 = extract_products_from_pdf_v2(str(save_path))

        elif ext == '.docx':

            from src.core.doc_parser import extract_products_from_docx

            df2 = extract_products_from_docx(str(save_path))


        #  Step 3: 择优 — 评分系统，质量优先于数量
        if df2 is not None and len(df2) > 0:

            if df is None or len(df) == 0:

                df = df2

                parse_source = 'specialized'

            else:
                # 用评分系统比较两个解析器的输出质量
                uni_score = score_dataframe(df)
                spec_score = score_dataframe(df2)

                # 比较"优质产品数"而非总分，避免"2个完美品 > 25个合格品"
                from score import score_product_row as _spr
                def _count_good(df_):
                    if df_ is None or len(df_) == 0:
                        return 0
                    return sum(1 for _, r_ in df_.iterrows()
                               if _spr(str(r_.get('model','')), r_.get('price_rmb'), str(r_.get('spec_zh',''))) > 0)

                uni_good = _count_good(df)
                spec_good = _count_good(df2)

                if spec_good > uni_good or (spec_good == uni_good and spec_score['score'] > uni_score['score']):
                    df = df2
                    parse_source = 'specialized'

        if df is None or len(df) == 0:

            raise HTTPException(400, "文件中未找到产品数据")

        # 统一图片匹配：选赢家后只做一次（放入线程池避免阻塞事件循环）
        if ext in ('.xlsx', '.xls') and '_row' in df.columns:
            try:
                from src.core.image import match_images_to_products as _match_img
                df = await loop.run_in_executor(None, _match_img, df, str(save_path))
            except Exception:
                pass
        elif ext == '.docx' and '_row' in df.columns:
            try:
                from src.core.image import match_images_to_products_docx as _match_docx
                df = await loop.run_in_executor(None, _match_docx, df, str(save_path))
            except Exception:
                pass
        # PDF 图片匹配已在 pdf_parser.py 内部完成


        # 后处理:PDF model 空值向前填充(共享上一行型号)

        if ext == '.pdf' and 'model' in df.columns:

            df['model'] = df['model'].fillna('').astype(str).str.strip()

            df['model'] = df.groupby((df['model'] != '').cumsum())['model'].transform('first')

        # 币种兜底：如果所有产品都没有币种，检查数据中是否含USD/FOB信号
        if 'currency' not in df.columns or df['currency'].isna().all() or (df['currency'] == '').all():
            # 检查有价格的产品，如果价格 > 500 且 spec 不含明显人民币信号，则设为 USD
            has_price_col = 'price_rmb' in df.columns
            if has_price_col:
                prices = [p for p in df['price_rmb'] if isinstance(p, (int, float)) and p > 0]
                if prices:
                    avg_price = sum(prices) / len(prices)
                    if avg_price > 500:
                        df['currency'] = 'USD'
                    else:
                        df['currency'] = 'CNY'

        products = df.fillna('').to_dict(orient='records')

        

        # 上传成功计数（仅登录用户）
        if user:
            from auth import increment_upload
            await increment_upload(user, db)

        

        return {
            "products": products,
            "count": len(products),
            "parse_source": parse_source,
            "cache_key": cache_key,
        }

    except HTTPException:

        raise

    except Exception as e:

        raise HTTPException(500, f"Parse failed: {str(e)}")

    finally:

        if save_path.exists():

            os.remove(save_path)


#  AI-enhanced parse 

@app.post("/api/parse/with-ai")

async def parse_with_ai(file: UploadFile = File(...), ai_backend: str = Form("gemini"), user: dict = Depends(get_current_user_optional)):

    ext = Path(file.filename).suffix.lower()

    if ext not in ('.xlsx', '.xls', '.pdf', '.docx'):

        raise HTTPException(400, f"不支持的文件格式:{ext},仅支持 Excel/PDF/Word")

    if ext in ('.pdf', '.docx'):

        # AI 模型不支PDF/DOCX 直接解析,引导用户使用传统解析器

        raise HTTPException(400, "AI 解析暂不支持 PDF/DOCX 格式,请使用传统解析 /api/parse")


    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    save_path = UPLOAD_DIR / f"{ts}_{file.filename}"

    content = await limit_upload_size(file)

    with open(save_path, "wb") as f:

        f.write(content)


    try:

        wb = load_workbook(str(save_path), data_only=True)

        ws = wb.active

        if ws is None:

            raise HTTPException(400, "文件为空,请检查文件")


        header_row = detect_header_row(ws)

        headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, min(ws.max_column + 1, 20))]

        cache_key = ai_cache_key(headers)


        # Check cache

        cache = load_cache()

        if cache_key in cache:

            col_map = cache[cache_key]

            header_row_2 = detect_header_row(ws)

            df = parse_with_colmap(ws, header_row_2, col_map)

            # 暂不 close — 低覆盖率需回退到 AI 检测，仍需要 ws
            if compute_coverage(df) >= 0.5:

                wb.close()

                products = df.fillna('').to_dict(orient='records')

                return {"products": products, "count": len(products), "parse_source": "cached_ai"}

            df = pd.DataFrame()


        # AI column detection

        md_table = sheet_to_markdown(ws, max_rows=15)

        wb.close()

        col_map = ai_detect_columns(md_table, backend=ai_backend)

        if not col_map or not any(v is not None for v in col_map.values()):

            raise HTTPException(400, "AI 无法识别此文件格")


        # Cache

        cache[cache_key] = col_map

        save_cache(cache)


        # Parse with AI column map
        header_row_3 = detect_header_row(ws)
        df = parse_with_colmap(ws, header_row_3, col_map)

        if df is None or len(df) == 0:

            raise HTTPException(400, "AI 解析后未找到产品数据")


        products = df.fillna('').to_dict(orient='records')

        return {"products": products, "count": len(products), "parse_source": "ai"}

    except HTTPException:

        raise

    except Exception as e:

        raise HTTPException(500, f"AI 解析失败: {str(e)}")

    finally:

        if save_path.exists():

            os.remove(save_path)


@app.post("/api/parse-text-products")

async def parse_text_products(data: dict = Body(...), user: User = Depends(get_current_user_optional)):
    """从自由文本中提取结构化产品（智能粘贴功能 — 仅 Pro）"""
    if not user:
        raise HTTPException(401, "请先登录")
    require_pro(user)
    text = data.get('text', '')
    if not text or not text.strip():
        return {"products": [], "count": 0}
    try:
        products = parse_text_to_products(text, backend='deepseek')
        return {"products": products, "count": len(products)}
    except Exception as e:
        raise HTTPException(422, f"文本解析失败: {str(e)}")


#  Document Templates 

TEMPLATE_DIR = Path.home() / ".product_tool" / "templates"

TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)


@app.post("/api/template/document/{doc_type}")

async def upload_doc_template(doc_type: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):

    if doc_type not in ('quotation', 'pi', 'packing', 'invoice'):

        raise HTTPException(400, "无效的文档类")

    ext = Path(file.filename).suffix.lower()

    if ext != '.xlsx':

        raise HTTPException(400, "仅支.xlsx 模板文件")

    save_path = TEMPLATE_DIR / f"{doc_type}.xlsx"

    content = await limit_upload_size(file)

    with open(save_path, "wb") as f:

        f.write(content)

    return {"status": "ok", "path": str(save_path)}


@app.get("/api/template/document/{doc_type}")

async def get_doc_template(doc_type: str):

    if doc_type not in ('quotation', 'pi', 'packing', 'invoice'):

        raise HTTPException(400, "无效的文档类")

    save_path = TEMPLATE_DIR / f"{doc_type}.xlsx"

    return {"exists": save_path.exists(), "size": save_path.stat().st_size if save_path.exists() else 0}


@app.delete("/api/template/document/{doc_type}")

async def delete_doc_template(doc_type: str, user: dict = Depends(get_current_user)):

    if doc_type not in ('quotation', 'pi', 'packing', 'invoice'):

        raise HTTPException(400, "无效的文档类")

    save_path = TEMPLATE_DIR / f"{doc_type}.xlsx"

    if save_path.exists():

        save_path.unlink()

    return {"status": "deleted"}


#  Generate quotation 

@app.post("/api/quotation")

async def generate_quotation(

    products: str = Form(...),

    lang: str = Form("bilingual"),

    trade_terms: str = Form("EXW"),

    company_name: str = Form(""),

    company_contact: str = Form(""),

    company_phone: str = Form(""),

    payment_terms: str = Form(DEFAULT_PAYMENT_TERMS),

    currency: str = Form("CNY"),

    with_images: str = Form("1"),

    contract_no: str = Form(""),
    po_no: str = Form(""),
    lc_no: str = Form(""),
    hs_code: str = Form(""),
    shipping_marks: str = Form(""),
    freight: str = Form(""),
    insurance: str = Form(""),
    handling: str = Form(""),
    delivery_time: str = Form(""),
    validity_days: str = Form(""),

    db: AsyncSession = Depends(get_session),

    authorization: str = Header(None)):

    
    import pandas as pd


    items = json.loads(products) 

    

    # Build company info dict (user input優先, fallback to config)
    cfg_company = _load_company_config()

    company_info = {} 

    company_info['name'] = company_name or cfg_company.get('name_en', '') or cfg_company.get('name', '') 

    company_info['address'] = cfg_company.get('address_en', '') or cfg_company.get('address', '') 

    company_info['contact'] = company_contact or cfg_company.get('contact_person', '') 

    company_info['tel'] = company_phone or cfg_company.get('tel', '') 

    company_info['email'] = cfg_company.get('email', '') 


        # Image fallback: build SKU→path index once (cached 60s), then O(1) lookup per item
    DATA_IMG_DIR = PROJECT_ROOT / "product_tool" / "data" 

    TEMP_IMG_DIR = PROJECT_ROOT / "product_tool" / "temp_images" 

    global _SKU_INDEX_CACHE, _SKU_INDEX_CACHE_TS
    import time as _time
    now_ts = _time.time()
    if not _SKU_INDEX_CACHE or now_ts - _SKU_INDEX_CACHE_TS > 60:
        import glob as _glob
        _SKU_INDEX_CACHE = {}
        _SKU_INDEX_CACHE_TS = now_ts
        for _dir in [str(TEMP_IMG_DIR), str(DATA_IMG_DIR)]:
            if os.path.isdir(_dir):
                for _ext in ['*.jpg', '*.jpeg', '*.png', '*.bmp', '*.gif']:
                    for _f in _glob.glob(os.path.join(_dir, '**', _ext), recursive=True):
                        _name = os.path.splitext(os.path.basename(_f).lower())[0]
                        if _name not in _SKU_INDEX_CACHE:
                            _SKU_INDEX_CACHE[_name] = _f
    _sku_index = _SKU_INDEX_CACHE

    for item in items: 

        if not item.get('_image_path') and not item.get('image_path'): 

            sku = item.get('model', '') or item.get('sku', '') 

            if sku: 
                sku_lower = sku.lower().strip()
                img = _sku_index.get(sku_lower) or next((v for k, v in _sku_index.items() if sku_lower in k or k in sku_lower), None)

                if img: 

                    item['_image_path'] = img 


    ts = datetime.now().strftime("%Y%m%d%H%M%S") 

    output_path = OUTPUT_DIR / f"quotation_{ts}.xlsx" 

    # 图片预热：提前 resize 所有图片，避免生成阶段逐张 PIL 解码大图
    _resize_temp_files = []
    if with_images == "1":
        try:
            from src.core.image import resize_image as _pre_resize
            for item in items:
                img_p = item.get('_image_path') or item.get('image_path', '')
                if img_p and os.path.exists(img_p):
                    resized = _pre_resize(img_p)
                    if hasattr(resized, 'read'):
                        import tempfile as _tf
                        _tmp = _tf.NamedTemporaryFile(suffix='.jpg', delete=False)
                        _tmp.write(resized.read())
                        _tmp.close()
                        item['_image_path'] = _tmp.name
                        _resize_temp_files.append(_tmp.name)
        except Exception:
            pass

    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            None,
            functools.partial(
                create_quotation,
                items, str(output_path),
                lang=lang,
                with_images=(with_images == "1"),
                company_info=company_info if company_info else None,
                payment_terms=payment_terms,
                currency=currency,
                contract_no=contract_no,
                po_no=po_no,
                lc_no=lc_no,
                hs_code=hs_code,
                shipping_marks=shipping_marks,
                freight=freight,
                insurance=insurance,
                handling=handling,
                delivery_time=delivery_time,
                validity_days=validity_days,
            ),
        )

    except Exception as e:

        logger.error("报价单生成失败: %s", e, exc_info=True)
        raise HTTPException(500, f"报价单生成失败: {e}")
    finally:
        # Clean up pre-resized temp files
        for _tf in _resize_temp_files:
            try: os.unlink(_tf)
            except OSError: pass

    # Auto-save to quotation history if user is logged in
    quotation_id = None
    user = await get_current_user_optional(authorization, db)
    if user:
        try: 
            from product_repo import save_quotation
            uid = user.id
            fname = f"报价单_{ts}.xlsx"
            quotation_id = save_quotation(uid, json.dumps(items), fname, str(output_path))

        except Exception as e: 
            logger.warning("报价auto-save 失败: %s", e)


    if quotation_id:
        return {"status": "ok", "id": quotation_id, "name": f"报价单_{ts}.xlsx"}
    return FileResponse(
        str(output_path), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"报价单_{ts}.xlsx",
    )



@app.get("/api/quotations/{id}/download")

async def download_quotation(id: int, user: dict = Depends(get_current_user)):

    from product_repo import get_quotation
    row = get_quotation(id, user.id)

    if not row:

        raise HTTPException(404, "报价记录不存")

    file_path = row["file_path"]

    if not file_path or not os.path.isfile(file_path):

        raise HTTPException(404, "文件已丢失或已被删除")

    return FileResponse(file_path, filename=row["file_name"])


# Delete quotation from history 

@app.delete("/api/quotations/{id}")

async def delete_quotation(id: int, user: dict = Depends(get_current_user)):

    from product_repo import delete_quotation as repo_delete_q
    file_path = repo_delete_q(id, user.id)

    if file_path and os.path.isfile(file_path):
        try:
            os.remove(file_path)
        except Exception:
            pass

    return {"status": "deleted"}


@app.post("/api/quotations/batch-delete")

async def batch_delete_quotations(ids: str = Form(...), user: dict = Depends(get_current_user)):

    id_list = json.loads(ids)

    from product_repo import batch_delete_quotations as repo_batch_q
    count = repo_batch_q([int(x) for x in id_list], user.id)
    return {"status": "deleted", "count": count}


#  Generate PI (Proforma Invoice) 

@app.post("/api/quotation/pdf")

async def generate_quotation_pdf(

    products: str = Form(...),

    lang: str = Form("bilingual"),

    company_name: str = Form(""),

    company_contact: str = Form(""),

    company_phone: str = Form(""),

    trade_terms: str = Form("EXW"),

    payment_terms: str = Form(DEFAULT_PAYMENT_TERMS),

    currency: str = Form("CNY"),

    with_images: str = Form("1"),

    db: AsyncSession = Depends(get_session),

    authorization: str = Header(None)):

    items = json.loads(products)

    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    output_path = OUTPUT_DIR / f"quotation_pdf_{ts}.pdf"

    cfg = _load_company_config()

    c_name = company_name or cfg.get('name_en', '') or cfg.get('name', '')

    c_contact = company_contact or cfg.get('contact_person', '')

    success = create_quote_pdf(items, str(output_path), company=c_name, contact=c_contact, lang=lang, payment_terms=payment_terms, currency=currency, with_images=(with_images == "1"))

    if not success:

        raise HTTPException(500, "PDF 报价单生成失败")

    pdf_qid = None
    user = await get_current_user_optional(authorization, db)
    if user:
        try:
            from product_repo import save_quotation
            pdf_qid = save_quotation(user.id, json.dumps(items),
                f"报价单PDF_{ts}.pdf", str(output_path))
        except Exception as e:
            logger.warning("PDF报价auto-save 失败: %s", e)

    if pdf_qid:
        return {"status": "ok", "id": pdf_qid, "name": f"报价单PDF_{ts}.pdf"}
    return FileResponse(str(output_path), media_type="application/pdf", filename=f"报价单_{ts}.pdf")


@app.post("/api/pi")
async def generate_pi(
    products: str = Form(...),
    lang: str = Form("chinese"),
    with_images: str = Form("1"),
    trade_terms: str = Form("exw"),
    currency: str = Form("CNY"),
    company_info: str = Form("{}"),
    buyer_info: str = Form(""),
    payment_terms: str = Form(""),
    shipping_marks: str = Form(""),
    bank_info: str = Form(""),
    db: AsyncSession = Depends(get_session),

    user: User = Depends(get_current_user)):
    # Pro 功能校验
    _pro_user = await get_current_user_optional(authorization, db)
    if not _pro_user:
        raise HTTPException(401, "请先登录")
    require_pro(_pro_user)

    items = json.loads(products)

    if not items:

        raise HTTPException(400, "产品列表不能为空")

    # 从统一公司信息配置加载卖家信息

    seller_config = _load_company_config()

    if bank_beneficiary:

        seller_config.setdefault('bank', {})

        if bank_beneficiary: seller_config['bank']['beneficiary'] = bank_beneficiary

        if bank_name: seller_config['bank']['bank_name'] = bank_name

        if bank_address: seller_config['bank']['bank_address'] = bank_address

        if bank_account: seller_config['bank']['account_no'] = bank_account

        if bank_swift: seller_config['bank']['swift_code'] = bank_swift

    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    output_path = OUTPUT_DIR / f"PI_{ts}.xlsx"

    result_path = generate_pi_xlsx(

        items=items,

        output_path=str(output_path),

        buyer_name=buyer_name,

        buyer_address=buyer_address,

        seller_config=seller_config,

        trade_terms=trade_terms,

        payment_terms=payment_terms,

        port_destination=port_destination,

        brand_name=brand_name,

        currency=currency,

        lang=lang,

        with_images=(with_images == "1"),

    )

    if not result_path or not os.path.isfile(result_path):

        raise HTTPException(500, "形式发票生成失败")

    pi_qid = None
    try:

        if user:

            from product_repo import save_quotation
            pi_qid = save_quotation(user.id, json.dumps(items),
                f"形式发票_{ts}.xlsx", str(result_path))

    except Exception as e:

        logger.warning("PI auto-save 失败: %s", e)

    if pi_qid:
        return {"status": "ok", "id": pi_qid, "name": f"形式发票_{ts}.xlsx"}
    return FileResponse(str(result_path), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"形式发票_{ts}.xlsx") 


#  Generate Packing List 

@app.post("/api/packing")

async def generate_packing(

    products: str = Form(...),

    lang: str = Form("chinese"),

    buyer_name: str = Form(""),

    buyer_address: str = Form(""),

    trade_terms: str = Form("FOB"),

    port_loading: str = Form("Qingdao"),

    port_discharge: str = Form(""),

    vessel: str = Form(""),

    bl_no: str = Form(""),

    origin_country: str = Form("China"),

    packing_type: str = Form("Carton"),

    packing_qty: str = Form(""),

    db: AsyncSession = Depends(get_session),

    authorization: str = Header(None),

):



    # Pro 功能校验
    _pro_user = await get_current_user_optional(authorization, db)
    if not _pro_user:
        raise HTTPException(401, "请先登录")
    require_pro(_pro_user)

    items = json.loads(products)

    if not items:

        raise HTTPException(400, "产品列表不能为空")

    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    inv_no = f"INV-{ts}"

    output_path = OUTPUT_DIR / f"packing_{ts}.xlsx"


    company_config = _load_company_config()

    company_config['city'] = company_config.get('city', '')

    result_path = generate_packing_list(items, inv_no, ts[:8], output_path=str(output_path), company_config=company_config,

        lang=lang, buyer_name=buyer_name, buyer_address=buyer_address,

        port_loading=port_loading, port_discharge=port_discharge, vessel=vessel, bl_no=bl_no,

        packing_type=packing_type, packing_qty=packing_qty)

    if not result_path or not os.path.isfile(result_path):

        raise HTTPException(500, "装箱单生成失败")

    pk_qid = None
    try:

        if _pro_user:

            from product_repo import save_quotation
            pk_qid = save_quotation(_pro_user.id, json.dumps(items),
                f"装箱单_{ts}.xlsx", str(result_path))

    except Exception as e:

        logger.warning("装箱单auto-save 失败: %s", e)

    if pk_qid:
        return {"status": "ok", "id": pk_qid, "name": f"装箱单_{ts}.xlsx"}
    return FileResponse(result_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"装箱单_{ts}.xlsx")


#  Generate Commercial Invoice 

@app.post("/api/invoice")

async def generate_invoice(

    products: str = Form(...),

    lang: str = Form("chinese"),

    buyer_name: str = Form(""),

    buyer_address: str = Form(""),

    trade_terms: str = Form("FOB"),

    payment_terms: str = Form(DEFAULT_PAYMENT_TERMS),

    currency: str = Form("CNY"),

    port_loading: str = Form("Qingdao"),

    port_discharge: str = Form(""),

    vessel: str = Form(""),

    bl_no: str = Form(""),

    origin_country: str = Form("China"),

    bank_beneficiary: str = Form(""),

    bank_name: str = Form(""),

    bank_address: str = Form(""),

    bank_account: str = Form(""),

    bank_swift: str = Form(""),

    db: AsyncSession = Depends(get_session),

    authorization: str = Header(None),

):



    # Pro 功能校验
    _pro_user = await get_current_user_optional(authorization, db)
    if not _pro_user:
        raise HTTPException(401, "请先登录")
    require_pro(_pro_user)

    items = json.loads(products)

    if not items:

        raise HTTPException(400, "产品列表不能为空")

    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    inv_no = f"INV-{ts}"

    output_path = OUTPUT_DIR / f"invoice_{ts}.xlsx"

    company_config = _load_company_config()

    # 前端传来的银行信息覆
    if bank_beneficiary:
        company_config.setdefault('bank', {})
        if bank_beneficiary: company_config['bank']['beneficiary'] = bank_beneficiary
        if bank_name: company_config['bank']['bank_name'] = bank_name
        if bank_address: company_config['bank']['bank_address'] = bank_address
        if bank_account: company_config['bank']['account_no'] = bank_account
        if bank_swift: company_config['bank']['swift_code'] = bank_swift

    result_path = generate_commercial_invoice(items, inv_no, ts[:8], output_path=str(output_path), company_config=company_config,

        lang=lang, buyer_name=buyer_name, buyer_address=buyer_address,

        trade_terms=trade_terms, payment_terms=payment_terms, currency=currency,

        port_loading=port_loading, port_discharge=port_discharge,

        vessel=vessel, bl_no=bl_no, origin_country=origin_country) 

    if not result_path or not os.path.isfile(result_path): 

        raise HTTPException(500, "商业发票生成失败")

    inv_qid = None
    try:

        if _pro_user:

            from product_repo import save_quotation
            inv_qid = save_quotation(_pro_user.id, json.dumps(items),
                f"商业发票_{ts}.xlsx", str(result_path))

    except Exception as e:

        logger.warning("商业发票 auto-save 失败: %s", e)

    if inv_qid:
        return {"status": "ok", "id": inv_qid, "name": f"商业发票_{ts}.xlsx"}
    return FileResponse(result_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"商业发票_{ts}.xlsx")


