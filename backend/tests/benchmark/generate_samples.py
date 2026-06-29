"""
用 openpyxl 生成评测用样本 Excel 文件
"""
import openpyxl
from openpyxl.styles import Font
from pathlib import Path

SAMPLES_DIR = Path(__file__).parent / "samples"

def create_sample_01():
    """标准外贸报价单 — 3 个产品，带型号/价格/规格"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Headers
    headers = ['型号', '产品名称', '规格', '单价(RMB)', 'MOQ', '包装']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True)
    
    # Data
    data = [
        ['BT-001', '蓝牙耳机', '蓝牙5.3 / 续航8h / IPX5防水 / Type-C充电', 45.00, 100, '彩盒'],
        ['BP-200', '移动电源 20000mAh', '22.5W快充 / USB-C+双USB-A / LED电量显示', 68.00, 50, '白盒'],
        ['WK-500', '智能手表', '1.43英寸AMOLED / 心率血氧监测 / IP68防水 / 14天续航', 129.00, 30, '彩盒'],
    ]
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    
    wb.save(SAMPLES_DIR / "sample_01_standard.xlsx")
    print(f"Created: sample_01_standard.xlsx")

def create_sample_02():
    """复杂格式 — 价格含美元和人民币两列，带币种标记"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PriceList"
    
    headers = ['Item No.', 'Description', 'Specification', 'FOB Price(USD)', 'EXW Price(RMB)', 'Packing']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True)
    
    data = [
        ['LED-100', 'LED Panel Light 100W', '100W / 5000K / 120lm/W / CE RoHS', 12.50, 85.00, 'Carton/10pcs'],
        ['LED-200', 'LED Panel Light 200W', '200W / 4000K / 130lm/W / CE RoHS IP65', 22.00, 145.00, 'Carton/5pcs'],
    ]
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    
    wb.save(SAMPLES_DIR / "sample_02_dual_price.xlsx")
    print(f"Created: sample_02_dual_price.xlsx")

def create_sample_03():
    """含无效行的表格 — 前2行是标题/空行，有订单号干扰"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Row 1: title
    ws.cell(1, 1, "Shanghai Factory Price List 2026")
    # Row 2: empty
    # Row 3: headers
    headers = ['Model', 'Name', 'Specs', 'Price', 'Qty']
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h).font = Font(bold=True)
    # Row 4: an order number line (should be skipped)
    ws.cell(4, 1, "Order PO-2026001")
    ws.cell(4, 2, "Confirmation")
    # Rows 5-7: products
    data = [
        ['SP-100', 'Bluetooth Speaker', '20W / IPX7 / TWS / 12h battery', 89.00, 200],
        ['SP-200', 'Portable Speaker Mini', '5W / IPX5 / Bluetooth 5.0 / USB-C', 35.00, 500],
        ['SP-300', 'Party Speaker', '50W / Karaoke / LED lights / 8h battery', 199.00, 50],
    ]
    for r, row in enumerate(data, 5):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    
    wb.save(SAMPLES_DIR / "sample_03_noisy.xlsx")
    print(f"Created: sample_03_noisy.xlsx")

if __name__ == '__main__':
    create_sample_01()
    create_sample_02()
    create_sample_03()
