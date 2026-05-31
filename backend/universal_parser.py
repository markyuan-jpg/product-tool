# -*- coding: utf-8 -*-
"""
universal_parser.py — 通用 Excel 解析器 V2
三层策略：KV 布局 → 表格布局 → 内容驱动，评分择优。
输出含 price_cny（人民币换算价）和 currency（原始币种）。
"""
import re, os, hashlib, math, logging
from typing import Optional
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

    # 非产品行过滤模式 — 这些模式匹配的行不应被当作产品
    _NON_PRODUCT_MODEL_PATTERNS = [
        # English field labels: "CONTRACT NO.:", "SELLER:", "Email: ..."
        re.compile(r'^(contract|seller|buyer|payment|shipping|delivery|transshipment|remarks?|note|terms|conditions?|address|tel[.:\s]|fax[.:\s]|email|phone|website|bank|account|beneficiary|swift|contact|signature|date|invoice|validity|description)', re.I),
        # Totals and subtotals
        re.compile(r'^(total\s+amount|total\s+payment|grand\s+total|sub\s*total)', re.I),
        # Numbered clauses: "5. Transshipment:", "8. Bank Information:"
        re.compile(r'^\d+[.、\s]\s*(transshipment|payment|delivery|packing|insurance|bank|inspection|arbitration|force\s*majeure|shipping|terms?|conditions?)', re.I),
        # Chinese field labels: "付款方式：", "合同编号："
        re.compile(r'^(合同|卖方|买方|付款|交货|运输|包装|条款|备注|说明|地址|电话|邮箱|日期|受益|银行|账户|签名|签字|合计|总计|金额|小计|编号|序号)', re.I),
        # Bare price/currency lines: "$ 1,000", "USD 500"
        re.compile(r'^(\u00a5|\$|eur|usd|cny)\s*[\d,]+', re.I),
        # Company info lines: "公司名称", "供应商", "客户"
        re.compile(r'^(company|supplier|customer|buyer|seller)(\s|$)', re.I),
        # Chinese numbered clauses: "1. 本合同签订...", "4. 摩托车品牌..."
        re.compile(r'^\d+[.、]\s*[（(]?\s*(本合|支付|交[货付]|运[输送]|包[装]|条[款]|备[注]|说[明]|地[址]|电[话]|日[期]|银[行]|账[户]|签[名字]|仲裁|保险)', re.I),
        # Document titles and address lines
        re.compile(r'^(proforma\s+invoice|sales\s+contract|to\s*:|the\s+(seller|buyer)|sign(ed|ature)|date\s+of\s+)', re.I),
        # Total payment text lines
        re.compile(r'^total\s+payment', re.I),
        # Port/brand/shipping clause lines
        re.compile(r'^\d+[.、]\s*(port\s+of|brand\s+name|handling\s+method|other\s+notices|motorcycle\s+brand)', re.I),
        # Signature and stamp lines
        re.compile(r'^\(signed|\(stamp|\(seal|signature\s+by|authorized\s+sign', re.I),
        # Auto-generated model placeholders
        re.compile(r'^(产品_r\d+|商品_×\d+)$'),
        # Document/order reference numbers (look like model codes but aren't products)
        re.compile(r'^(contract\s*no|order\s*no|po\s*no|invoice\s*no|ref\s*no|payment\s*no)\s*[:\-\.]?\s*[\w\-]+', re.I),
        re.compile(r'^(shipment\s*no|delivery\s*no|doc\s*no|document\s*no|quotation\s*no|quote\s*no)\s*[:\-\.]?\s*[\w\-]+', re.I),
        re.compile(r'^(PO|SO|CO|DO|WO|IV|INV|CT|CN|QT|RFQ|PI|DN|GRN)[\d\-]{8,20}$'),  # 常见文档编号前缀
        # Packing list headers
        re.compile(r'^(packing\s*(list|detail|info)|装箱(单|明细|清单))', re.I),
        # Notification / declaration lines
        re.compile(r'^(notify\s+party|consignee|carrier|forwarder|agent)', re.I),
        # Quality / test report identifiers
        re.compile(r'^(test\s+report|inspection\s+report|certificate\s+of)', re.I),
        # Unit of measure lines
        re.compile(r'^(unit\s+(of|in)|uom|计量单位|单位[：:])', re.I),
        # Country of origin
        re.compile(r'^(country\s+of\s+origin|made\s+in|原产地|manufacturer)', re.I),
    ]]


def _filter_non_product_rows(df):
    """过滤明显不是产品的行（合同条款、字段名、纯中文句子等）"""
    if df is None or df.empty:
        return df

    if 'model' not in df.columns:
        return df

    mask = pd.Series(True, index=df.index)
    for idx, row in df.iterrows():
        m = str(row.get('model', '')).strip()
        if not m:
            mask.at[idx] = False
            continue

        # Normalize non-breaking spaces
        m_norm = m.replace('\xa0', ' ')

        for pat in _NON_PRODUCT_MODEL_PATTERNS:
            if pat.search(m_norm):
                mask.at[idx] = False
                break

        if not mask.at[idx]:
            continue

        # Chinese sentence check (not a product)
        chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', m))
        if chinese_chars > 8 and len(m) > 12:
            mask.at[idx] = False

    return df[mask]


# ─── 关键词映射（来自共享常量） ───

try:
    from shared_keywords import (COLUMN_SIGNALS, COLUMN_SIGNALS_FLAT,
                                  SKIP_COLUMN_SIGNALS, PRICE_KEYWORDS,
                                  SKIP_HEADER_KEYWORDS,
                                  PRODUCT_SKIP_WORDS as SKIP_WORDS,
                                  CONTENT_SPEC_KEYWORDS)
except ImportError:
    # 回退：直接定义（防止独立运行时找不到 product_tool 模块）
    COLUMN_SIGNALS = {}  # 空值，运行时改为直接使用
    COLUMN_SIGNALS_FLAT = []
    SKIP_COLUMN_SIGNALS = []
    PRICE_KEYWORDS = []
    SKIP_HEADER_KEYWORDS = []
    CONTENT_SPEC_KEYWORDS = ['规格', '尺寸', '参数', '材质', '颜色']

# ─── 汇率 ───

def _get_usd_rate() -> float:
    try:
        from src.rates import get_rate
        return get_rate('USD', 'CNY') or 7.2
    except Exception:
        return 7.2

# ─── 价格提取 ───

def extract_price_from_value(text: str) -> Optional[dict]:
    """从文本提取价格，返回 {price_rmb, price_cny, currency} 或 None。"""
    if not text:
        return None
    text = str(text).strip()
    if not text:
        return None

    def _parse_num(s):
        s = s.strip().replace(',', '')
        return float(s) if s.replace('.', '').isdigit() else None

    # USD pattern
    m = re.search(r'(?:USD|US\s*Dollar)\s*([\d,]+(?:\.\d+)?)', text, re.I)
    if not m:
        m = re.search(r'([\d,]+(?:\.\d+)?)\s*(?:USD|US\s*Dollar)', text, re.I)
    if m:
        val = _parse_num(m.group(1))
        if val:
            rate = _get_usd_rate()
            return {'price_rmb': val, 'price_cny': round(val * rate, 2), 'currency': 'USD'}

    # CNY pattern
    m = re.search(r'(?:¥|CNY|RMB|人民币)\s*([\d,]+(?:\.\d+)?)', text, re.I)
    if not m:
        m = re.search(r'([\d,]+(?:\.\d+)?)\s*(?:元|RMB|CNY)', text, re.I)
    if m:
        val = _parse_num(m.group(1))
        if val:
            return {'price_rmb': val, 'price_cny': val, 'currency': 'CNY'}

    # Plain number
    m = re.search(r'(?<!\d)([\d,]+(?:\.\d+)?)(?!\d*\s*[WVKA%a-z]|\s*(?:Kg|KM|Nm|mm|Hz|rpm))', text)
    if m:
        val = _parse_num(m.group(1))
        if val and 0.01 < val < 1000000:
            return {'price_rmb': val, 'price_cny': val, 'currency': 'CNY'}
    return None


def clean_price_text(text) -> Optional[float]:
    if text is None:
        return None
    if isinstance(text, (int, float)):
        return float(text) if 0.01 < float(text) < 1000000 else None
    s = str(text).replace(',', '').replace('，', '')
    # 优先取 price/$/¥ 等标识后的数字（避免 'Size:15x10cm, Price:12.5' 取到 15）
    for prefix in ['price', '单价', '价格', '售价', '金额', 'total', '总额', '小计', '合计', '$', '¥', 'usd']:
        idx = s.lower().find(prefix)
        if idx >= 0:
            after = s[idx + len(prefix):]
            m = re.search(r'[\d]+(?:\.\d+)?', after)
            if m:
                try:
                    return float(m.group())
                except ValueError:
                    pass
    # 兜底: 取第一个数字
    m = re.search(r'(?:(?<![\d.])|^)[\d]+(?:\.\d+)?', s)
    if m:
        try:
            return float(m.group())
        except ValueError:
            return None
    return None


def sheet_to_markdown(ws, max_rows=15) -> str:
    """Convert worksheet to markdown (used by AI parser)."""
    rows = []
    col_limit = min(ws.max_column + 1, 15)
    for r in range(1, min(max_rows + 1, ws.max_row + 1)):
        row = [str(ws.cell(r, c).value or '').strip()
               for c in range(1, col_limit)]
        if any(row):
            rows.append(row)
    if not rows:
        return ''
    lines = ['| ' + ' | '.join(rows[0]) + ' |']
    lines.append('| ' + ' | '.join(['---'] * len(rows[0])) + ' |')
    for row in rows[1:]:
        lines.append('| ' + ' | '.join(row) + ' |')
    return '\n'.join(lines)


def _actual_cols(ws, row: int = 1) -> int:
    """返回指定行实际非空列数（上限 50 列）。"""
    limit = min(ws.max_column + 1, 50)
    return sum(1 for c in range(1, limit)
               if ws.cell(row, c).value is not None)


def _detect_currency(ws, header_row: int, price_col: int) -> str:
    """检查价格列头和数据值，判断币种。"""
    if price_col is None:
        return 'CNY'
    header = str(ws.cell(header_row, price_col + 1).value or '').lower()
    if any(kw in header for kw in ['usd', '$', 'fob', 'cif', 'cfr', 'dap', 'ddp']):
        return 'USD'
    # 价格列前几行数据含 $/USD 标记 → USD
    for r in range(header_row + 1, min(header_row + 5, ws.max_row + 1)):
        val = str(ws.cell(r, price_col + 1).value or '')
        if '$' in val or 'usd' in val.lower():
            return 'USD'
    return 'CNY'

# ─── 表头检测 ───

def _row_max_cols(ws, max_rows=15) -> int:
    """扫描前 N 行，返回最大实际列数（确定扫描宽度）。"""
    best = 0
    for r in range(1, min(ws.max_row + 1, max_rows + 1)):
        n = _actual_cols(ws, r)
        if n > best:
            best = n
    return best


def detect_header_row(ws) -> int:
    """扫描行 1-30，关键词匹配 + 非空列数评分。
    
    策略：
    - 根据前 15 行的最大列数确定扫描宽度
    - 长文本行（>30 字符）降权（极可能是公司信息/条款，非表头）
    - 非空列数太少且得分靠单一长文本的行大幅降权
    """
    best_row, best_score = 1, 0
    max_col = _row_max_cols(ws, 15) + 2
    if max_col < 4:
        max_col = 10
    if max_col > 60:
        max_col = 60
    for r in range(1, min(ws.max_row + 1, 30)):
        score = 0
        non_empty = 0
        long_text_penalty = 0
        for c in range(1, min(max_col, ws.max_column + 1)):
            val = str(ws.cell(r, c).value or '').strip()
            if val:
                non_empty += 1
                if len(val) > 30:
                    long_text_penalty += 3  # 长文本降权（公司介绍、条款等）
                vl = val.lower()
                if vl in [v.lower() for v in COLUMN_SIGNALS_FLAT]:
                    score += 2
                elif any(kw in vl for kw in ['model', 'name', 'price', 'spec', 'qty', '规格', '型号']):
                    score += 1
                if any(kw in vl for kw in PRICE_KEYWORDS):
                    score += 10 if r <= 15 else 3
        # 长文本降权: 减去长文本数
        score -= long_text_penalty
        # 如果非空列太少但得分高，说明靠单一长文本命中关键词 → 降权
        if non_empty <= 2 and long_text_penalty > 0:
            score *= 0.3
        if r > 1 and score >= 10 and non_empty >= 3:
            return r
        if score > best_score:
            best_score = score
            best_row = r
    return best_row if best_score >= 2 else 1


def _is_header_like(val: str) -> bool:
    """判断是否为非产品表头行（公司信息、条款等）。"""
    if not val:
        return False
    vl = val.lower()
    if any(kw in vl for kw in SKIP_HEADER_KEYWORDS):
        return True
    if len(val) > 30:
        return True
    return False

# ─── 列映射 ───

def _is_serial_col(ws, col: int, max_rows=20) -> bool:
    """判断某列是否为序号列（全部为纯数字 1..N）。"""
    count, num_count = 0, 0
    for r in range(1, min(ws.max_row + 1, max_rows + 1)):
        v = ws.cell(r, col).value
        if v is not None:
            count += 1
            s = str(v).strip()
            if s.isdigit() and 1 <= int(s) <= 9999:
                num_count += 1
    return count >= 3 and num_count >= count * 0.8


def map_columns(ws, header_row: int) -> dict:
    """关键词匹配（子串匹配）+ position 兜底。
    
    规则：
    - 不覆盖已匹配的列
    - 跳过含图片/序号关键词的列
    - 跳过全数字的序号列（不映射为 model）
    """
    col_map = {}
    actual = _actual_cols(ws, header_row)
    max_col = min(actual + 3, 60)
    # 按优先级迭代：model→spec→name→price→qty→packing→category→remark
    # name 最后才检测，避免 description 列被误判为 name
    _COLUMN_ORDER = ('model', 'spec', 'name', 'price', 'qty', 'packing', 'category', 'remark')
    for c in range(1, min(max_col, ws.max_column + 1)):
        val = str(ws.cell(header_row, c).value or '').strip().lower()
        # 跳过图片/序号列
        if any(sk in val for sk in SKIP_COLUMN_SIGNALS):
            continue
        for key in _COLUMN_ORDER:
            signals = COLUMN_SIGNALS.get(key, [])
            k = key + '_col'
            if k in col_map:
                continue
            if any(kw.lower() in val for kw in signals):
                # model 跳过序号列
                if key in ('model', 'name') and _is_serial_col(ws, c):
                    continue
                # model↔spec 冲突消解：表头同时含 model/spec 关键词且含强 spec 指示词
                if key == 'model' and any(kw.lower() in val for kw in COLUMN_SIGNALS['spec']):
                    strong_spec = ['specification', 'specifications', '规格型号', '参数', 'description']
                    strong_model = ['型号', '物料编码', '料号', '货号', 'product no',
                                    'part no', 'p/n', 'sku', 'item no', '产品编号', '料号']
                    if any(kw.lower() in val for kw in strong_spec) and not any(kw.lower() in val for kw in strong_model):
                        continue  # 是 spec 列，不是 model 列
                col_map[k] = c - 1
                break

    # Position 兜底（含图片列跳过）
    if 'name_col' not in col_map and 'model_col' in col_map:
        col_map['name_col'] = col_map['model_col']

    # model_col 不设默认值 → 避免把名称/文档列当型号
    if 'model_col' not in col_map:
        # 尝试从内容推断 model 列（找含最多型号代码的列）
        best_model_col = -1
        best_model_count = 0
        for c in range(min(max_col, 30)):
            model_count = 0
            content_count = 0
            for r in range(header_row + 1, min(header_row + 10, ws.max_row + 1, header_row + 30)):
                v = str(ws.cell(r, c + 1).value or '').strip()
                if v:
                    content_count += 1
                    if re.search(r'^[A-Za-z0-9][A-Za-z0-9\-_\./]+$', v) and not str(v).isdigit():
                        model_count += 1
            if content_count >= 2 and model_count >= content_count * 0.5 and model_count > best_model_count:
                best_model_count = model_count
                best_model_col = c
        if best_model_col >= 0:
            col_map['model_col'] = best_model_col

    if 'price_col' not in col_map:
        # 价格列兜底：找最后一列有数字数据的列
        available = sorted(set(range(max_col)) - {col_map.get(k) for k in col_map})
        if available:
            # 从后往前找含数字的列
            for c in reversed(available):
                for r in range(header_row + 1, min(header_row + 5, ws.max_row + 1)):
                    v = ws.cell(r, c + 1).value
                    if v is not None and isinstance(v, (int, float)):
                        col_map['price_col'] = c
                        break
                if 'price_col' in col_map:
                    break
            if 'price_col' not in col_map:
                col_map['price_col'] = available[-1]  # 兜底取最后一列

    # 币种检测：从表头行找 USD/FOB 标记
    price_col = col_map.get('price_col')
    if price_col is not None:
        header_text = str(ws.cell(header_row, price_col + 1).value or '').lower()
        if any(kw in header_text for kw in ['usd', 'fob', '$']):
            col_map['currency'] = 'USD'
        elif any(kw in header_text for kw in ['rmb', 'cny', '¥', '元']):
            col_map['currency'] = 'CNY'

    # spec_col 无检测时不强制兜底 → 避免型号/名称被塞进规格
    # 旧代码: if 'spec_col' not in col_map and 'name_col' in col_map: col_map['spec_col'] = name_col + 1
    # 改为: 只在确认 spec 有内容时才保留
    return col_map

# ─── KV 布局检测 ───

def _detect_kv_col(ws, header_row: int) -> int:
    """检测 KV 布局的起始列号（1-based），A 列空则返回 B 列。"""
    has_a = any(str(ws.cell(r, 1).value or '').strip()
                for r in range(header_row, min(header_row + 5, ws.max_row + 1)))
    if has_a:
        return 1
    has_b = any(str(ws.cell(r, 2).value or '').strip()
                for r in range(header_row, min(header_row + 5, ws.max_row + 1)))
    return 2 if has_b else 1


def is_kv_layout(ws, header_row: int) -> bool:
    """检测是否为 KV 布局（参数名含冒号）。
    
    支持数据从 B 列开始的情况（A 列为空，自动检测）。
    """
    kv_col = _detect_kv_col(ws, header_row)
    count_colon = 0
    count_total = 0
    for r in range(header_row, min(header_row + 10, ws.max_row + 1)):
        val = str(ws.cell(r, kv_col).value or '').strip()
        if val:
            count_total += 1
            if val.endswith(':') or val.endswith('：'):
                count_colon += 1
    return count_total >= 3 and count_colon >= count_total * 0.6


def _kv_offset(ws, header_row: int) -> int:
    """返回 KV 布局的 0-based 列偏移（A=0, B=1）。"""
    return _detect_kv_col(ws, header_row) - 1


def extract_kv_product(ws, header_row: int) -> pd.DataFrame:
    """提取单产品 KV 格式（支持列偏移）。"""
    off = _kv_offset(ws, header_row)
    model, price, specs = None, None, []
    for r in range(header_row, min(ws.max_row + 1, 300)):
        c1 = str(ws.cell(r, off + 1).value or '').strip()
        c2 = ws.cell(r, off + 2).value
        c3 = ws.cell(r, off + 3).value
        c2s = str(c2).strip() if c2 else ''
        c3s = str(c3).strip() if c3 else ''

        if not c1 and not c2s:
            continue

        # Model 检测
        if c1.lower() in ('model:', '型号:', 'item:'):
            model = c2s or c1
            continue

        # 价格检测
        price_info = extract_price_from_value(c1)
        if price_info:
            price = price_info
            continue
        price_info = extract_price_from_value(c2s)
        if price_info:
            price = price_info
            continue

        # 规格收集
        if c1 and (c2s or c3s):
            specs.append(f"{c1}: {c2s or c3s}")

    if not model:
        return pd.DataFrame()

    row = {'model': model, 'name_zh': model, 'spec_zh': '\n'.join(specs[:50]),
           'price_rmb': price['price_rmb'] if price else None,
           'price_cny': price['price_cny'] if price else None,
           'currency': price['currency'] if price else 'CNY', '_row': header_row}
    return pd.DataFrame([row])


def extract_multi_kv_products(ws, header_row: int) -> pd.DataFrame:
    """提取多产品 KV 格式（Model: 标记切分，支持列偏移）。"""
    off = _kv_offset(ws, header_row)
    products = []
    current = {}
    for r in range(header_row, min(ws.max_row + 1, 300)):
        c1 = str(ws.cell(r, off + 1).value or '').strip()
        c2 = ws.cell(r, off + 2).value
        c2s = str(c2).strip() if c2 else ''

        if not c1 and not c2s:
            continue

        if c1.lower() in ('model:', '型号:', 'item:'):
            if current.get('model'):
                products.append(current)
            current = {'model': c2s or c1, 'specs': [], 'price': None}
            continue

        if not current.get('model'):
            continue

        pi = extract_price_from_value(c1) or extract_price_from_value(c2s)
        if pi:
            current['price'] = pi
            continue

        if c1 and c2s:
            current['specs'].append(f"{c1}: {c2s}")

    if current.get('model'):
        products.append(current)

    result = []
    for p in products:
        pr = p.get('price') or {}
        result.append({'model': p['model'], 'name_zh': p['model'],
                       'spec_zh': '\n'.join(p.get('specs', [])[:50]),
                       'price_rmb': pr.get('price_rmb'),
                       'price_cny': pr.get('price_cny'),
                       'currency': pr.get('currency', 'CNY'), '_row': header_row})
    return pd.DataFrame(result)

# ─── 表格提取 ───

def _get_cell_val(ws, r, c, _cm):
    """获取单元格值（支持合并单元格传播）。_cm 为 _build_merge_map 返回的 cell_map。"""
    if _cm:
        val = _cm.get((r, c))
        if val is not None:
            return val
    return ws.cell(r, c).value


def _build_merge_map(ws):
    """构建合并单元格映射（含 O(1) 查找表）。"""
    mm = {}
    cell_map = {}
    try:
        for mg in ws.merged_cells.ranges:
            val = ws.cell(mg.min_row, mg.min_col).value
            mm[(mg.min_row, mg.min_col, mg.max_row, mg.max_col)] = val
            # 构建 O(1) 查找：每个子单元格 → 值
            for r in range(mg.min_row, mg.max_row + 1):
                for c in range(mg.min_col, mg.max_col + 1):
                    cell_map[(r, c)] = val
    except Exception:
        pass
    return mm, cell_map


def _is_product_row(first_cell: str, qty_val, price_val) -> bool:
    """判断一行是否是产品行。
    
    检测顺序：正向（价格/数量/型号）→ 跳词过滤 → 备选检测
    避免跳过词误杀包含关键词的产品名（如 'Total 1000'）。
    支持：
    - 中文品名（'棉签(Cotton swab)'）
    - 混合数量（'360袋/箱'）
    - 含单位价格（'0.15/片', '2.4元/片'）
    - 纯数字序号+品名列
    """
    if not first_cell or len(first_cell) > 200:
        return False
    # 跳过纯图片公式行
    if first_cell.startswith('=') and 'dispimg' in first_cell.lower():
        pass  # 靠下面的价格判断决定
    
    # ─── 快速否决：文档字段/编号模式 ───
    fc_lower = first_cell.lower().strip()
    # 订单号/合同号/参考号模式（这些看起来像型号但实际是文档编号）
    _doc_number_patterns = [
        re.compile(r'^(contract|order|po|p\.?o\.?|invoice|quote|ref|reference|date|payment|delivery|shipment|document|doc)(\s*(no|#|number|\.?))?\s*[:\-\.]?\s*[\w\-\d]+', re.I),
        re.compile(r'^(contract\s*no|order\s*no|po\s*no|invoice\s*no|ref\s*no|reference\s*no|doc\s*no|payment\s*no|shipment\s*no|delivery\s*no)', re.I),
        re.compile(r'^(合同编号|订单编号|订单号|发票号|参考号|文档号|付款编号|运输编号)', re.I),
        re.compile(r'^(date|dates?)\s*[:\-\.]?\s*[\d\-\/]+', re.I),
        # Payment method / delivery info lines
        re.compile(r'^(payment|delivery|shipment|shipping|port\s+of|discharge|loading)\s', re.I),
        # Company / seller / buyer info (these often appear as first column "model")
        re.compile(r'^(seller|buyer|supplier|customer|consignee|notify\s+party)\s*[:\-]', re.I),
        # Shipping marks / container info
        re.compile(r'^(marks?\s*(&|\s*no|\s*#)|shipping\s*mark|container\s*(no|#))', re.I),
        # Document type labels
        re.compile(r'^(proforma|commercial)\s+(invoice|inv)', re.I),
        re.compile(r'^(certificate\s+of|origin\s+certificate|fumigation|inspection\s+(report|cert))', re.I),
        # Chinese contract clauses
        re.compile(r'^(卖方|买方|供应商|客户|收货人|通知方|承运人)', re.I),
        # Numbered clauses in contracts
        re.compile(r'^\d+[.、\s]\s*(transshipment|payment|delivery|packing|insurance|bank|inspection|arbitration|force\s*majeure|shipping|terms?|conditions?|warranty|validity|quality)', re.I),
        # Pure clause text lines
        re.compile(r'^(仲裁|保险|商检|产地|原产|信用证|l/?c|t/?t|不可抗力)', re.I),
        # Signature / approval lines
        re.compile(r'^(sign(ed|ature)?|approv(ed|al)?|authoriz(ed|ation)?|seal|stamp|公章|签字|签名|盖章|审批)', re.I),
        # Packing list headers misidentified as products
        re.compile(r'^(packing\s*(list|detail|info)|装箱(单|明细|清单)|mark\s*(nos?|#))', re.I),
    ]
    for pat in _doc_number_patterns:
        if pat.search(fc_lower):
            return False
    
    # ─── 正向检测（优先于跳过词，防止误杀） ───
    # 型号模式：字母+数字
    if re.search(r'[A-Za-z]+\d+', first_cell) and len(first_cell) < 40:
        # 额外检查：如果是订单号/合同号风格（字母部分在头部，然后是数字），过滤
        # 如 "PO20250516", "CT20250516", "INV20250516"
        if re.match(r'^[A-Za-z]{2,4}[\d\-]{6,15}$', first_cell) and len(first_cell) <= 20:
            # 可能: "PO20250516" -> 这是订单号，不是产品。只有短型号才放过（≤12字符）
            if len(first_cell) > 12:
                return False
        return True
    # 有价格（用 clean_price_text 支持 '0.15/片'）
    clean_p = clean_price_text(price_val)
    if clean_p is not None and clean_p > 0:
        return True
    # 数量中提取数字
    try:
        qty_str = str(qty_val)
        qty_nums = re.findall(r'\d+', qty_str)
        if qty_nums and int(qty_nums[0]) > 0 and len(first_cell) < 100:
            return True
    except (ValueError, TypeError):
        pass
    # ─── 跳过词（避免非产品行混入，来自共享常量） ───
    try:
        from shared_keywords import PRODUCT_SKIP_WORDS
        _skip = PRODUCT_SKIP_WORDS
    except ImportError:
        _skip = ['条款', '备注', '说明', '合计', 'total', 'subtotal',
                 '小计', '汇总', '总计', '报关', '合同', 'header',
                 'warranty', 'oem', 'validity', 'payment',
                 '包邮', '执行', '经销价', '件以下', '件起', '非偏远',
                 '系统价格', '零售价', '批发价', '不含税', '含税价']
    if any(kw in fc_lower for kw in _skip):
        return False
    # ─── 定价条款预检（数字起头 + 以下/以上/包邮 → 非产品） ───
    if re.search(r'^\d+', first_cell) and any(kw in fc_lower for kw in ['以下', '以上', '包邮', '执行']):
        return False
    # ─── 备选检测 ───
    # 第一格含常见产品标识（括号、型号风格、英文品名含空格两位以上）
    if re.search(r'[（(][^）)]*[)）]', first_cell) and len(first_cell) < 60:
        return True
    if re.search(r'[A-Za-z]+[\d\-/]+', first_cell):
        return True
    if re.search(r'[A-Z][a-z]+\s[A-Z][a-z]+', first_cell):  # 英文品名如 'Mouthwash tablets'
        return True
    return False


def _is_battery_or_charger(first_cell: str) -> bool:
    """判断是否为电池/充电器行（仅检查第一个单元格，避免配件描述含 'battery' 误匹配）。"""
    if not first_cell:
        return False
    rl = first_cell.lower()
    return any(kw in rl for kw in ['电池', '充电器', 'battery', 'charger'])


def _is_remark_row(first_cell: str) -> bool:
    if not first_cell:
        return True
    rl = first_cell.lower()
    return any(kw in rl for kw in ['备注', 'remark', 'remarks', '条款', '包装'])


def parse_with_colmap(ws, header_row: int, col_map: dict) -> pd.DataFrame:
    """从表格布局提取产品（支持合并单元格、备注附着、电池行）。
    
    当 header_row=0 时，视为无表头文件，所有行均为数据行。
    """
    if header_row < 1:
        header_row = 0
        actual = _row_max_cols(ws, max_rows=15)
    else:
        actual = _actual_cols(ws, header_row)
    max_col = min(actual + 3, 60)
    if max_col < 2:
        max_col = min(ws.max_column + 1, 10)
    merge_map, _merge_cells = _build_merge_map(ws)
    currency = _detect_currency(ws, max(header_row, 1), col_map.get('price_col'))
    packing_col = col_map.get('packing_col')
    remark_col = col_map.get('remark_col')
    name_col = col_map.get('name_col')
    price_col = col_map.get('price_col')
    model_col = col_map.get('model_col')
    spec_col = col_map.get('spec_col')
    qty_col = col_map.get('qty_col')

    result = []
    _empty_run = 0
    _start = 1 if header_row < 1 else header_row + 1
    _row_limit = min(ws.max_row + 1, _start + 2000)
    for r in range(_start, _row_limit):
        vals = [str(_get_cell_val(ws, r, c, _merge_cells) or '').strip() for c in range(1, min(max_col, ws.max_column + 1))]

        # 连续 20 空行 → 终止
        if not any(vals):
            _empty_run += 1
            if _empty_run >= 20:
                break
            continue
        _empty_run = 0

        name = vals[name_col] if name_col is not None and name_col < len(vals) else ''
        model = vals[model_col] if model_col is not None and model_col < len(vals) else (name or '')
        # 安全网：model_col 和 spec_col 指向同一列 → 用 name 做 model
        if model_col is not None and spec_col is not None and model_col == spec_col:
            if name_col is not None and name_col != model_col:
                model = name
        # 如果 model 远超 name 长度（如取到了规格列），回退到 name
        if model and name and len(model) > len(name) * 3:
            model = name
        # 如果 model 是纯数字序号而 name 有内容，用 name
        if model and name and str(model).strip().isdigit():
            model = name
        # 如果 model 是图片公式，用 name 或下一有内容的列
        if model and ('dispimg' in model.lower() or model.startswith('=')):
            model = name or (vals[min(name_col+1 if name_col else 1, len(vals)-1)] if len(vals) > 1 else '')

        price_raw = _get_cell_val(ws, r, price_col + 1, _merge_cells) if price_col is not None else None
        price = clean_price_text(price_raw)
        qty = vals[qty_col] if qty_col is not None and qty_col < len(vals) else ''
        spec = vals[spec_col] if spec_col is not None and spec_col < len(vals) else ''
        # 提取 packing/箱规信息追加到 spec_zh
        packing_parts = []
        if packing_col is not None and packing_col < len(vals):
            pv = vals[packing_col]
            if pv:
                packing_parts.append(str(pv))
        first_cell = model or name or next((v for v in vals if v), '')

        row_text = ' '.join(v for v in vals if v)
        remark_text = ' '.join(v for i, v in enumerate(vals) if v and i not in {price_col, name_col, model_col} if i is not None)

        if _is_remark_row(first_cell) and result:
            result[-1]['remark'] = (result[-1].get('remark', '') + ' ' + remark_text).strip()
            continue
        if _is_battery_or_charger(first_cell) and result:
            label = '电池' if any(kw in row_text.lower() for kw in ['电池', 'battery']) else '充电器'
            result[-1]['remark'] = (result[-1].get('remark', '') + f' [{label}] {remark_text}').strip()
            continue

        if not _is_product_row(first_cell, qty, price_raw):
            continue

        price_cny = round(price * _get_usd_rate(), 2) if currency == 'USD' and price else (price or 0)
        result.append({
            'model': model or name or f"产品_R{r}",
            'name_zh': name or model or '',
            'spec_zh': '; '.join(filter(None, [spec] + packing_parts)) or '',
            'price_rmb': price,
            'price_cny': price_cny,
            'qty': qty,
            'currency': currency,
            'remark': '',
            'category': '',
            '_row': r,
            '_sheet': ws.title,
        })

    return pd.DataFrame(result)


# ─── 内容驱动列推断 ───

def _classify_columns_by_content(ws, header_row: int) -> dict:
    """根据列内容分布推断列角色。"""
    if header_row < 1:
        start_row = 1
        actual = _row_max_cols(ws, max_rows=15)
    else:
        start_row = header_row + 1
        actual = _actual_cols(ws, header_row)
    max_col = actual + 3
    if max_col > 50 or max_col < 3:
        max_col = 30
    col_scores = {c: {'model': 0, 'price': 0, 'spec': 0, 'qty': 0, 'skip': 0, 'total': 0} for c in range(1, max_col + 1)}

    for r in range(start_row, min(ws.max_row + 1, start_row + 200)):
        for c in range(1, max_col + 1):
            val = str(ws.cell(r, c).value or '').strip()
            if not val:
                continue
            col_scores[c]['total'] += 1
            vl = val.lower()
            # 型号检测：字母开头+至少2位数字，长度<30，不含中文/单位
            if (re.search(r'^[A-Za-z]+[-]?\d{2,}$', val) and len(val) < 30
                    and not re.search(r'[\u4e00-\u9fff]|cm|mm|kg|inch|volt|amp|watt|hz|rpm|pcs|set|unit', val.lower())):
                col_scores[c]['model'] += 2
            price_m = re.search(r'[\d,]+(?:\.\d+)?', val)
            if price_m:
                try:
                    p = float(price_m.group().replace(',', ''))
                    if 0.01 < p < 1000000:
                        col_scores[c]['price'] += 1
                except ValueError:
                    pass
            if any(kw in vl for kw in CONTENT_SPEC_KEYWORDS):
                col_scores[c]['spec'] += 1
            if val.isdigit() and 1 <= int(val) <= 9999:
                col_scores[c]['qty'] += 1
            if any(kw in vl for kw in ['序号', 'no.', 'serial', '图片', 'photo', 'image']):
                col_scores[c]['skip'] += 5

    col_map = {}
    for c, scores in col_scores.items():
        if scores['skip'] > scores['total'] * 0.3 and scores['total'] > 2:
            continue
        best = max(['model', 'price', 'spec', 'qty'], key=lambda k: scores[k] + (scores['total'] * 0.1))
        if scores[best] >= 2:
            col_map[best + '_col'] = c - 1
    return col_map


def parse_by_content(ws, header_row: int) -> pd.DataFrame:
    """无视表头，按内容分析提取产品。"""
    col_map = _classify_columns_by_content(ws, header_row)
    return parse_with_colmap(ws, header_row, col_map)


# ─── 评分 ───

def score_result(df) -> float:
    """评分解析结果。"""
    if not isinstance(df, pd.DataFrame):
        return -1
    if df is None or df.empty:
        return -1
    n = len(df)
    has_model = sum(1 for _, r in df.iterrows() if r.get('model') and len(str(r['model']).strip()) > 1)
    has_price = sum(1 for _, r in df.iterrows() if r.get('price_rmb') is not None and r['price_rmb'] > 0)
    models = set(str(r.get('model', '')).strip().lower() for _, r in df.iterrows() if r.get('model'))
    diversity = len(models) / max(has_model, 1)
    score = n * 0.3 + has_model * 0.3 + has_price * 0.2 + diversity * 10 * 0.2
    # 产品_ 前缀惩罚
    prefix_penalty = sum(1 for _, r in df.iterrows() if str(r.get('model', '')).startswith('产品_'))
    if prefix_penalty > n * 0.5:
        score *= 0.5
    return score


# ─── 入口 ───

def _parse_sheet(ws) -> pd.DataFrame:
    """对单个 sheet 执行 4 层策略解析（带提前终止）。"""
    # 跳过生成的文件
    first_text = str(ws.cell(1, 1).value or '').lower()
    if 'foreign trade quotation' in first_text or 'proforma invoice' in first_text:
        return pd.DataFrame()

    header_row = detect_header_row(ws)
    candidates = []

    # 策略1: KV 布局
    if is_kv_layout(ws, header_row):
        df_kv = extract_multi_kv_products(ws, header_row)
        if df_kv.empty:
            df_kv = extract_kv_product(ws, header_row)
        if not df_kv.empty:
            score = score_result(df_kv)
            candidates.append((df_kv, 'kv', score))
            if score >= 5 or (len(df_kv) >= 5 and score >= 3):
                df_best = max(candidates, key=lambda c: c[2])[0]
                df_best = _filter_non_product_rows(df_best)
                if not df_best.empty:
                    return df_best

    # 策略2: 表格布局（关键词映射）
    col_map = map_columns(ws, header_row)
    if col_map:
        df_tbl = parse_with_colmap(ws, header_row, col_map)
        if not df_tbl.empty:
            score = score_result(df_tbl)
            candidates.append((df_tbl, 'table', score))
            if score >= 5 or (len(df_tbl) >= 5 and score >= 3):
                df_best = max(candidates, key=lambda c: c[2])[0]
                df_best = _filter_non_product_rows(df_best)
                if not df_best.empty:
                    return df_best

    # 策略3: 内容驱动
    df_content = parse_by_content(ws, header_row)
    if not df_content.empty:
        score = score_result(df_content)
        candidates.append((df_content, 'content', score))
        if score >= 5 or (len(df_content) >= 5 and score >= 3):
            df_best = max(candidates, key=lambda c: c[2])[0]
            df_best = _filter_non_product_rows(df_best)
            if not df_best.empty:
                return df_best

    # 策略4: 无表头模式
    if header_row <= 1:
        df_no_header = parse_by_content(ws, 0)
        if not df_no_header.empty:
            score = score_result(df_no_header)
            candidates.append((df_no_header, 'content', score))

    candidates = [(df, t) for df, t, _ in candidates if isinstance(df, pd.DataFrame) and not df.empty]
    if candidates:
        best = max(candidates, key=lambda c: score_result(c[0]))[0]
        best = _filter_non_product_rows(best)
        return best
    return pd.DataFrame()


def parse(file_path: str):
    """
    主入口：遍历所有 Sheet，每 Sheet 多策略并行 → 合并结果。
    返回 (DataFrame, parse_type, count, cache_key)。
    """
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        logger.error(f"load_workbook failed: {e}")
        return pd.DataFrame(), 'error', 0, ''

    if not wb.sheetnames:
        wb.close()
        return pd.DataFrame(), 'empty', 0, ''

    all_results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws is None:
            continue
        df_sheet = _parse_sheet(ws)
        if not df_sheet.empty:
            if '_sheet' not in df_sheet.columns:
                df_sheet['_sheet'] = sheet_name
            all_results.append(df_sheet)

    wb.close()

    if not all_results:
        return pd.DataFrame(), 'empty', 0, ''

    if len(all_results) == 1:
        best_df = all_results[0]
    else:
        best_df = pd.concat(all_results, ignore_index=True)

    best_type = 'multi' if len(all_results) > 1 else 'table'

    return best_df, best_type, len(best_df), ''
