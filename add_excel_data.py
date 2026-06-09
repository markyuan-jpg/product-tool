import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

src = r'C:\Users\Administrator\Desktop\报价单_20260520195903.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb['Quotation']

# 20尺柜 data (same order as before)
container20_data = [
    270,    # 0  Free examination gloves
    238,    # 1  syringes (1ml)
    222,    # 2  syringes (3ml)
    202,    # 3  syringes (5ml)
    238,    # 4  syringes (10ml)
    900,    # 5  kids' tablet
    490,    # 6  handheld fans 行4
    361,    # 7  handheld fans 行5
    490,    # 8  handheld fans 行6
    900,    # 9  HIV cassette
    900,    # 10 HIV strip
    900,    # 11 malaria
    540,    # 12 small speakers 行10
    540,    # 13 small speakers 行11
    947,    # 14 tongue depressors 行12
    947,    # 15 tongue depressors 行13
    441,    # 16 cotton balls 100pc
    441,    # 17 cotton balls 200pc
    259,    # 18 phone HMT-MT002
    798,    # 19 phone SC-ZM003
    429,    # 20 phone SC-CFK005
    211,    # 21 phone SC-XP002
    900,    # 22 charging 5V2A
    900,    # 23 charging 5V1A
    900,    # 24 Ear phones
]

mapping = {
    6: [0], 7: [1,2,3,4], 8: [5],
    9: [6], 10: [7], 11: [8],
    12: [9], 13: [10], 14: [11],
    15: [12], 16: [13], 17: [14], 18: [15],
    19: [16,17],
    20: [18], 21: [19], 22: [20], 23: [21],
    24: [22], 25: [23], 26: [24],
}

# ========== 1. Add 20尺柜 column (K) ==========
ws.cell(row=5, column=11).value = '20尺柜装载量（个外箱）'
for e_row, indices in mapping.items():
    vals = [str(container20_data[i]) for i in indices]
    ws.cell(row=e_row, column=11).value = ' / '.join(vals)

# ========== 2. Format columns H, I, J, K to match existing style ==========

# Existing header style from column A5
header_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
header_fill = PatternFill(patternType='solid', fgColor='FF4472C4')
header_align = Alignment(horizontal='center', vertical='center', wrapText=True)
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

# Data font/style (matching row 6 col A style)
data_font = Font(name='Arial', size=10, color='FF000000')
data_align = Alignment(horizontal='center', vertical='center', wrapText=True)

# Alternating row fills
fill_white = PatternFill(patternType=None)
fill_gray = PatternFill(patternType='solid', fgColor='FFE7E3E6')

# Format headers (row 5, columns H-K)
for col in [8, 9, 10, 11]:
    cell = ws.cell(row=5, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Format data rows (6-26, columns H-K)
for row_num in range(6, 27):
    row_fill = fill_white if (row_num % 2 == 0) else fill_gray
    for col in [8, 9, 10, 11]:
        cell = ws.cell(row=row_num, column=col)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border
        cell.fill = row_fill

# Also format empty rows 27+ for H-K (to match theme)
for row_num in range(27, 37):
    for col in [8, 9, 10, 11]:
        cell = ws.cell(row=row_num, column=col)
        cell.border = Border(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
        )

# ========== 3. Set column widths ==========
ws.column_dimensions['H'].width = 20
ws.column_dimensions['I'].width = 18
ws.column_dimensions['J'].width = 22
ws.column_dimensions['K'].width = 22

# ========== 4. Number formats ==========
# H, I (Carton/Box Size): decimal with 4 places
for row_num in range(6, 27):
    cell_h = ws.cell(row=row_num, column=8)
    if cell_h.value and ' / ' not in str(cell_h.value):
        cell_h.number_format = '0.0000'
    cell_i = ws.cell(row=row_num, column=9)
    if cell_i.value and ' / ' not in str(cell_i.value):
        cell_i.number_format = '0.0000'
    cell_j = ws.cell(row=row_num, column=10)
    if cell_j.value and ' / ' not in str(cell_j.value):
        cell_j.number_format = '#,##0'
    cell_k = ws.cell(row=row_num, column=11)
    if cell_k.value and ' / ' not in str(cell_k.value):
        cell_k.number_format = '#,##0'

wb.save(src)

# Verify
print('=== Verification ===')
for e_row in sorted(mapping.keys()):
    name = (ws.cell(row=e_row, column=3).value or '')[:30]
    h = ws.cell(row=e_row, column=8).value or ''
    i = ws.cell(row=e_row, column=9).value or ''
    j = ws.cell(row=e_row, column=10).value or ''
    k = ws.cell(row=e_row, column=11).value or ''
    print('Row {:2d} | {:<30s} | C={:<25s} | B={:<25s} | 40={:<20s} | 20={}'.format(
        e_row, name, str(h), str(i), str(j), str(k)))

print('\nDone! 4 columns (H-K) added and formatted.')
print('H: Carton Size (m3) - 4 decimal places')
print('I: Box Size (m3) - 4 decimal places')
print('J: 40尺柜装载量 - integer format')
print('K: 20尺柜装载量 - integer format')
print('Style: Arial 10, Blue header, alternating rows, thin borders')
