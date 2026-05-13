# -*- coding: utf-8 -*-

# -*- coding: utf-8 -*-

"""

Packing list & commercial invoice generator

"""

import os, re, logging

from datetime import datetime

from typing import Dict, List, Optional

from openpyxl import Workbook

from openpyxl.utils import get_column_letter

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.output.doc_shared import translate_items, payment_by_lang, get_seller_info



#   

FONT_TITLE = Font(name='Arial', size=14, bold=True, color='1F4E79')

FONT_SECTION = Font(name='Arial', size=10, bold=True)

FONT_NORMAL = Font(name='Arial', size=9)

FONT_BOLD = Font(name='Arial', size=9, bold=True)

FONT_SMALL = Font(name='Arial', size=8)

FONT_TABLE_HEADER = Font(name='Arial', size=9, bold=True, color='FFFFFF')

FILL_BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

FILL_LIGHT = PatternFill(start_color='F5F8FC', end_color='F5F8FC', fill_type='solid')

FILL_TOTAL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)

THIN = Side(style='thin', color='CCCCCC')

BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)





def generate_packing_list(

    pi_items: List[Dict],

    invoice_no: str,

    invoice_date: str,

    buyer_name: str = "",

    buyer_address: str = "",

    output_path: str = None,

    company_config: dict = None,

    lang: str = "chinese",

    port_loading: str = "",

    port_discharge: str = "XXXXX",

    vessel: str = "",

    bl_no: str = "",

    packing_type: str = "CARTON",

    packing_qty: str = "",

) -> str:

    """Generate Packing List"""

    import traceback

    try:

        if not output_path:

            output_path = f"PackingList_{invoice_no}.xlsx"

        pi_items = translate_items(pi_items, lang)

        logging.info(f"[PACKING] lang={lang}, items={len(pi_items)}")

    except Exception as e:

        logging.error(f"[PACKING] init failed: {e}\n{traceback.format_exc()}")

        raise



    try:

        from src.utils.translator import translate_doc as _td

    except Exception:

        _td = lambda t, l: t



    try:

        from src.output.excel_template import get_template_path, apply_template

        tmpl = get_template_path('packing')

        if tmpl and apply_template(pi_items, tmpl, str(output_path)):

            return output_path

    except Exception:

        pass



    seller = get_seller_info(company_config)

    seller_name = seller['company']

    seller_addr = seller['address']



    wb = Workbook()

    ws = wb.active

    ws.title = "Packing List"

    row = 1



    ws.merge_cells('A1:I1')

    ws.cell(1, 1).value = _td('PACKING LIST', lang)

    ws.cell(1, 1).font = FONT_TITLE

    ws.cell(1, 1).alignment = ALIGN_CENTER

    ws.row_dimensions[1].height = 28

    row = 3



    ws.cell(row, 1).value = _td('1. Shipper (Exporter):', lang); ws.cell(row, 1).font = FONT_SECTION; ws.merge_cells(f'A{row}:D{row}'); row += 1

    ws.cell(row, 1).value = seller_name; ws.cell(row, 1).font = FONT_BOLD; ws.merge_cells(f'A{row}:D{row}'); row += 1

    ws.cell(row, 1).value = seller_addr; ws.cell(row, 1).font = FONT_NORMAL; ws.merge_cells(f'A{row}:D{row}'); row += 1

    ws.cell(row-3, 5).value = _td('Packing List No.', lang); ws.cell(row-3, 5).font = FONT_SECTION

    ws.cell(row-3, 6).value = invoice_no; ws.cell(row-3, 6).font = FONT_NORMAL

    ws.cell(row-2, 5).value = _td('Date', lang); ws.cell(row-2, 5).font = FONT_SECTION

    ws.cell(row-2, 6).value = invoice_date; ws.cell(row-2, 6).font = FONT_NORMAL

    ws.cell(row-1, 5).value = _td('Invoice No.', lang); ws.cell(row-1, 5).font = FONT_SECTION

    ws.cell(row-1, 6).value = invoice_no; ws.cell(row-1, 6).font = FONT_NORMAL; row += 1

    ws.cell(row, 1).value = _td('2. Consignee (Buyer):', lang); ws.cell(row, 1).font = FONT_SECTION; ws.merge_cells(f'A{row}:D{row}'); row += 1

    ws.cell(row, 1).value = buyer_name or 'XXXXX'; ws.cell(row, 1).font = FONT_NORMAL; ws.merge_cells(f'A{row}:D{row}'); row += 1

    if buyer_address: ws.cell(row, 1).value = buyer_address; ws.cell(row, 1).font = FONT_NORMAL; ws.merge_cells(f'A{row}:D{row}'); row += 1

    row += 1

    ws.cell(row, 1).value = _td('3. Transport Details:', lang); ws.cell(row, 1).font = FONT_SECTION; ws.merge_cells(f'A{row}:I{row}'); row += 1

    ws.cell(row, 1).value = _td('Port of Loading:', lang); ws.cell(row, 1).font = FONT_BOLD

    ws.cell(row, 2).value = port_loading or 'XXXXX'; ws.cell(row, 2).font = FONT_NORMAL

    ws.cell(row, 4).value = _td('Vessel/Flight:', lang); ws.cell(row, 4).font = FONT_BOLD

    ws.cell(row, 5).value = vessel or 'XXXXX'; ws.cell(row, 5).font = FONT_NORMAL

    ws.cell(row, 7).value = _td('B/L No.:', lang); ws.cell(row, 7).font = FONT_BOLD

    ws.cell(row, 8).value = bl_no or 'XXXXX'; ws.cell(row, 8).font = FONT_NORMAL; row += 1

    ws.cell(row, 1).value = _td('Port of Discharge:', lang); ws.cell(row, 1).font = FONT_BOLD

    ws.cell(row, 2).value = port_discharge or 'XXXXX'; ws.cell(row, 2).font = FONT_NORMAL; row += 2

    ws.cell(row, 1).value = _td('4. Shipping Marks:', lang); ws.cell(row, 1).font = FONT_SECTION; row += 1

    ws.cell(row, 1).value = _td('N/M', lang); ws.cell(row, 1).font = FONT_NORMAL; row += 2

    ws.cell(row, 1).value = _td('5. No. of Packages:', lang); ws.cell(row, 1).font = FONT_SECTION; row += 1



    headers = [_td(h, lang) for h in ['Marks & Nos.', 'Description of Goods', 'Qty', 'NW (kg)', 'GW (kg)', 'Meas (m\u00b3)', 'Carton Size (cm)', 'Qty/Carton']]

    widths = [14, 40, 12, 12, 12, 12, 18, 14]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):

        ws.cell(row, ci).value = h; ws.cell(row, ci).font = FONT_TABLE_HEADER; ws.cell(row, ci).fill = FILL_BLUE

        ws.cell(row, ci).alignment = ALIGN_CENTER; ws.cell(row, ci).border = BORDER

        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[row].height = 28; row += 1



    total_qty=0; total_nw=0; total_gw=0; total_meas=0; carton_count=0

    for i, item in enumerate(pi_items, 1):

        qty = int(item.get('qty', item.get('quantity', 1)))

        nw_raw = item.get('net_weight') or item.get('nw', 0)

        nw = float(nw_raw) if nw_raw else None

        gw_raw = item.get('gross_weight') or item.get('gw', 0)

        gw = float(gw_raw) if gw_raw else None

        meas = float(item.get('cbm', 0)) if item.get('cbm') else None

        ctn_size = str(item.get('carton_size', '') or '')

        upc_raw = item.get('units_per_carton', 0)

        upc = int(upc_raw) if upc_raw else 0

        ctns = max(1, (qty+upc-1)//upc) if upc>0 else None

        dims = ctn_size or ''  # ?

        model = item.get('model',''); desc = item.get('name_zh','') or model

        vals = ['N/M', f'{model}\n{desc}', qty,

                round(nw*qty,2) if nw is not None else '',

                round(gw*qty,2) if gw is not None else '',

                round(meas*ctns,3) if meas is not None and ctns else '',

                dims, str(upc) if upc else '']

        for ci, v in enumerate(vals, 1):

            ws.cell(row, ci).value = v; ws.cell(row, ci).font = FONT_NORMAL; ws.cell(row, ci).border = BORDER

            ws.cell(row, ci).alignment = ALIGN_CENTER if ci==1 else ALIGN_LEFT if ci==2 else ALIGN_RIGHT

        if i%2==0:

            for ci in range(1,9): ws.cell(row, ci).fill = FILL_LIGHT

        total_qty+=qty

        if nw is not None: total_nw+=nw*qty

        if gw is not None: total_gw+=gw*qty

        if meas is not None and ctns: total_meas+=meas*ctns

        if ctns: carton_count+=ctns

        row += 1



    total_nw_display = round(total_nw,2) if total_nw else ''

    total_gw_display = round(total_gw,2) if total_gw else ''

    total_meas_display = round(total_meas,3) if total_meas else ''

    total_vals = ['', _td('TOTAL', lang), total_qty, total_nw_display, total_gw_display, total_meas_display, '', '']

    for ci, v in enumerate(total_vals, 1):

        ws.cell(row, ci).value = v; ws.cell(row, ci).font = Font(name='Arial',size=9,bold=True); ws.cell(row, ci).border = BORDER

        ws.cell(row, ci).fill = FILL_TOTAL; ws.cell(row, ci).alignment = ALIGN_RIGHT if ci>2 else ALIGN_CENTER

    row += 2

    pk = packing_qty or (str(carton_count) if carton_count else '')
    pt = packing_type or 'CARTON'

    ws.cell(row, 1).value = _td(f'6. Total Packages (in words): SAY {pk} {pt.upper()}{"S" if pk.isdigit() and int(pk)>1 else ""} ONLY.', lang)

    ws.cell(row, 1).font = FONT_SECTION; ws.merge_cells(f'A{row}:H{row}'); row += 2

    ws.cell(row, 1).value = _td('7. Remarks:', lang); ws.cell(row, 1).font = FONT_SECTION; row += 1

    ws.cell(row, 1).value = '[Palletized]'; ws.cell(row, 1).font = Font(name='Arial',size=8,color='999999',italic=True); ws.merge_cells(f'A{row}:H{row}'); row += 2

    ws.cell(row, 1).value = _td('8. Signature:', lang); ws.cell(row, 1).font = FONT_SECTION; row += 1

    ws.cell(row, 1).value = seller_name; ws.cell(row, 1).font = FONT_BOLD; row += 1

    ws.cell(row, 1).value = _td('(signed & stamped)', lang); ws.cell(row, 1).font = Font(name='Arial',size=8,color='888888')

    wb.save(output_path)

    logging.info(f"[PACKING] successfully generated: {output_path}")

    return output_path





def generate_commercial_invoice(

    pi_items: List[Dict],

    invoice_no: str,

    invoice_date: str,

    buyer_name: str = "",

    buyer_address: str = "",

    output_path: str = None,

    company_config: dict = None,

    lang: str = "chinese",

    trade_terms: str = "FOB XXXXX",

    payment_terms: str = "",

    currency: str = "CNY",

    port_loading: str = "",

    port_discharge: str = "XXXXX",

    vessel: str = "",

    bl_no: str = "",

    origin_country: str = "",

) -> str:

    """Generate Commercial Invoice"""

    import traceback

    try:

        if not output_path:

            output_path = f"CommercialInvoice_{invoice_no}.xlsx"

        pi_items = translate_items(pi_items, lang)

        logging.info(f"[INVOICE] lang={lang}, items={len(pi_items)}")

    except Exception as e:

        logging.error(f"[INVOICE] init failed: {e}\n{traceback.format_exc()}")

        raise



    try:

        from src.utils.translator import translate_doc as _td

    except Exception:

        _td = lambda t, l: t



    try:

        from src.output.excel_template import get_template_path, apply_template

        tmpl = get_template_path('invoice')

        if tmpl and apply_template(pi_items, tmpl, str(output_path)):

            return output_path

    except Exception:

        pass



    seller = get_seller_info(company_config)

    seller_name = seller['company']

    seller_addr = seller['address']

    seller_tel = seller['phone']

    seller_email = seller.get('email', '') or 'XXXXX'



    wb = Workbook()

    ws = wb.active

    ws.title = "Commercial Invoice"

    row = 1

    ws.merge_cells('A1:H1')

    ws.cell(1, 1).value = _td('COMMERCIAL INVOICE', lang)

    ws.cell(1, 1).font = FONT_TITLE

    ws.cell(1, 1).alignment = ALIGN_CENTER

    ws.row_dimensions[1].height = 28

    row = 3

    ws.cell(row, 1).value = _td('1. Seller:', lang); ws.cell(row, 1).font = FONT_SECTION

    ws.cell(row, 5).value = _td('Invoice No.', lang); ws.cell(row, 5).font = FONT_SECTION

    ws.cell(row, 6).value = invoice_no; ws.cell(row, 6).font = FONT_NORMAL; row += 1

    ws.cell(row, 1).value = seller_name; ws.cell(row, 1).font = FONT_BOLD; ws.merge_cells(f'A{row}:D{row}')

    ws.cell(row, 5).value = _td('Date', lang); ws.cell(row, 5).font = FONT_SECTION

    ws.cell(row, 6).value = invoice_date; ws.cell(row, 6).font = FONT_NORMAL; row += 1

    ws.cell(row, 1).value = seller_addr; ws.cell(row, 1).font = FONT_NORMAL; ws.merge_cells(f'A{row}:D{row}')

    ws.cell(row, 5).value = _td('S/C No.', lang); ws.cell(row, 5).font = FONT_SECTION

    ws.cell(row, 6).value = invoice_no; ws.cell(row, 6).font = FONT_NORMAL; row += 1

    ws.cell(row, 1).value = f'T: {seller_tel}  E: {seller_email}'; ws.cell(row, 1).font = FONT_NORMAL; ws.merge_cells(f'A{row}:D{row}')

    ws.cell(row, 5).value = _td('L/C No.', lang); ws.cell(row, 5).font = FONT_SECTION

    ws.cell(row, 6).value = ''; ws.cell(row, 6).font = FONT_NORMAL; row += 2

    ws.cell(row, 1).value = _td('2. Buyer:', lang); ws.cell(row, 1).font = FONT_SECTION; ws.merge_cells(f'A{row}:D{row}'); row += 1

    ws.cell(row, 1).value = buyer_name or 'XXXXX'; ws.cell(row, 1).font = FONT_NORMAL; ws.merge_cells(f'A{row}:D{row}'); row += 1

    ws.cell(row, 1).value = buyer_address or ''; ws.cell(row, 1).font = FONT_NORMAL; ws.merge_cells(f'A{row}:D{row}'); row += 2

    ws.cell(row, 1).value = _td('3. Transport Details:', lang); ws.cell(row, 1).font = FONT_SECTION; ws.merge_cells(f'A{row}:H{row}'); row += 1

    ws.cell(row, 1).value = _td('Port of Loading:', lang); ws.cell(row, 1).font = FONT_BOLD

    ws.cell(row, 2).value = port_loading or 'XXXXX'; ws.cell(row, 2).font = FONT_NORMAL

    ws.cell(row, 5).value = _td('Payment Terms:', lang); ws.cell(row, 5).font = FONT_BOLD

    pt_value = payment_by_lang(payment_terms, lang) if payment_terms else _td('T/T 30% deposit + 70% before shipment', lang)

    ws.cell(row, 6).value = pt_value; ws.cell(row, 6).font = FONT_NORMAL; ws.merge_cells(f'F{row}:H{row}'); row += 1

    ws.cell(row, 1).value = _td('Port of Discharge:', lang); ws.cell(row, 1).font = FONT_BOLD

    ws.cell(row, 2).value = port_discharge or 'XXXXX'; ws.cell(row, 2).font = FONT_NORMAL

    ws.cell(row, 5).value = _td('Incoterms:', lang); ws.cell(row, 5).font = FONT_BOLD

    ws.cell(row, 6).value = trade_terms or _td('FOB XXXXX', lang); ws.cell(row, 6).font = FONT_NORMAL; ws.merge_cells(f'F{row}:H{row}'); row += 2

    ws.cell(row, 1).value = _td('4. Marks & No.:', lang); ws.cell(row, 1).font = FONT_SECTION; row += 1

    ws.cell(row, 1).value = _td('N/M', lang); ws.cell(row, 1).font = FONT_NORMAL; row += 2

    cur_sym = '$' if currency == 'USD' else ('¥' if currency in ('CNY', 'RMB', '') else currency)

    headers = [_td(h, lang) for h in ['Item No.', 'Description of Goods', 'Specification', 'Quantity', 'Unit', f'Unit Price ({cur_sym})', f'Total Amount ({cur_sym})']]

    widths = [10, 30, 30, 12, 8, 16, 18]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):

        ws.cell(row, ci).value = h; ws.cell(row, ci).font = FONT_TABLE_HEADER; ws.cell(row, ci).fill = FILL_BLUE

        ws.cell(row, ci).alignment = ALIGN_CENTER; ws.cell(row, ci).border = BORDER

        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[row].height = 30; row += 1

    total_amount=0; total_qty=0
    # 商业发票：CNY→USD 换算（同报价单/PI）
    _rate = None
    if currency == 'USD':
        try:
            from src.rates import get_rate as _get_rate
            _rate = _get_rate('CNY', 'USD') or 0.14
        except Exception:
            _rate = 0.14

    for i, item in enumerate(pi_items, 1):

        qty=int(item.get('qty',item.get('quantity',1)))

        price=float(item.get('price_rmb',item.get('unit_price',0)))
        if _rate:
            price = round(price * _rate, 2)

        total=qty*price; total_qty+=qty; total_amount+=total

        model=item.get('model',''); name=item.get('name_zh',''); spec=item.get('spec_zh','')

        vals=[i,model or name,spec,qty,_td('pcs', lang),price,total]

        for ci,v in enumerate(vals,1):

            ws.cell(row,ci).value=v; ws.cell(row,ci).font=FONT_NORMAL; ws.cell(row,ci).border=BORDER

            ws.cell(row,ci).alignment=ALIGN_CENTER if ci<=2 else ALIGN_RIGHT

        if i%2==0:

            for ci in range(1,8): ws.cell(row,ci).fill=FILL_LIGHT

        row+=1

    ws.cell(row,6).value=_td('Subtotal:', lang); ws.cell(row,6).font=FONT_BOLD; ws.cell(row,6).alignment=ALIGN_RIGHT; ws.cell(row,6).border=BORDER

    ws.cell(row,7).value=total_amount; ws.cell(row,7).font=FONT_BOLD; ws.cell(row,7).border=BORDER; ws.cell(row,7).fill=FILL_TOTAL; row+=1

    ws.cell(row,1).value=_td('6. Freight & Charges:', lang); ws.cell(row,1).font=FONT_SECTION; ws.merge_cells(f'A{row}:F{row}'); row+=1

    ws.cell(row,1).value=_td('Freight:', lang); ws.cell(row,1).font=FONT_NORMAL; ws.cell(row,4).value=_td('Insurance:', lang); ws.cell(row,4).font=FONT_NORMAL; row+=1

    ws.cell(row,1).value=_td('Handling:', lang); ws.cell(row,1).font=FONT_NORMAL; ws.cell(row,4).value=_td('Others:', lang); ws.cell(row,4).font=FONT_NORMAL; row+=2

    ws.cell(row,1).value=_td(f'7. Total Amount ({trade_terms}):', lang); ws.cell(row,1).font=FONT_SECTION; ws.merge_cells(f'A{row}:F{row}')

    ws.cell(row,7).value=total_amount; ws.cell(row,7).font=Font(name='Arial',size=11,bold=True); ws.cell(row,7).border=BORDER; ws.cell(row,7).fill=FILL_TOTAL; row+=2

    try:

        from src.output.pi_generator import _num_to_words

        words=_td(f'8. Total Amount in Words: {_num_to_words(total_amount,"USD")}', lang)

    except Exception:

        words=_td('8. Total Amount in Words: USD XXXXX ONLY.', lang)

    ws.cell(row,1).value=words; ws.cell(row,1).font=FONT_SECTION; ws.merge_cells(f'A{row}:G{row}'); row+=2

    ws.cell(row,1).value=_td('9. Country of Origin:', lang); ws.cell(row,1).font=FONT_SECTION

    ws.cell(row,3).value=origin_country or 'XXXXX'; ws.cell(row,3).font=FONT_NORMAL

    ws.cell(row,5).value=_td('HS Code:', lang); ws.cell(row,5).font=FONT_SECTION

    ws.cell(row,6).value='XXXXX'; ws.cell(row,6).font=FONT_NORMAL; row+=2

    bank=company_config.get('bank',{})

    ws.cell(row,1).value=_td('10. Bank Information:', lang); ws.cell(row,1).font=FONT_SECTION; ws.merge_cells(f'A{row}:G{row}'); row+=1

    for label,key in [('Beneficiary:','beneficiary'),('Bank name:','bank_name'),('Bank add.:','bank_address'),('Account No.:','account_no'),('Swift Code:','swift_code')]:

        val=bank.get(key,'') or 'XXXXX'

        ws.cell(row,1).value=_td(f'   {label}  {val}', lang); ws.cell(row,1).font=FONT_NORMAL; ws.merge_cells(f'A{row}:G{row}'); row+=1

    row+=1

    ws.cell(row,1).value=_td('11. Remarks: All disputes subject to jurisdiction of China.', lang); ws.cell(row,1).font=FONT_SECTION; ws.merge_cells(f'A{row}:G{row}'); row+=2

    ws.cell(row,1).value=_td('12. Signature:', lang); ws.cell(row,1).font=FONT_SECTION; row+=1

    ws.cell(row,1).value=seller_name; ws.cell(row,1).font=FONT_BOLD; row+=1

    ws.cell(row,1).value=_td('(signed & stamped)', lang); ws.cell(row,1).font=Font(name='Arial',size=8,color='888888')

    wb.save(output_path)

    logging.info(f"[INVOICE] successfully generated: {output_path}")

    return output_path





def create_packing_and_invoice(

    pi_items, invoice_no, invoice_date, buyer_name="", buyer_address="", output_dir="output",

):

    os.makedirs(output_dir, exist_ok=True)

    packing_path = os.path.join(output_dir, f"PackingList_{invoice_no}.xlsx")

    invoice_path = os.path.join(output_dir, f"CommercialInvoice_{invoice_no}.xlsx")

    generate_packing_list(pi_items, invoice_no, invoice_date, buyer_name, buyer_address, output_path=packing_path)

    generate_commercial_invoice(pi_items, invoice_no, invoice_date, buyer_name, buyer_address, output_path=invoice_path)

    return {'packing_list': packing_path, 'commercial_invoice': invoice_path}



