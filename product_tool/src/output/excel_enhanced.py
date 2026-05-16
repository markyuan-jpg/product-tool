# -*- coding: utf-8 -*-
"""
Excel enhanced - freeze/filter/template
"""
import os
import logging
from typing import List, Dict, Optional

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    raise ImportError("openpyxl needed")

logging.basicConfig(level=logging.INFO)


def set_freeze_panes(ws, cell: str = 'B2'):
    """冻结窗格"""
    ws.freeze_panes = cell


def set_autofilter(ws, range: str = None):
    """"""
    if range:
        ws.auto_filter.ref = range
    else:
        ws.auto_filter.ref = ws.dimensions


def apply_conditional_format(ws, range: str, formula: str, fill: str = 'FFEB9C'):
    """条件格式"""
    rule = FormulaRule(formula=[formula], fill=PatternFill(start_color=fill, fill_type='solid'))
    ws.conditional_formatting.add(range, rule)


def protect_sheet(ws, password: str = None):
    """"""
    ws.protection.sheet = True
    if password:
        ws.protection.password = password


def add_footer(ws, left: str = None, center: str = None, right: str = None):
    """添加页脚"""
    if left:
        ws.oddFooter.left.text = left
    if center:
        ws.oddFooter.center.text = center
    if right:
        ws.oddFooter.right.text = right


def set_column_width(ws, col: str, width: float):
    """设置列宽"""
    ws.column_dimensions[col].width = width


def hide_column(ws, col: str):
    """"""
    ws.column_dimensions[col].hidden = True


def add_data_validation(ws, range: str, formula: str, title: str = "List"):
    """数据验证(下拉菜单)"""
    from openpyxl.worksheet.datavalidation import DataValidation
    
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(range)
    ws.add_data_validation(dv)


def apply_template(input_path: str, output_path: str, template: str = None) -> str:
    """"""
    wb = load_workbook(input_path)
    ws = wb.active
    
    # 冻结首行
    ws.freeze_panes = 'B2'
    
    # 自动筛???    ws.auto_filter.ref = ws.dimensions
    
    # 自动列宽
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value or '')) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        width = min(max(max_len + 2, 10), 50)
        ws.column_dimensions[col_letter].width = width
    
    wb.save(output_path)
    return output_path


def merge_excel_files(files: List[str], output_path: str) -> str:
    """合并多个Excel文件"""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    row = 1
    for file in files:
        if not os.path.exists(file):
            continue
        
        wb2 = load_workbook(file)
        ws2 = wb2.active
        
        for r in ws2.iter_rows():
            for cell in r:
                new_cell = ws.cell(row=row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.alignment = cell.alignment.copy()
            row += 1
    
    wb.save(output_path)
    return output_path


def export_to_template(data: List[Dict], output_path: str, template_path: str = None) -> str:
    """"""
    if template_path and os.path.exists(template_path):
        wb = load_workbook(template_path)
    else:
        wb = Workbook()
    
    ws = wb.active
    
    # 写入表头
    headers = list(data[0].keys()) if data else []
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    
    # 写入数据
    for row_idx, item in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, item.get(header, ''))
    
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    print('Excel Enhanced Functions')
