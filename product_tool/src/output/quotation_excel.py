# -*- coding: utf-8 -*-

"""

Quotation Excel generator

"""

import os

import re

import glob as g

import logging

import warnings

import pandas as pd

from datetime import datetime, timedelta

from typing import List, Dict, Optional

from openpyxl import Workbook

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from openpyxl.utils import get_column_letter



try:

    from src.utils.translator import translate_text, bilingual_text, translate_spec_safe

except ImportError:

    from utils.translator import translate_text, bilingual_text, translate_spec_safe



try:

    from src.parsers.spec_cleaner import clean_spec

except ImportError:

    def clean_spec(x): return x or ''



from .doc_shared import payment_by_lang, filter_by_lang, has_chinese, is_already_bilingual





# ====================  ====================



HEADER_FONT = Font(name='Arial', size=14, bold=True, color='000000')

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

HEADER_FONT_WHITE = Font(name='Arial', size=12, bold=True, color='FFFFFF')



SUBHEADER_FONT = Font(name='Arial', size=11, bold=True, color='000000')

SUBHEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')



TABLE_HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')

TABLE_HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')



DATA_FONT = Font(name='Arial', size=10, color='000000')

DATA_ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

DATA_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

DATA_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')



EVEN_ROW_FILL = PatternFill(start_color='E7E3E6', end_color='E7E3E6', fill_type='solid')

ODD_ROW_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')



THIN = Side(style='thin', color='000000')

BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)



PRICE_FONT = Font(name='Arial', size=10, bold=False, color='000000')

TOTAL_FONT = Font(name='Arial', size=11, bold=True, color='000000')





# ==================== Quotation Excel ====================



class QuotationExcel:


    

    def __init__(

        self,

        supplier: str = '',

        quotation_no: str = '',

        valid_days: int = 30,

        trade_terms: str = 'FOB Qingdao',

        payment_terms: str = '30% deposit by T/T, 70% balance before shipment within 60 days upon receipt of payment.',

        currency: str = 'CNY',

        lang: str = 'chinese',

        company_info: dict = None,

    ):

        self.supplier = supplier

        self.company_info = company_info or {}

        self.quotation_no = quotation_no or f"Q-{datetime.now().strftime('%Y%m%d')}"

        self.valid_days = valid_days

        self.trade_terms = trade_terms

        self.payment_terms = payment_terms

        self.date = datetime.now().strftime('%Y-%m-%d')

        self.currency = currency.upper()

        self.lang = lang

    

    def _currency_symbol(self) -> str:
        """Method"""
        """Method"""
        return '$' if self.currency in ('USD', 'US') else ''

    

    def _translate_doc(self, text: str) -> str:
        """Method"""
        """Method"""
        try:

            from src.utils.translator import translate_doc

            return translate_doc(text, self.lang)

        except Exception:

            return text

    

    @staticmethod

    def _format_model_name(model: str, name: str) -> str:
        """Method"""
        """Method"""
        model = str(model).strip() if model else ''

        name = str(name).strip() if name else ''

        if not model and not name:

            return ''

        if not name or name.lower() in ('nan', 'none', ''):

            return model

        if model == name:

            return model

        return f"{model}\n{name}"

    

    def add_products(self, df: pd.DataFrame) -> pd.DataFrame:
        """Method"""
        """Method"""
        result = []

        

        # 批处理翻译（收集所有 name/spec 一次翻译，避免逐条调用）
        if self.lang == 'bilingual':
            from src.utils.translator import batch_translate
            _names_batch = [str(r['name_zh']) for _, r in df.iterrows() if pd.notna(r.get('name_zh')) and str(r['name_zh']).strip().lower() not in ['', 'nan', 'none'] and not r.get('name_en', '')]
            _specs_batch = [str(r['spec_zh']) for _, r in df.iterrows() if pd.notna(r.get('spec_zh')) and str(r['spec_zh']).strip().lower() not in ['', 'nan', 'none']]
            _name_xlat = batch_translate(_names_batch, 'zh_en')
            _spec_xlat = batch_translate(_specs_batch, 'zh_en')
        else:
            _name_xlat = _spec_xlat = {}

        for idx, (_, row) in enumerate(df.iterrows()):

            # model 

            model = row.get('model', '')

            if pd.isna(model) or str(model).strip() in ['', 'nan']:

                continue

            

            # Clean model - remove any parameter content that got mixed in

            model_str = str(model).strip()

            # Split on any newline to remove leaked params

            model_str = model_str.split('\n')[0]

            

            # Split on ; delimiter too - sometimes params are after ;

            model_str = model_str.split(';')[0]

            

            # Also check for early colon pattern (params leak)

            colon_pos = model_str.find(':')

            if colon_pos > 0 and colon_pos < 20:

                model_str = model_str[:colon_pos].strip()

            

            # Use cleaned model

            model = model_str

            

            # name_zh possible empty or NaN, fallback to model

            name_raw = row.get('name_zh')

            if pd.isna(name_raw) or str(name_raw).strip().lower() in ['', 'nan', 'none']:

                name = model_str  # Use cleaned model as name fallback

            else:

                name = str(name_raw).strip()

            

            # 

            category = str(row.get('category', '')).strip()

            if category == '':

                name = f"[] {name}"

            

            #  lang / name

            if self.lang == 'bilingual':

                name_en = row.get('name_en', '')

                if name_en and str(name_en).strip():

                    ne = str(name_en).strip()

                    has_cjk = any('\u4e00' <= c <= '\u9fff' for c in ne)

                    has_suspicious_ascii = any(c.isascii() and c.isalpha() for c in ne) and has_cjk

                    if not has_suspicious_ascii:

                        name = bilingual_text(name, ne)

                else:

                    translated = _name_xlat.get(name, name) if _name_xlat else translate_spec_safe(name)

                    if translated and translated != name:

                        name = bilingual_text(name, translated)

            else:

                name = filter_by_lang(name, self.lang)

            

            # spec 

            spec_raw = row.get('spec_zh', '')

            if pd.isna(spec_raw) or str(spec_raw).strip().lower() in ['', 'nan', 'none']:

                spec = ''

            else:

                spec = clean_spec(spec_raw)

                if self.lang == 'bilingual':

                    spec_en = _spec_xlat.get(spec, spec) if _spec_xlat else translate_spec_safe(spec)

                    if spec_en and spec_en != spec:

                        spec = bilingual_text(spec, spec_en)

                else:

                    spec = filter_by_lang(spec, self.lang)

            

            qty = row.get('qty', row.get('quantity', ''))

            if qty is None or (isinstance(qty, str) and qty.strip().lower() in ['', 'nan', 'none']):

                qty = ''

            else:

                try:

                    qty = int(float(str(qty)))

                except (ValueError, TypeError):

                    qty = ''

            

            # 

            unit_price_raw = row.get('unit_price') or row.get('price_rmb')

            if pd.isna(unit_price_raw) or unit_price_raw is None:

                unit_price = 0

            else:

                try:

                    unit_price = float(unit_price_raw)

                    # USD CNY auto conversion

                    if self.currency == 'USD' and unit_price > 0:

                        try:

                            from src.rates import get_rate

                            rate = get_rate('CNY', 'USD') or 0.14

                            unit_price = round(unit_price * rate, 2)

                        except Exception:

                            pass

                except (ValueError, TypeError):

                    unit_price = 0

            

            result.append({

                '_seq': idx + 1,

                'model': model,

                'name_zh': name,

                'spec_zh': spec,

                'qty': qty,

                'unit_price': unit_price,

                'total': unit_price * qty if unit_price and qty else 0,

                '_image_path': row.get('_image_path') or row.get('image_path', '') or '',

                'remark': row.get('remark', ''),

                'category': row.get('category', ''),

            })

        

        return pd.DataFrame(result)

    

    def write(

        self,

        data: List[Dict],

        output_path: str,

        image_dir: str = None,

        with_images: bool = False,

    ) -> str:

        """Create quotation output"""
        df = pd.DataFrame(data)

        

        # Clean NaN
        if 'model' in df.columns:

            df = df[df['model'].notna() & (df['model'] != '') & (df['model'] != 'nan')]

        

        if df.empty:

            # 

            wb = Workbook()

            wb.active.title = 'Quotation'

            wb.save(output_path)

            return output_path

        

        embed_images = with_images

        

        # : add_products()???NaN

        df = self.add_products(df)

        

        valid_until = (datetime.now() + timedelta(days=self.valid_days)).strftime('%Y-%m-%d')

        

        # workbook

        wb = Workbook()

        

        ws = wb.active

        ws.title = 'Quotation'

        

        row = 1

        

        # -----  -----

        ws.merge_cells(f'A{row}:G{row}')

        cell = ws[f'A{row}']

        cell.value = self._translate_doc('FOREIGN TRADE QUOTATION')

        cell.font = HEADER_FONT

        cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[row].height = 30

        row += 1

        

        # ----- Company info block (after title, before info row) -----

        company = self.company_info or {}

        company_name = company.get('name_en', '') or company.get('name', '') or self.supplier or 'XXXXX'

        

        if company_name:

            ws.merge_cells(f'A{row}:G{row}')

            cell = ws[f'A{row}']

            cell.value = company_name

            cell.font = Font(name='Arial', size=11, bold=True, color='1a5fb4')

            cell.alignment = Alignment(horizontal='left', vertical='center')

            ws.row_dimensions[row].height = 22

            row += 1

        

        company_addr = company.get('address', '') or 'XXXXX'

        company_tel = company.get('tel', '') or 'XXXXX'

        company_email = company.get('email', '') or 'XXXXX'

        if company_addr or company_tel or company_email:

            ws.merge_cells(f'A{row}:G{row}')

            parts = []

            if company_addr:

                parts.append(company_addr)

            contact_parts = []

            if company_tel:

                contact_parts.append(f"{self._translate_doc('Tel')}: {company_tel}")

            if company_email:

                contact_parts.append(f"{self._translate_doc('Email')}: {company_email}")

            if contact_parts:

                parts.append(' | '.join(contact_parts))

            cell = ws[f'A{row}']

            cell.value = ' | '.join(parts)

            cell.font = Font(name='Arial', size=9, color='555555')

            cell.alignment = Alignment(horizontal='left', vertical='center')

            ws.row_dimensions[row].height = 18

            row += 1

        

        # ----- ???-----

        info_labels = [self._translate_doc('Quotation No.'), self._translate_doc('Date'), self._translate_doc('Valid Until')]

        info_values = [

            self.quotation_no,

            self.date,

            valid_until,

        ]

        

        for i in range(len(info_labels)):

            ws.cell(row, i*2+1, info_labels[i])

            ws.cell(row, i*2+2, info_values[i])

            ws.cell(row, i*2+1).font = Font(bold=True, size=9)

            ws.cell(row, i*2+2).font = Font(size=9)

        row += 1

        

            # ----- ???-----

        cur_sym = '$' if self.currency in ('USD', 'US') else ('¥' if self.currency in ('CNY', 'RMB', '') else self.currency)

        headers = [

            self._translate_doc('No.'),

            self._translate_doc('Photo'),

            self._translate_doc('Model / Product Name'),

            self._translate_doc('Specifications'),

            self._translate_doc('Qty'),

            f"{self._translate_doc('Unit Price')} ({cur_sym})",

            f"{self._translate_doc('Total')} ({cur_sym})",

        ]

        for col_idx, header in enumerate(headers, 1):

            cell = ws.cell(row, col_idx)

            cell.value = header

            cell.font = TABLE_HEADER_FONT

            cell.fill = TABLE_HEADER_FILL

            cell.alignment = DATA_ALIGN_CENTER

            cell.border = BORDER

        row += 1

        

        # ----- ???-----

        data_start_row = row

        total_amount = 0

        for idx, (_, record) in enumerate(df.iterrows()):

            model = record.get('model', '')



            # Clean model - same logic as add_products

            if model:

                model = str(model).strip()

                if '\n' in model:

                    model = model.split('\n')[0]

                if ';' in model:

                    model = model.split(';')[0]

                colon_pos = model.find(':')

                if colon_pos > 0 and colon_pos < 20:

                    model = model[:colon_pos].strip()

            

            # Skip empty model

            if not model:

                continue

            

            is_even = (idx % 2 == 0)

            row_fill = EVEN_ROW_FILL if is_even else ODD_ROW_FILL

            

            qty_val = record.get('qty', 0)

            if qty_val == '' or qty_val is None:

                qty_val = 0

            else:

                try:

                    qty_val = int(qty_val)

                except (ValueError, TypeError):

                    qty_val = 0

            

            price_val = record.get('unit_price', 0) or 0

            total_amount += qty_val * price_val

            

            row_data = [

                idx + 1,

                '',

                self._format_model_name(record.get('model', ''), record.get('name_zh', '')),

                record.get('spec_zh', ''),

                record.get('qty', 0),

                record.get('unit_price', 0),

                record.get('total', 0),

            ]

            

            img_path = record.get('_image_path') or record.get('image_path', '')

            has_img = bool(img_path)

            if has_img and embed_images and os.path.exists(img_path):

                try:

                    from openpyxl.drawing.image import Image as XLImage

                    from src.core.image import resize_image

                    IMG_W, IMG_H = 100, 80

                    _resized = resize_image(img_path)

                    img = XLImage(_resized if hasattr(_resized, 'read') else img_path)

                    img.width = IMG_W; img.height = IMG_H

                    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, 90)

                    ws.add_image(img, f'B{row}')

                except Exception:

                    pass

            row_data[1] = self._translate_doc('Photo') if has_img and not embed_images else ''

            

            for col_idx, value in enumerate(row_data, 1):

                cell = ws.cell(row, col_idx)

                

                if col_idx == 1:  # 

                    cell.value = value

                    cell.alignment = DATA_ALIGN_CENTER

                elif col_idx == 2:  # Photo (???

                    cell.alignment = DATA_ALIGN_CENTER

                elif col_idx == 3:  #  +  ()

                    cell.value = value

                    cell.alignment = DATA_ALIGN_LEFT

                elif col_idx == 4:  #  - 

                    cell.value = value

                    cell.alignment = DATA_ALIGN_LEFT

                    if value and len(str(value)) > 0:

                        text = str(value)

                        explicit_lines = text.count('\n') + 1

                        # ??????5)

                        chars_per_line = 45

                        wrapped_lines = sum(max(1, len(line) // chars_per_line + 1) for line in text.split('\n'))

                        total_lines = max(explicit_lines, wrapped_lines)

                        est_height = max(ws.row_dimensions[row].height or 15, total_lines * 14)

                        ws.row_dimensions[row].height = est_height

                elif col_idx == 5:  # 

                    cell.value = value

                    cell.alignment = DATA_ALIGN_CENTER

                elif col_idx == 6:  #  ???

                    if value and value > 0:

                        if value == int(value):

                            cell.value = int(value)

                        else:

                            cell.value = value

                        cell.number_format = '#,##0.##'

                    else:

                        cell.value = '-'

                    cell.alignment = DATA_ALIGN_RIGHT

                    cell.font = PRICE_FONT

                elif col_idx == 7:  # Total

                    if value and value > 0:

                        cell.value = value

                    else:

                        cell.value = 0

                    cell.alignment = DATA_ALIGN_RIGHT

                    cell.font = PRICE_FONT

                    cell.number_format = '#,##0.##'

                

                cell.fill = row_fill

                cell.border = BORDER

                if col_idx not in [6, 7]:

                    cell.font = DATA_FONT

            

            #  attached remark???

            remark = record.get('remark', '')

            if remark:

                row += 1

                ws.merge_cells(f'A{row}:G{row}')

                cell = ws.cell(row, 1)

                cell.value = f"???{remark}"

                cell.font = Font(name='Arial', size=8, color='888888', italic=True)

                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                ws.row_dimensions[row].height = max(20, len(str(remark)) // 2)

                # Apply light fill to distinguish

                for cc in range(1, 8):

                    ws.cell(row, cc).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

            

            row += 1

        

        # ----- Total???-----

        ws.merge_cells(f'A{row}:F{row}')

        cell = ws.cell(row, 1)

        cell.value = self._translate_doc('TOTAL:')

        cell.font = TOTAL_FONT

        cell.alignment = Alignment(horizontal='right', vertical='center')

        cell.fill = SUBHEADER_FILL

        cell.border = BORDER

        

        # Apply fill/border to merged cells too

        for c in range(2, 7):

            ws.cell(row, c).fill = SUBHEADER_FILL

            ws.cell(row, c).border = BORDER

        

            cell = ws.cell(row, 7)

            cell.value = total_amount if total_amount > 0 else 0

            cell.font = TOTAL_FONT

            cell.alignment = DATA_ALIGN_RIGHT

            cell.fill = SUBHEADER_FILL

            cell.border = BORDER

            cell.number_format = '#,##0.00'

        row += 1

        

        # -----  -----

        row += 1

        

        # -----  -----

        # Enhance trade terms with more details

        enhanced_terms = self.trade_terms

        if 'FOB' in enhanced_terms.upper() and 'qingdao' not in enhanced_terms.lower():

            enhanced_terms = 'FOB Qingdao, China'

        elif 'CIF' in enhanced_terms.upper() and 'qingdao' not in enhanced_terms.lower():

            enhanced_terms = 'CIF [Destination Port], China'

            

        terms = [

            f"{self._translate_doc('Trade Terms')}: {enhanced_terms}",

            f"{self._translate_doc('Payment Terms')}: {payment_by_lang(self.payment_terms, self.lang)}",

            f"{self._translate_doc('Packing')}: {self._translate_doc('Standard export packing')}",

            f"{self._translate_doc('Delivery')}: {self._translate_doc('15-25 days after deposit')}",

            f"{self._translate_doc('Validity')}: {self._translate_doc('Please confirm within validity period')}"

        ]

        

        for term in terms:

            ws.merge_cells(f'A{row}:G{row}')

            cell = ws.cell(row, 1)

            cell.value = term

            cell.font = Font(size=9, color='333333')

            cell.alignment = Alignment(horizontal='left', vertical='center')

            row += 1

        

        # -----  -----

        # Calculate max width for each column based on content

        widths = [6, 14, 30, 45, 8, 14, 14]  # No., Photo(14=100px), Model/Name, Specs, Qty, Unit Price, Total

        

        # Check header widths

        headers = ['No.', 'Photo', 'Model / Product Name', 'Specifications', 'Qty', 'Unit Price', 'Total']

        for i, h in enumerate(headers):

            widths[i] = max(widths[i], len(h) + 2)

        

        # Check data row widths (sample first 50 rows)

        for idx, (_, record) in enumerate(df.head(50).iterrows()):

            widths[0] = max(widths[0], len(str(idx+1)) + 2)  # No.

            widths[1] = max(widths[1], 12)  # Photo (fixed width)

            model_name = record.get('name_zh', '') or record.get('model', '')

            widths[2] = max(widths[2], len(str(model_name)) + 2)  # Model

            specs = record.get('spec_zh', '')

            widths[3] = max(widths[3], min(len(str(specs)) // 3 + 20, 80))  # Specs - allow wider

            widths[4] = max(widths[4], len(str(record.get('qty', ''))) + 2)  # Qty

            widths[5] = max(widths[5], len(str(record.get('unit_price', ''))) + 2)  # Price

            widths[6] = max(widths[6], len(str(record.get('total', ''))) + 2)  # Total

        

        # Apply column widths (min 10, max 80)

        final_widths = [max(8, min(w, 80)) for w in widths]

        for col_idx, w in enumerate(final_widths, 1):

            ws.column_dimensions[get_column_letter(col_idx)].width = w

        

        # ----- ??????0pt ???-----

        for r in range(data_start_row, row):

            if ws.row_dimensions[r].height and ws.row_dimensions[r].height >= 80:

                continue

            cell_val = ws.cell(r, 4).value  # Specifications column

            if cell_val:

                # Calculate needed height based on content length and column width

                text_len = len(str(cell_val))

                col_width = final_widths[3]

                # Estimate lines needed (assuming ~8 chars per cell width)

                lines = max(1, text_len // (col_width * 0.5))

                ws.row_dimensions[r].height = min(20 + lines * 15, 200)  # Max 200 height

        

        # 

        ws.freeze_panes = 'A2'

        

        # Sheet 2: 

        ws2 = wb.create_sheet(title='Raw Data')

        

        # 

        raw_headers = list(df.columns)

        for col_idx, header in enumerate(raw_headers, 1):

            cell = ws2.cell(1, col_idx)

            cell.value = header

            cell.font = Font(bold=True)

            cell.fill = SUBHEADER_FILL

        

        for row_idx, (_, record) in enumerate(df.iterrows(), 2):

            for col_idx, header in enumerate(raw_headers, 1):

                cell = ws2.cell(row_idx, col_idx)

                cell.value = record.get(header, '')

        

        # 

        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

        wb.save(output_path)

        

        return output_path





def create_quotation(

    data: List[Dict],

    output_path: str,

    supplier: str = '',

    image_dir: str = None,

    **kwargs

) -> str:

    """Generate quotation Excel file"""
    if image_dir and data:

        for item in data:

            img = item.get('_image_path') or item.get('image_path', '')

            if img and not os.path.isabs(img):

                item['_image_path'] = os.path.join(image_dir, img)

    

    with_images = kwargs.pop('with_images', False)

    

    # 

    try:

        from .excel_template import get_template_path, apply_template

        tmpl = get_template_path('quotation')

        if tmpl and apply_template(data, tmpl, output_path):

            return output_path

    except Exception:

        pass

    qt = QuotationExcel(supplier=supplier, **kwargs)

    return qt.write(data, output_path, with_images=with_images)





def df_to_quotation(

    df: pd.DataFrame,

    output_path: str,

    supplier: str = '',

    **kwargs

) -> str:


    data = df.to_dict('records')

    return create_quotation(data, output_path, supplier, **kwargs)





def create_quotation_from_library(

    # product_ids: List[int],  # orphaned

    # quantities: List[int],  # orphaned

    # output_path: str,  # orphaned

    # currency: str = 'RMB',  # orphaned

    # company_config: dict = None,  # orphaned

    # trade_terms: str = 'FOB Qingdao',  # orphaned

    # payment_terms: str = '30% deposit by T/T, 70% balance before shipment within 60 days upon receipt of payment.',  # orphaned

    # user_id: str = 'local',  # orphaned

    # use_term_calculation: bool = False,  # orphaned

    # destination_country: str = "",  # orphaned

    # volume_cbm: float = 0,  # orphaned

    # image_search_dirs: list = None,  # orphaned

    # with_images: bool = False,  # orphaned

    # lang: str = 'chinese',  # orphaned

    # industry: str = None,  # orphaned

    # include_optional: bool = True,  # orphaned

) -> str:

    """
# Args:  # orphaned
    """

        # product_ids: ID  # orphaned
        # quantities:  # orphaned

        # output_path:  # orphaned

        # currency:  (RMB/USD)  # orphaned

        # company_config:  (None = )  # orphaned

        # trade_terms:  (EXW/FOB/CFR/CIF/DAP/DDP)  # orphaned

        # payment_terms:  # orphaned

        # user_id: ID  # orphaned

        # use_term_calculation:  # orphaned

        # destination_country: ???()  # orphaned

        # volume_cbm: ???CBM  # orphaned

        

    # Returns:  # orphaned

    # Import from product_manage

    try:

        from src.product_manage.repository import get_products_by_ids

        from src.product_manage.db import init_db

    except ImportError:

        from product_manage.repository import get_products_by_ids

        from product_manage.db import init_db

    

    #  - ???.2

    rate_usd_to_cny = 7.2  # ???    

    try:

        from src.output.pricing import TieredPricing, validate_moq

        use_pricing = True

    except ImportError:

        use_pricing = False

    

    # Trade terms (???

    use_terms = False

    use_rates = False

    term_info = {}

    if use_term_calculation:

        try:

            from src.rates import convert as rate_convert

            from src.terms import calculate_price, get_term_info

            use_rates = True

            use_terms = True

            term_info = get_term_info(trade_terms.split()[0]) if trade_terms else {}

        except ImportError:

            pass

    

    # Initialize DB

    init_db()

    

    # Load company config

    if company_config is None:

        try:

            from src.company import load_company

            company_config = load_company()

        except ImportError:

            from company import load_company

            company_config = load_company()



    # Build data from products

    data = []

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    products_by_id = get_products_by_ids(product_ids, user_id)

    

    # Only build image cache if any product needs image resolution

    image_cache = {}

    if image_search_dirs and products_by_id and any(not p.image_path for p in products_by_id.values()):

        for search_dir in image_search_dirs:

            if os.path.isdir(search_dir):

                for ext in ['*.jpg', '*.jpeg', '*.png']:

                    for fpath in g.glob(os.path.join(search_dir, '**', ext), recursive=True):

                        fname = os.path.splitext(os.path.basename(fpath).lower())[0]

                        image_cache[fname] = fpath



    for pid, qty in zip(product_ids, quantities):

        product = products_by_id.get(pid)

        if not product:

            continue

        

        # ???        base_rmb = product.price_rmb or 0

        if include_optional and product.prices:

            # DB ???prices  JSON ???            prices_dict = product.prices

            if isinstance(prices_dict, str):

                try:

                    prices_dict = json.loads(prices_dict)

                except (json.JSONDecodeError, TypeError):

                    prices_dict = {}

            if not isinstance(prices_dict, dict):

                prices_dict = {}

            for vals in prices_dict.values():

                for val, currency in vals:

                    if currency == 'CNY' or currency == 'RMB':

                        base_rmb += val

                    else:

                        try:

                            from src.rates import convert as rate_convert

                            base_rmb += rate_convert(val, currency, 'CNY')

                        except Exception:

                            from src.rates import get_rate

                            rate = get_rate(currency, 'CNY')

                            base_rmb += val * rate if rate else val

        

        # 1)  (FOB/CIF/DAP/DDP)

        if use_terms:

            calc = calculate_price(

                base_price=base_rmb,

                quantity=qty,

                term=trade_terms.split()[0],

                currency='CNY',  # Always use CNY for calculation, convert later if needed

                destination_country=destination_country,

                volume_cbm=volume_cbm / len(product_ids) if volume_cbm else 0.5,

            )

            if trade_terms.startswith('FOB'):

                price = calc.fob_price / qty

            elif trade_terms.startswith('CIF'):

                price = calc.cif_price / qty

            elif trade_terms.startswith('DAP'):

                price = calc.dap_price / qty

            elif trade_terms.startswith('DDP'):

                price = calc.ddp_price / qty

            else:

                price = base_rmb

            # ???USD???            if currency.upper() == 'USD':

                if use_rates:

                    price = rate_convert(price, 'CNY', 'USD')

                else:

                    price = price / rate_usd_to_cny

        else:

            # 2) ???            price = base_rmb

            if currency.upper() == 'USD':

                if product.price_usd:

                    price = product.price_usd

                elif use_rates:

                    price = rate_convert(base_rmb, 'CNY', 'USD')

                else:

                    price = base_rmb / rate_usd_to_cny

        

        # 3) MOQ

        if use_pricing and price > 0 and product.moq > 1 and qty < product.moq:

            qty = product.moq  # MOQ

        

        # FIX: Clean SKU - remove any parameter content after the actual SKU name

        # Strategy: SKU should NOT contain any Chinese characters followed by : or numbers after :

        sku_raw = str(product.sku).strip() if product.sku else ''

        

        # Split on newline and take first part, then remove any trailing param pattern

        sku_parts = sku_raw.split('\n')

        sku_clean = sku_parts[0]

        

        # If first part contains patterns like "??? 2000" or similar, extract just the SKU part

        # Pattern: take text before any "X: number" pattern appears

        param_pattern = re.compile(r'[\u4e00-\u9fa5]+:\s*[\d]+')  # Chinese:text + colon + digits

        match = param_pattern.search(sku_clean)

        if match:

            sku_clean = sku_clean[:match.start()].strip()

        

        # Final cleanup - remove any trailing punctuation

        sku_clean = sku_clean.rstrip(';:???')

        

        # Build spec string: prefer more complete source (spec_zh vs specs dict)

        def _build_spec_from_dict(s):
            """Method"""
            """Method"""
            p = []

            for k, v in s.items():

                vc = str(v).replace('\n', ' ').strip()

                p.append(f"{k}: {vc}")

            return "\n".join(p)



        spec_from_specs = _build_spec_from_dict(product.specs) if product.specs else ""

        spec_zh_len = len(product.spec_zh or "")

        specs_len = len(spec_from_specs)



        if product.spec_zh and spec_zh_len > max(specs_len, 50):

            spec_str = product.spec_zh

        elif product.specs:

            spec_str = spec_from_specs

        else:

            spec_str = ""

        

        spec_str = re.sub(r'[:]\s*[:]', ':', spec_str)

        

        # Append secondary prices to spec_zh

        if product.prices:

            from ..price_config import get_industry_config, get_secondary_labels

            ic = get_industry_config(industry)

            sl = get_secondary_labels(ic)

            lines = []

            for ptype, vals in product.prices.items():

                label = sl.get(ptype, ptype.capitalize() + ' Price')

                items = [f"{cur} {v:,.2f}" for v, cur in vals]

                lines.append(f"{label}: {', '.join(items)}")

            if lines:

                spec_str += '\n\n[Optional Accessories]\n' + '\n'.join(lines)

        

        # 

        stored_img = product.image_path

        resolved_img = ''



        if stored_img:

            if os.path.isfile(stored_img):

                resolved_img = stored_img

            elif os.path.isabs(stored_img):

                logging.warning(f"Image path not found for {product.sku}: {stored_img}")

            else:

                candidate = os.path.join(BASE_DIR, stored_img)

                if os.path.isfile(candidate):

                    resolved_img = candidate

                else:

                    candidate2 = os.path.join(BASE_DIR, 'data', stored_img)

                    if os.path.isfile(candidate2):

                        resolved_img = candidate2

                    else:

                        logging.warning(f"Image not found for {product.sku}: {stored_img}")

        

        data.append({

            'model': sku_clean,

            'name_zh': product.sku if (not product.name_zh or str(product.name_zh).strip().lower() in ('nan', 'none', '')) else str(product.name_zh).split(' / ')[0],

            'name_en': '' if (not product.name_en or str(product.name_en).strip().lower() in ('nan', 'none', '')) else str(product.name_en).split(' / ')[0],

            'spec_zh': spec_str,

            'quantity': qty,

            'unit_price': price,

            'total': price * qty,

            '_image_path': resolved_img,

        })

    

    if not data:

        raise ValueError("No valid products found")

    

    # Get supplier name from company config

    supplier = company_config.get('name', '') if company_config else ''

    

    # Create quotation

    qt = QuotationExcel(

        supplier=supplier,

        trade_terms=trade_terms,

        payment_terms=payment_terms,

        currency=currency,

        lang=lang,

        company_info=company_config,

    )

    return qt.write(data, output_path, with_images=with_images)





# ====================  ====================



if __name__ == '__main__':

    test_data = [

        {'model': 'XP', 'name_zh': 'Electric Motorcycle', 'spec_zh': '3000W/4000W', 'quantity': 1, 'unit_price': 4900, 'total': 4900},

        {'model': 'BOX', 'name_zh': 'Electric Motorcycle', 'spec_zh': '3000W', 'quantity': 2, 'unit_price': 4900, 'total': 9800},

        {'model': 'GCD', 'name_zh': 'Electric Motorcycle', 'spec_zh': '4000W Direct', 'quantity': 1, 'unit_price': 3970, 'total': 3970},

    ]

    

    output = r'C:\Users\marky\Desktop\production tool\product_tool\output\test_quotation.xlsx'

    create_quotation(test_data, output, supplier='SONLINK')

    print(f"Generated: {output}")

