# -*- coding: utf-8 -*-
"""
Excel Template Engine
从用户上传的 .xlsx 模板读取布局，替换数据行生成新文档
"""
import os, shutil, logging
from typing import List, Dict, Optional
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font

TEMPLATE_DIR = Path.home() / ".product_tool" / "templates"


def get_template_path(doc_type: str) -> Optional[str]:
    """获取指定文档类型的模板路径"""
    p = TEMPLATE_DIR / f"{doc_type}.xlsx"
    return str(p) if p.exists() else None


def find_data_table(ws):
    """扫描工作表，找到数据区域（行数最多的连续区域）"""
    max_row = ws.max_row
    max_col = min(ws.max_column, 10)
    if max_row is None or max_row < 3:
        return None, None, None
    # 找第一条完整的数据行（所有列都有内容）
    for r in range(1, max_row + 1):
        filled = sum(1 for c in range(1, max_col + 1) if ws.cell(r, c).value is not None)
        if filled >= 3:
            return r, max_row, max_col
    return None, None, None


def apply_template(data: List[Dict], template_path: str, output_path: str) -> bool:
    """从模板生成文档：复制模板 → 清空旧数据 → 填入新数据

    Args:
        data: 产品数据列表
        template_path: 模板文件路径
        output_path: 输出文件路径

    Returns:
        True 表示成功，False 表示失败（模板格式不匹配等）
    """
    if not data or not os.path.exists(template_path):
        return False
    try:
        # 复制模板
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        if ws is None:
            return False

        # 找到数据起始行和列
        data_start, data_end, max_col = find_data_table(ws)
        if data_start is None:
            return False

        # 清空旧数据（保留表头）
        for r in range(data_start, data_end + 1):
            for c in range(1, max_col + 1):
                ws.cell(r, c).value = None

        # 填入新数据
        current_row = data_start
        for i, item in enumerate(data, 1):
            model = item.get('model', '')
            name = item.get('name_zh', '') or model
            spec = item.get('spec_zh', '')
            qty = item.get('qty', item.get('quantity', 1))
            price = item.get('price_rmb', item.get('unit_price', 0))
            total = int(qty) * float(price) if qty and price else 0

            # 写入各列（根据模板原有样式自动继承）
            ws.cell(current_row, 1).value = i           # No.
            ws.cell(current_row, 2).value = model        # Model
            ws.cell(current_row, 3).value = spec         # Spec
            ws.cell(current_row, 4).value = qty          # Qty
            ws.cell(current_row, 5).value = price        # Unit Price
            ws.cell(current_row, 6).value = total        # Total

            current_row += 1

        wb.save(output_path)
        return True
    except Exception as e:
        logging.error(f"Template application failed: {e}")
        return False
