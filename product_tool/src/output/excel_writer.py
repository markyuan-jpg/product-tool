# -*- coding: utf-8 -*-
"""
Excel输出 - 图片嵌入 + 样式美化
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from PIL import Image


# ==================== 样式定义 ====================

# 表头样式
HEADER_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')

# 数据样式
DATA_FONT = Font(name='Microsoft YaHei', size=10)
DATA_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 交替行样式
EVEN_ROW_FILL = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
ODD_ROW_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# 边框
THIN = Side(style='thin', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ==================== 核心函数 ====================

def write_excel_styled(
    output_path: str,
    data: list,
    sheet_name: str = 'Sheet1',
    image_dir: str = None,
    image_column: str = 'image_path'
) -> str:
    """
    带样式和图片的Excel写入
    
    Args:
        output_path: 输出路径
        data: 数据列表
        sheet_name: 工作表名
        image_dir: 图片目录 (可选)
        image_column: 图片列名
    
    Returns:
        输出路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    
    if not data:
        wb.save(output_path)
        return output_path
    
    # 清理数据 - 移除无效列
    clean_data = []
    for item in data:
        clean_item = {}
        for k, v in item.items():
            # 跳过内部字段
            if k in ['specs_raw']:
                continue
            # 处理值
            if v is None:
                clean_item[k] = ''
            elif isinstance(v, (list, dict)):
                clean_item[k] = str(v)
            elif not isinstance(v, (str, int, float)):
                clean_item[k] = str(v)
            else:
                clean_item[k] = v
        clean_data.append(clean_item)
    
    data = clean_data
    
    # 准备数据
    df = pd.DataFrame(data)
    columns = list(df.columns)
    
    # 列索引映射
    img_col_idx = None
    if image_column and image_column in columns:
        img_col_idx = columns.index(image_column) + 1
    
    # 写入表头 (第1行)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(1, col_idx)
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    
    # 写入数据 (从第2行开始)
    for row_idx, row_data in enumerate(data, 2):
        is_even = (row_idx % 2 == 0)
        row_fill = EVEN_ROW_FILL if is_even else ODD_ROW_FILL
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row_idx, col_idx)
            val = row_data.get(col_name)
            
            # 处理值 - 确保是字符串或数字
            if val is None:
                val = ''
            elif isinstance(val, (list, dict)):
                # 列表/字典转为字符串
                val = str(val)
            elif not isinstance(val, (str, int, float)):
                val = str(val)
            
            # 样式
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = BORDER
            cell.fill = row_fill
            
            # 嵌入图片
            if img_col_idx and col_idx == img_col_idx:
                img_path = row_data.get(image_column)
                if img_path and os.path.exists(img_path):
                    try:
                        # 缩放图片
                        img = XLImage(img_path)
                        # 限制最大尺寸
                        max_w, max_h = 80, 60
                        if img.width > max_w:
                            ratio = max_w / img.width
                            img.width = max_w
                            img.height = img.height * ratio
                        if img.height > max_h:
                            ratio = max_h / img.height
                            img.height = max_h
                            img.width = img.width * ratio
                        # 嵌入单元格
                        ws.add_image(img, f'{get_column_letter(col_idx)}{row_idx}')
                    except Exception as e:
                        print(f"图片嵌入失败: {img_path} - {e}")
    
    # 自动列宽
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, len(data) + 2):
            val = ws.cell(row_idx, col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    return output_path


def to_grouped_excel(
    data: list,
    output_path: str,
    group_by: str = 'category',
    sort_by: str = None,
    image_dir: str = None
) -> str:
    """
    按分组输出Excel
    """
    if not data:
        return output_path
    
    # 清理数据
    clean_data = []
    for item in data:
        clean_item = {}
        for k, v in item.items():
            if k in ['specs_raw']:
                continue
            if v is None:
                clean_item[k] = ''
            elif isinstance(v, (list, dict)):
                clean_item[k] = str(v)
            elif not isinstance(v, (str, int, float)):
                clean_item[k] = str(v)
            else:
                clean_item[k] = v
        clean_data.append(clean_item)
    
    data = clean_data
    df = pd.DataFrame(data)
    
    # 排序
    if sort_by and sort_by in df.columns:
        df = df.sort_values(sort_by)
    
    # 分组
    groups = {}
    for item in data:
        cat = item.get(group_by, '未分类')
        if cat not in groups:
            groups[cat] = []
        groups[cat].append(item)
    
    # 创建workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # 按分组创建sheet
    for sheet_name, items in sorted(groups.items()):
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        
        if not items:
            continue
        
        df_sheet = pd.DataFrame(items)
        cols = list(df_sheet.columns)
        
        # 表头
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(1, col_idx)
            cell.value = col_name
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = BORDER
        
        # 数据
        for row_idx, row_data in enumerate(items, 2):
            is_even = (row_idx % 2 == 0)
            row_fill = EVEN_ROW_FILL if is_even else ODD_ROW_FILL
            
            for col_idx, col_name in enumerate(cols, 1):
                cell = ws.cell(row_idx, col_idx)
                val = row_data.get(col_name)
                cell.value = val if val else ''
                cell.font = DATA_FONT
                cell.alignment = DATA_ALIGN
                cell.border = BORDER
                cell.fill = row_fill
        
        # 列宽
        for col_idx in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    return output_path


def to_single_excel(
    data: list,
    output_path: str,
    sort_by: str = None,
    image_dir: str = None
) -> str:
    """单Sheet输出"""
    if sort_by:
        df = pd.DataFrame(data)
        if sort_by in df.columns:
            df = df.sort_values(sort_by)
            data = df.to_dict('records')
    
    return write_excel_styled(output_path, data, image_dir=image_dir)


# ==================== 兼容接口 ====================

def to_grouped_excel_with_styles(data: list, output_path: str, group_by: str = 'category') -> str:
    """兼容旧接口"""
    return to_grouped_excel(data, output_path, group_by)


def write_excel_with_styles(output_path: str, data: list, sheet_name: str = 'Sheet', image_mapping: dict = None) -> str:
    """兼容旧接口"""
    return write_excel_styled(output_path, data, sheet_name, image_dir=None)


# ==================== 测试 ====================

if __name__ == "__main__":
    test_data = [
        {'model': 'G5000', 'name_zh': '电动车', 'spec_zh': '参数1\n参数2', 'price_rmb': 5000, 'category': '电动车'},
        {'model': 'S500', 'name_zh': '电动车', 'spec_zh': '参数3', 'price_rmb': 3000, 'category': '电动车'},
    ]
    
    output = r'C:\Users\marky\Desktop\production tool\product_tool\output\test_styled.xlsx'
    write_excel_styled(output, test_data)
    print(f"生成: {output}")