# -*- coding: utf-8 -*-
"""
PI Generator — 用 openpyxl 生成 Proforma Invoice (xlsx)
格式参照 SONLINK PI 模板
"""
import os
import re
import json
import random
from datetime import datetime
from typing import Dict, List, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from .doc_shared import translate_items, get_seller_info, payment_by_lang
from ..parsers.spec_cleaner import clean_spec

# ─── 样式 ───
FONT_TITLE = Font(name='Arial', size=16, bold=True)
FONT_HEADER = Font(name='Arial', size=10, bold=True)
FONT_NORMAL = Font(name='Arial', size=9)
FONT_SMALL = Font(name='Arial', size=8)
FONT_BOLD = Font(name='Arial', size=9, bold=True)
FONT_TOTAL = Font(name='Arial', size=10, bold=True)
FILL_TITLE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
FONT_TITLE_WHITE = Font(name='Arial', size=16, bold=True, color='FFFFFF')
FILL_HEADER = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
FILL_ODD = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_EVEN = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
FILL_TOTAL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _num_to_words(n: float, currency: str = 'USD', lang: str = 'chinese') -> str:
    """数字金额转英文大写（简化版）"""
    try:
        from src.utils.translator import translate_doc as _td
    except Exception:
        _td = lambda t, l: t
    if not n:
        return f'{_td("ZERO", lang)} {currency}'
    ones = ['', 'ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX', 'SEVEN', 'EIGHT', 'NINE',
            'TEN', 'ELEVEN', 'TWELVE', 'THIRTEEN', 'FOURTEEN', 'FIFTEEN', 'SIXTEEN',
            'SEVENTEEN', 'EIGHTEEN', 'NINETEEN']
    tens = ['', '', 'TWENTY', 'THIRTY', 'FORTY', 'FIFTY', 'SIXTY', 'SEVENTY', 'EIGHTY', 'NINETY']
    def _convert_chunk(num):
        if num == 0: return ''
        res = ''
        if num >= 100:
            res += ones[num // 100] + ' HUNDRED '
            num %= 100
        if num >= 20:
            res += tens[num // 10] + ' '
            num %= 10
        if num > 0:
            res += ones[num] + ' '
        return res.strip()
    integer_part = int(n)
    decimal_part = round((n - integer_part) * 100)
    if integer_part >= 1000000:
        millions = integer_part // 1000000
        thousands = (integer_part % 1000000) // 1000
        remainder = integer_part % 1000
        words = _convert_chunk(millions) + ' MILLION'
        if thousands:
            words += ' ' + _convert_chunk(thousands) + ' THOUSAND'
        if remainder:
            words += ' ' + _convert_chunk(remainder)
    elif integer_part >= 1000:
        thousands = integer_part // 1000
        remainder = integer_part % 1000
        words = _convert_chunk(thousands) + ' THOUSAND'
        if remainder:
            words += ' ' + _convert_chunk(remainder)
    else:
        words = _convert_chunk(integer_part)
    if decimal_part > 0:
        words += f' AND CENTS {decimal_part:02d}/100'
    return f'{words} {currency} {_td("ONLY.", lang)}'


def generate_pi_xlsx(
    items: List[Dict],
    output_path: str,
    buyer_name: str = '',
    buyer_address: str = '',
    seller_config: dict = None,
    invoice_no: str = '',
    trade_terms: str = 'FOB XXXXX',
    payment_terms: str = '',
    port_destination: str = '',
    brand_name: str = '',
    currency: str = 'USD',
    freight: float = 0,
    lang: str = 'chinese',
    port_loading: str = 'Qingdao',
    with_images: bool = True,
) -> str:
    """生成 Proforma Invoice xlsx（SONLINK 格式）

    Args:
        items: 产品列表，每项含 model, name_zh, spec_zh, price_rmb, qty, currency
        output_path: 输出路径
        buyer_name: 买方名称
        buyer_address: 买方地址
        seller_config: 卖方配置字典
        invoice_no: 发票号（自动生成）
        trade_terms: 贸易术语
        payment_terms: 付款条款
        port_destination: 目的港
        brand_name: 品牌名
        currency: 币种
        freight: 运费
    Returns:
        输出路径
    """
    items = translate_items(items, lang)
    try:
        from src.utils.translator import translate_doc as _td
    except Exception:
        _td = lambda t, l: t
    seller = get_seller_info(seller_config)
    # 检查是否有用户上传模板
    try:
        from .excel_template import get_template_path, apply_template
        tmpl = get_template_path('pi')
        if tmpl and apply_template(items, tmpl, str(output_path)):
            return output_path
    except Exception:
        pass
    # 提前导入汇率（不在循环里反复导）
    try:
        from src.rates import get_rate as _get_rate
    except Exception:
        _get_rate = None
    if not invoice_no:
        today = datetime.now().strftime("%Y%m%d")
        rand = f"{random.randint(0, 999):03d}"
        invoice_no = f"XSL-{today}-{rand}"
    inv_date = datetime.now().strftime("%d/%m/%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = 'PI'
    ws.sheet_properties.pageSetUpPr = None

    # 列宽
    col_widths = {'A': 10, 'B': 5, 'C': 45, 'D': 14, 'E': 18, 'F': 20}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ─── Header: Logo + 公司信息 ───
    # Logo预留位
    logo_path = seller.get('logo_path', '')
    if logo_path and os.path.exists(logo_path):
        try:
                from openpyxl.drawing.image import Image as XLImage

                from src.core.image import resize_image

                _logo_resized = resize_image(logo_path, max_w=300)

                logo = XLImage(_logo_resized if hasattr(_logo_resized, 'read') else logo_path)

                logo.width = 60; logo.height = 30

                ws.add_image(logo, f'A{row}')
        except Exception:
            ws.cell(row, 1).value = '[LOGO]'
            ws.cell(row, 1).font = Font(name='Arial', size=8, color='CCCCCC')
    else:
        ws.cell(row, 1).value = '[LOGO]'
        ws.cell(row, 1).font = Font(name='Arial', size=8, color='CCCCCC')
    ws.column_dimensions['A'].width = 12

    header_text = seller.get('company', '') or _td('[Company Name]', lang)
    addr = seller.get('address', '') or _td('[Address]', lang)
    header_text += f' | Add: {addr}'
    ws.merge_cells(f'B{row}:F{row}')
    cell = ws[f'B{row}']
    cell.value = header_text
    cell.font = Font(name='Arial', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 2

    # ─── Title ───
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = _td('PROFORMA INVOICE', lang)
    cell.font = Font(name='Arial', size=18, bold=True, color='1F4E79')
    cell.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 32
    row += 2

    # ─── Invoice Info ───
    cur_label = 'USD' if currency == 'USD' else 'RMB'
    ws.cell(row, 4).value = f"{_td('Invoice No.: ', lang)}{invoice_no}"
    ws.cell(row, 4).font = FONT_NORMAL
    ws.cell(row, 4).alignment = ALIGN_LEFT
    ws.cell(row, 5).value = f"{_td('Date: ', lang)}{inv_date}"
    ws.cell(row, 5).font = FONT_NORMAL
    ws.cell(row, 5).alignment = ALIGN_LEFT
    row += 1

    # ─── Buyer ───
    ws.cell(row, 1).value = f"{_td('To: ', lang)}{buyer_name}" if buyer_name else _td('To: ', lang) + 'XXXXX'
    ws.cell(row, 1).font = FONT_BOLD
    ws.merge_cells(f'A{row}:F{row}')
    if buyer_address:
        row += 1
        ws.cell(row, 1).value = buyer_address
        ws.cell(row, 1).font = FONT_NORMAL
        ws.merge_cells(f'A{row}:F{row}')
    row += 2

    # ─── Column widths (7 columns: A-G) ───
    col_widths = {'A': 10, 'B': 5, 'C': 28, 'D': 28, 'E': 12, 'F': 16, 'G': 18}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ─── Table Header Row 1 (merged C:D = DESCRIPTION OF GOODS) ───
    ws.merge_cells(f'C{row}:D{row}')
    for ci in range(1, 8):
        h = [_td('Photo', lang), _td('No.', lang), _td('DESCRIPTION OF GOODS', lang), '', _td('Quantity\n(sets)', lang), f"{_td('Unit Price', lang)}\n({cur_label})", f"{_td('Total', lang)}\n({cur_label})"][ci - 1]
        if ci == 4:  # D 列已合并到 C，跳过
            continue
        cell = ws.cell(row, ci)
        cell.value = h
        cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        cell.fill = FILL_TITLE
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30
    row += 1

    # ─── Table Header Row 2 (sub-headers for C & D) ───
    sub_headers = ['', '', _td('Model / Product Name', lang), _td('Specifications', lang), '', '', '']
    for ci, h in enumerate(sub_headers, 1):
        cell = ws.cell(row, ci)
        cell.value = h
        cell.font = Font(name='Arial', size=8, bold=True, color='FFFFFF')
        cell.fill = FILL_TITLE
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    # ─── Product Rows ───
    total_qty = 0
    total_amount = 0

    for i, item in enumerate(items, 1):
        qty = int(item.get('qty', item.get('quantity', 1)))
        price = float(item.get('price_rmb', item.get('unit_price', 0)))
        # 目标币种是 USD 且来源币种非 USD 时，才从 CNY 转换为 USD
        source_currency = item.get('currency', item.get('source_currency', 'RMB'))
        if currency == 'USD' and source_currency not in ('USD', ''):
            try:
                if _get_rate:
                    rate = _get_rate('CNY', 'USD') or 0.14
                    price = round(price * rate, 2)
            except Exception:
                pass
        total = qty * price
        total_qty += qty
        total_amount += total

        model = item.get('model', '') or item.get('name_zh', '')
        name = item.get('name_zh', '') or model
        spec = clean_spec(item.get('spec_zh', ''))
        model_desc = f"{model}\n{name}" if name and name != model else model

        # 嵌入图片（处理路径含空格问题）
        img_path = item.get('_image_path', '') or item.get('image_path', '')
        img_path = img_path.strip() if img_path else ''
        has_img = bool(with_images and img_path and os.path.exists(img_path))

        row_data = ['', i, model_desc, spec, qty, price, total]
        is_odd = i % 2 == 1
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row, ci)
            cell.value = val
            cell.font = FONT_SMALL
            cell.border = THIN_BORDER
            cell.fill = FILL_ODD if is_odd else FILL_EVEN
            if ci in (5, 6, 7):
                cell.alignment = ALIGN_RIGHT
            elif ci in (3, 4):
                cell.alignment = ALIGN_LEFT
            else:
                cell.alignment = ALIGN_CENTER
        # 嵌入产品图片到Photo列
        if has_img:
            try:
                from openpyxl.drawing.image import Image as XLImage
                from src.core.image import resize_image
                _prod_resized = resize_image(img_path)
                img = XLImage(_prod_resized if hasattr(_prod_resized, 'read') else img_path)
                img.width = 50; img.height = 50
                ws.add_image(img, f'A{row}')
                ws.row_dimensions[row].height = max(80, len(spec.split('\n')) * 14)
            except Exception:
                pass
        row += 1

    # ─── Total Row ───
    city_name = seller.get('city', '').upper() or 'XXXXX'
    # 不能合并，因为 D/E/F 列各要写不同内容
    ws.cell(row, 1).value = f"{_td('TOTAL AMOUNT', lang)} {_td('FOB', lang)} {city_name}:"
    ws.cell(row, 1).font = FONT_BOLD
    ws.cell(row, 1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row, 1).border = THIN_BORDER
    ws.cell(row, 1).fill = FILL_TOTAL
    for ci in range(2, 7):
        ws.cell(row, ci).fill = FILL_TOTAL
        ws.cell(row, ci).border = THIN_BORDER
    ws.cell(row, 4).value = total_qty
    ws.cell(row, 4).font = FONT_BOLD
    ws.cell(row, 4).alignment = ALIGN_RIGHT
    ws.cell(row, 5).value = '-'
    ws.cell(row, 5).font = FONT_BOLD
    ws.cell(row, 5).alignment = ALIGN_CENTER
    ws.cell(row, 6).value = total_amount
    ws.cell(row, 6).font = FONT_BOLD
    ws.cell(row, 6).alignment = ALIGN_RIGHT
    ws.row_dimensions[row].height = 22
    row += 1

    # ─── Amount in Words ───
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row, 1).value = f"{_td('TOTAL PAYMENT', lang)} {_td('SAY', lang)} {_num_to_words(total_amount, currency, lang)}"
    ws.cell(row, 1).font = Font(name='Arial', size=8, italic=True)
    ws.cell(row, 1).alignment = ALIGN_LEFT
    row += 2

    # ─── Terms (SONLINK 标准格式) ───
    terms = [
        f"1. {payment_by_lang(payment_terms, lang) if payment_terms else _td('30% deposit by T/T, 70% balance before shipment within 60 days upon receipt of payment.', lang)}",
        f"2. {_td('Delivery date: 60 days upon receipt of payment.', lang)}",
        f"3. {_td('Port of destination:', lang)} {port_destination or '______________'}",
        f"4. {_td('Brand Name:', lang)} {brand_name or '______________'}",
        f"5. {_td('Transshipment: not allowed; Partial shipment: not allowed.', lang)}",
        f"6. {_td('Validity: This quotation is valid for 30 days from the date above.', lang)}",
        f"7. {_td('Insurance: To be covered by the buyer.', lang)}",
        f"8. {_td('Documents required: Commercial Invoice, Packing List, Bill of Lading, Certificate of Origin.', lang)}",
        f"9. {_td('Handling method on the expiry order: The contract will automatically become void if the buyer does not provide any instruction before the delivery date.', lang)}",
        f"10. {_td('Handling method on discrepancy of quality & quantity: In case of quality discrepancy, claim should be filed within 30 days after arrival of goods.', lang)}",
        f"11. {_td('Arbitration: All disputes shall be settled through friendly negotiation. If no settlement can be reached, the case shall be submitted to China International Economic and Trade Arbitration Commission for arbitration.', lang)}",
    ]

    for t in terms:
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row, 1).value = t
        ws.cell(row, 1).font = FONT_SMALL
        ws.cell(row, 1).alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = 20
        row += 1

    # ─── Bank Info（每行一个字段，空值用占位符） ───
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row, 1).value = f"8.  {_td('Bank Information:', lang)}"
    ws.cell(row, 1).font = Font(name='Arial', size=9, bold=True)
    ws.cell(row, 1).alignment = ALIGN_LEFT
    row += 1
    bank_labels = [
        (_td('Beneficiary:', lang), 'bank_beneficiary'),
        (_td('Bank name:', lang), 'bank_name'),
        (_td('Bank add.:', lang), 'bank_address'),
        (_td('Bank account no.:', lang), 'bank_account'),
        (_td('Swift code.:', lang), 'bank_swift'),
    ]
    for label, key in bank_labels:
        val = seller.get(key, '') or '______________'
        ws.cell(row, 1).value = f'   {label}  {val}'
        ws.cell(row, 1).font = FONT_SMALL
        ws.cell(row, 1).alignment = ALIGN_LEFT
        row += 1

    # Other Notices
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row, 1).value = f"9. {_td('Other Notices: The supplier provide the Certificate of Origin (C/O) to the buyer ONLY. Other documents or certificates will be charged additionally.', lang)}"
    ws.cell(row, 1).font = FONT_SMALL
    ws.cell(row, 1).alignment = ALIGN_LEFT
    row += 2

    # ─── Signature ───
    ws.cell(row, 1).value = _td('The seller:', lang)
    ws.cell(row, 1).font = FONT_BOLD
    ws.cell(row, 4).value = _td('The buyer:', lang)
    ws.cell(row, 4).font = FONT_BOLD
    row += 1
    ws.cell(row, 1).value = _td('(signed & stamped)', lang)
    ws.cell(row, 1).font = Font(name='Arial', size=8, color='888888')
    ws.cell(row, 4).value = _td('(signed & stamped)', lang)
    ws.cell(row, 4).font = Font(name='Arial', size=8, color='888888')
    row += 2

    # ─── Footer ───
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row, 1).value = seller.get('company', '')
    ws.cell(row, 1).font = Font(name='Arial', size=9, bold=True, color='1F4E79')
    ws.cell(row, 1).alignment = ALIGN_CENTER

    # Print Layout
    ws.print_area = f'A1:F{row}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    wb.save(output_path)
    return output_path


# 保持兼容旧接口
def generate_pi(
    df_or_items,
    buyer_info: Dict[str, str] = None,
    seller_info: Dict[str, str] = None,
    output_path: str = None,
    **kwargs
) -> str:
    """兼容入口：自动转成新格式"""
    import pandas as pd

    if isinstance(df_or_items, pd.DataFrame):
        df = df_or_items
        items = []
        for _, row in df.iterrows():
            items.append({
                'model': row.get('model', ''),
                'name_zh': row.get('name_zh', '') or row.get('model', ''),
                'spec_zh': row.get('spec_zh', ''),
                'price_rmb': row.get('price_rmb', row.get('price_usd', 0)),
                'qty': row.get('quantity', row.get('qty', 1)),
                'currency': row.get('currency', 'USD'),
            })
    else:
        items = df_or_items

    buyer_name = ''
    buyer_address = ''
    if buyer_info:
        buyer_name = buyer_info.get('company', '') or buyer_info.get('buyer_name', '')
        buyer_address = buyer_info.get('address', '') or buyer_info.get('buyer_address', '')

    if not output_path:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        output_path = f'PI_{ts}.xlsx'

    return generate_pi_xlsx(
        items=items,
        output_path=output_path,
        buyer_name=buyer_name,
        buyer_address=buyer_address,
        seller_config=seller_info,
        **kwargs
    )


if __name__ == '__main__':
    test_items = [
        {'model': 'G5000', 'name_zh': 'E-motorcycle', 'spec_zh': 'Motor: 5000W\nBattery: 72V 20AH', 'price_rmb': 7560, 'qty': 10},
        {'model': 'S1', 'name_zh': 'Electric Scooter', 'spec_zh': '450W 48V20Ah', 'price_rmb': 1325, 'qty': 50},
    ]
    out = generate_pi_xlsx(test_items, 'test_pi.xlsx', buyer_name='ABC Company Ltd.', trade_terms='FOB Qingdao')
    print(f'PI generated: {out}')
