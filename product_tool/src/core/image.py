# -*- coding: utf-8 -*-
"""
图片提取模块 - 统一图片提取与产品匹配引擎

三路匹配:
1. openpyxl _images (标准嵌入)
2. DISPIMG 公式 (WPS)
3. 文件夹 SKU 匹配 (回退)
"""
import os, io
import re
import hashlib
import glob
import logging
import zipfile
from typing import Optional, Dict, List, Tuple, Union
from xml.etree import ElementTree as ET
from PIL import Image as PILImage

TEMP_IMAGE_FOLDER = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'temp_images')

# ─── 图片压缩（生成时使用，避免原始分辨率嵌入） ───

MAX_IMAGE_WIDTH = 1200
JPEG_QUALITY = 95
_resize_cache = {}

def resize_image(path: str, max_w: int = MAX_IMAGE_WIDTH, quality: int = JPEG_QUALITY) -> Union[io.BytesIO, str]:
    global _resize_cache
    key = (path, max_w, quality)
    if key in _resize_cache:
        return _resize_cache[key]
    """压缩大图到指定宽度，保留原格式。宽 <= max_w 的图片原样返回。
    
    Args:
        path: 图片路径
        max_w: 最大宽度像素（默认 1200）
        quality: JPEG quality（默认 95，视觉无损）
    Returns:
        BytesIO（已压缩）或原始路径字符串（无需压缩时）
    """
    try:
        img = PILImage.open(path)
    except Exception:
        return path  # 打不开的图片原样返回
    
    # 小图不处理
    if img.width <= max_w:
        return path
    
    # 转换 RGBA/P 模式为 RGB
    if img.mode in ('RGBA', 'P'):
        img = img.convert('RGB')
    
    ratio = max_w / img.width
    new_h = int(img.height * ratio)
    img = img.resize((max_w, new_h), PILImage.LANCZOS)
    
    fmt = 'JPEG' if path.lower().endswith(('.jpg', '.jpeg')) else 'PNG'
    buf = io.BytesIO()
    save_kw = {'format': fmt, 'quality': quality} if fmt == 'JPEG' else {'format': fmt}
    img.save(buf, **save_kw)
    buf.seek(0)
    _resize_cache[key] = buf
    return buf

DRAWING_NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
}
SPREADSHEET_NS = {
    's': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
}
CELLIMG_NS = {
    'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData',
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
}


def get_image_hash(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()


def _save_image_unique(img_data: bytes, ext: str) -> str:
    """按内容 hash 保存图片，相同 hash 不重复写入"""
    os.makedirs(TEMP_IMAGE_FOLDER, exist_ok=True)
    h = get_image_hash(img_data)
    ext_clean = ext.lstrip('.')
    filename = f"{h}.{ext_clean}"
    img_path = os.path.join(TEMP_IMAGE_FOLDER, filename)
    if not os.path.exists(img_path):
        with open(img_path, 'wb') as f:
            f.write(img_data)
    return img_path


# ==================== 第1路: openpyxl _images ====================

def _detect_image_column(file_path: str) -> Optional[int]:
    """从表头行检测产品图片列位置（1-based）"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        for ws in wb.worksheets:
            for r in range(1, min(ws.max_row + 1, 5)):
                for c in range(1, min(ws.max_column + 1, 15)):
                    val = str(ws.cell(r, c).value or '').lower()
                    if any(kw in val for kw in ['产品图片', '商品图', '产品图', 'picture', '图片']):
                        wb.close()
                        return c
        wb.close()
    except Exception:
        pass
    return None


def _image_col_matches(anchor_col: int, image_col: Optional[int], tolerance: int = 1) -> bool:
    """判断图片锚点列是否在图片列附近"""
    if image_col is None:
        return True  # 没检测到图片列 → 不过滤
    return abs(anchor_col - image_col) <= tolerance


def extract_openpyxl_images(file_path: str, image_col: Optional[int] = None) -> Dict[str, Dict[int, str]]:
    """
    通过 openpyxl _images 提取标准嵌入图片
    返回: {sheet_name: {row(1-based): file_path}}
    image_col: 只提取该列附近的图片（1-based），None=全部提取
    """
    result = {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return result

    wb = None
    try:
        wb = load_workbook(file_path, data_only=False)
    except Exception:
        return result

    filename = os.path.basename(file_path).replace('.xlsx', '')

    try:
        for ws in wb:
            sheet_name = ws.title
            if not hasattr(ws, '_images') or not ws._images:
                continue
            for img in ws._images:
                try:
                    row = 0
                    col = None
                    if hasattr(img.anchor, '_from') and hasattr(img.anchor._from, 'row'):
                        row = img.anchor._from.row + 1
                        col = img.anchor._from.col + 1
                    if row <= 1:
                        continue
                    if not _image_col_matches(col, image_col):
                        continue
                    img_data = img._data()
                    if not img_data:
                        continue
                    ext = '.jpg' if getattr(img, 'format', '').lower() == 'jpeg' else '.png'
                    img_path = _save_image_unique(img_data, ext)
                    if sheet_name not in result:
                        result[sheet_name] = {}
                    result[sheet_name][row] = img_path
                except Exception:
                    continue
    finally:
        if wb:
            wb.close()
    return result


# ==================== 第2路: DISPIMG 公式解析 ====================

def _get_sheet_name_map(z: zipfile.ZipFile) -> Dict[int, str]:
    wb = z.read('xl/workbook.xml')
    root = ET.fromstring(wb)
    sheets = {}
    idx = 0
    for sheet in root.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
        name = sheet.attrib.get('name', f'Sheet{idx + 1}')
        sheets[idx] = name
        idx += 1
    return sheets


def _cell_ref_to_col(cell_ref: str) -> int:
    """将 Excel 列字母转为 1-based 列号: A→1, C→3, AA→27"""
    col = 0
    for ch in cell_ref:
        if ch.isalpha():
            col = col * 26 + (ord(ch.upper()) - ord('A') + 1)
    return col


def parse_dispimg_images(file_path: str, image_col: Optional[int] = None) -> Dict[str, Dict[int, str]]:
    """
    解析 DISPIMG (WPS) 公式图片
    返回: {sheet_name: {row(1-based): file_path}}
    image_col: 只提取该列附近的图片，None=全部提取
    """
    result = {}
    if not file_path or not file_path.endswith('.xlsx'):
        return result

    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            if 'xl/cellimages.xml' not in z.namelist():
                return result

            # step 1: cellimages.xml → (guid → rId)
            cellimg_root = ET.fromstring(z.read('xl/cellimages.xml'))
            guid_to_rid = {}
            for cellimg in cellimg_root.iter('{http://www.wps.cn/officeDocument/2017/etCustomData}cellImage'):
                nvpr = cellimg.find('.//xdr:cNvPr', CELLIMG_NS)
                blip = cellimg.find('.//a:blip', {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'})
                if nvpr is not None and blip is not None:
                    guid = nvpr.attrib.get('name', '')
                    rid = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', '')
                    if guid:
                        guid_to_rid[guid] = rid

            if not guid_to_rid:
                return result

            # step 2: rels → (rId → media_file)
            rels_root = ET.fromstring(z.read('xl/_rels/cellimages.xml.rels'))
            rid_to_media = {}
            for child in rels_root:
                rid = child.attrib.get('Id', '')
                target = child.attrib.get('Target', '').replace('media/', '')
                rid_to_media[rid] = target

            # step 3: guid → media_file
            guid_to_media = {}
            for guid, rid in guid_to_rid.items():
                media_name = rid_to_media.get(rid, '')
                media_file = f'xl/media/{media_name}'
                if media_file in z.namelist():
                    guid_to_media[guid] = media_file

            if not guid_to_media:
                return result

            # step 4: scan worksheets for DISPIMG formulas
            sheet_names = _get_sheet_name_map(z)
            for sheet_idx in range(len(sheet_names)):
                ws_path = f'xl/worksheets/sheet{sheet_idx + 1}.xml'
                if ws_path not in z.namelist():
                    continue
                ws_root = ET.fromstring(z.read(ws_path))
                sheet_name = sheet_names.get(sheet_idx, f'Sheet{sheet_idx + 1}')
                sheet_images = {}

                for row_elem in ws_root.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                    r = int(row_elem.attrib.get('r', '0'))
                    if r <= 1:
                        continue
                    for c in row_elem:
                        cell_ref = c.attrib.get('r', '')
                        cell_col = _cell_ref_to_col(cell_ref) if cell_ref else None
                        f = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f')
                        if f is not None and f.text and 'DISPIMG' in f.text.upper():
                            if not _image_col_matches(cell_col, image_col):
                                continue
                            m = re.search(r'"([^"]+)"', f.text)
                            if m:
                                guid = m.group(1)
                                media_file = guid_to_media.get(guid, '')
                                if media_file:
                                    img_data = z.read(media_file)
                                    ext = os.path.splitext(media_file)[1] or '.png'
                                    img_path = _save_image_unique(img_data, ext)
                                    sheet_images[r] = img_path

                if sheet_images:
                    result[sheet_name] = sheet_images

    except Exception as e:
        logging.warning(f"DISPIMG parse failed for {file_path}: {e}")

    return result


# ==================== 第3路: drawing XML ====================

def parse_drawing_images(file_path: str, image_col: Optional[int] = None) -> Dict[str, Dict[int, str]]:
    """
    解析标准 xlsx drawing XML 图片
    返回: {sheet_name: {row(1-based): file_path}}
    image_col: 只提取该列附近的图片，None=全部提取
    """
    result = {}
    if not file_path or not file_path.endswith('.xlsx'):
        return result

    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            # 获取 sheet → drawing 映射
            wb_rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
            rid_to_sheet = {}
            for child in wb_rels:
                target = child.attrib.get('Target', '')
                rid = child.attrib.get('Id', '')
                if 'worksheet' in target:
                    rid_to_sheet[rid] = target

            wb = ET.fromstring(z.read('xl/workbook.xml'))
            sheet_map = {}  # {drawing_name: sheet_name}
            for sheet in wb.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
                rid = sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id', '')
                name = sheet.attrib.get('name', '')
                ws_path = rid_to_sheet.get(rid, '')
                if ws_path:
                    ws_name = os.path.basename(ws_path)
                    ws_rels_path = f'xl/worksheets/_rels/{ws_name}.rels'
                    if ws_rels_path in z.namelist():
                        ws_rels = ET.fromstring(z.read(ws_rels_path))
                        for r in ws_rels:
                            if 'drawing' in r.attrib.get('Type', ''):
                                drawing_target = r.attrib.get('Target', '')
                                drawing_name = os.path.basename(drawing_target)
                                sheet_map[drawing_name] = name

            for drawing_name, sheet_name in sheet_map.items():
                draw_path = f'xl/drawings/{drawing_name}'
                rels_path = f'xl/drawings/_rels/{drawing_name}.rels'
                if draw_path not in z.namelist():
                    continue

                # rId → media_file
                rid_to_media = {}
                if rels_path in z.namelist():
                    rels_root = ET.fromstring(z.read(rels_path))
                    for child in rels_root:
                        rid = child.attrib.get('Id', '')
                        target = child.attrib.get('Target', '').replace('../media/', '')
                        rid_to_media[rid] = target

                draw_root = ET.fromstring(z.read(draw_path))
                sheet_images = {}

                for anchor in draw_root.iter('{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}twoCellAnchor'):
                    from_elem = anchor.find('xdr:from', DRAWING_NS)
                    if from_elem is None:
                        continue
                    row_el = from_elem.find('xdr:row', DRAWING_NS)
                    if row_el is None:
                        continue
                    row = int(row_el.text) + 1
                    if row <= 1:
                        continue
                    col_el = from_elem.find('xdr:col', DRAWING_NS)
                    anchor_col = (int(col_el.text) + 1) if col_el is not None else None
                    if not _image_col_matches(anchor_col, image_col):
                        continue
                    
                    blip = anchor.find('.//a:blip', DRAWING_NS)
                    if blip is None:
                        continue
                    rid = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', '')
                    media_name = rid_to_media.get(rid, '')
                    if not media_name:
                        continue
                    media_file = f'xl/media/{media_name}'
                    if media_file not in z.namelist():
                        continue

                    img_data = z.read(media_file)
                    ext = os.path.splitext(media_name)[1] or '.png'
                    img_path = _save_image_unique(img_data, ext)

                    if row not in sheet_images:
                        sheet_images[row] = img_path

                if sheet_images:
                    result[sheet_name] = sheet_images

    except Exception as e:
        logging.warning(f"Drawing parse failed for {file_path}: {e}")

    return result


# ==================== 三路统一入口 ====================

def extract_embedded_images(file_path: str, image_col: Optional[int] = None) -> Dict[str, Dict[int, str]]:
    """
    统一入口: 三路合并, 返回 {sheet_name: {row(1-based): file_path}}
    
    image_col: 只提取该列附近的图片（1-based），None=全部提取
    
    优先级:
    1. openpyxl _images (标准嵌入, 有 anchor)
    2. DISPIMG (WPS 公式, 覆盖第1路缺失的)
    3. drawing XML (标准 drawing, 覆盖前两路都缺失的)
    """
    result = {}

    try:
        r1 = extract_openpyxl_images(file_path, image_col=image_col)
        for s, rows in r1.items():
            result.setdefault(s, {}).update(rows)
    except Exception:
        pass

    try:
        r2 = parse_dispimg_images(file_path, image_col=image_col)
        for s, rows in r2.items():
            result.setdefault(s, {}).update(rows)
    except Exception:
        pass

    try:
        r3 = parse_drawing_images(file_path, image_col=image_col)
        for s, rows in r3.items():
            existing = result.get(s, {})
            for row, path in rows.items():
                if row not in existing:
                    result.setdefault(s, {})[row] = path
    except Exception:
        pass

    return result


def find_folder_images(file_path: str) -> List[str]:
    """从同目录 images/ 文件夹匹配图片"""
    file_dir = os.path.dirname(file_path)
    possible_folders = [
        os.path.join(file_dir, 'images'),
        os.path.join(file_dir, 'imgs'),
        os.path.join(file_dir, 'pics'),
    ]
    image_files = []
    for folder in possible_folders:
        if os.path.exists(folder):
            for ext in ['*.jpg', '*.jpeg', '*.png', '*.bmp']:
                image_files.extend(glob.glob(os.path.join(folder, ext)))
    return sorted(image_files)


def get_all_images(file_path: str) -> List[str]:
    """获取所有可用图片 (嵌入优先, 文件夹回退)"""
    embedded = extract_embedded_images(file_path)
    all_images = []
    if embedded:
        for sheet_name in sorted(embedded.keys()):
            for row in sorted(embedded[sheet_name].keys()):
                all_images.append(embedded[sheet_name][row])
    if not all_images:
        all_images = find_folder_images(file_path)
    return all_images


# ==================== 文件夹 SKU 匹配 ====================

def match_sku_folder(sku: str, image_dirs: List[str]) -> Optional[str]:
    if not sku or not image_dirs:
        return None
    sku_lower = sku.lower().strip()
    candidates = []
    for img_dir in image_dirs:
        if not os.path.isdir(img_dir):
            continue
        for ext in ['*.jpg', '*.jpeg', '*.png', '*.bmp', '*.gif']:
            for fpath in glob.glob(os.path.join(img_dir, '**', ext), recursive=True):
                fname = os.path.splitext(os.path.basename(fpath).lower())[0]
                score = 0
                if fname == sku_lower:
                    score = 3
                elif sku_lower in fname or fname in sku_lower:
                    score = 2
                elif any(part in fname for part in sku_lower.split('-')):
                    score = 1
                if score > 0:
                    candidates.append((score, fpath))
    if not candidates:
        return None
    candidates.sort(key=lambda x: -x[0])
    return candidates[0][1]


# ==================== DOCX 图片提取 ====================

def extract_images_from_docx(file_path: str) -> Dict[int, str]:
    """
    从 DOCX 提取图片, 按顺序返回 {index: file_path}
    
    DOCX 图片在 word/media/ 中, 按文件名排序后与产品顺序匹配
    """
    import zipfile as zf_local
    images = {}
    if not file_path or not file_path.endswith('.docx'):
        return images
    try:
        with zf_local.ZipFile(file_path, 'r') as z:
            media_files = sorted([
                n for n in z.namelist()
                if n.startswith('word/media/') and not n.endswith('/')
            ])
            if not media_files:
                return images
            os.makedirs(TEMP_IMAGE_FOLDER, exist_ok=True)
            for idx, mf in enumerate(media_files):
                img_data = z.read(mf)
                ext = os.path.splitext(mf)[1] or '.png'
                img_path = _save_image_unique(img_data, ext)
                images[idx] = img_path
    except Exception as e:
        logging.warning(f"DOCX image extraction failed: {e}")
    return images


def match_images_to_products_docx(df, file_path: str):
    """为 DOCX 解析的 DataFrame 按顺序分配图片"""
    images = extract_images_from_docx(file_path)
    if not images or df.empty:
        return df
    df['_image_path'] = ''
    for idx in range(min(len(df), len(images))):
        df.at[idx, '_image_path'] = images[idx]
    return df


# ==================== 统一产品匹配函数 ====================

def match_images_to_products(df, file_path: str) -> 'pd.DataFrame':
    """
    统一匹配入口: 根据解析结果中的 _row + _sheet 匹配图片

    输入: DataFrame(含 _row, _sheet, _source_file 列)
    输出: DataFrame(带 _image_path)

    匹配策略:
    a. 精确匹配: (sheet, _row)
    b. 容差 ±1 (anchor 偏移)
    c. 顺序匹配 (兜底)
    """
    import pandas as pd

    if df.empty:
        return df
    if '_row' not in df.columns:
        return df

    # 检测产品图片列，只提取该列的图片（过滤包装图/适合图）
    image_col = _detect_image_column(file_path)
    img_map = extract_embedded_images(file_path, image_col=image_col)
    if not img_map:
        return df

    def _sheet_matches(product_sheet: str, map_sheet: str) -> bool:
        ps = str(product_sheet or '').strip().lower()
        ms = str(map_sheet or '').strip().lower()
        return ps == ms or ps in ms or ms in ps

    all_imgs = []
    for s_name in sorted(img_map.keys()):
        for r in sorted(img_map[s_name].keys()):
            all_imgs.append((s_name, r, img_map[s_name][r]))

    def find_image(row):
        row_num = row.get('_row')
        sheet = row.get('_sheet', '')
        if row_num is None:
            return _order_match(row.name, all_imgs)

        for s_name, s_rows in img_map.items():
            if _sheet_matches(sheet, s_name):
                if row_num in s_rows:
                    return s_rows[row_num]

        for s_name, s_rows in img_map.items():
            if _sheet_matches(sheet, s_name):
                for delta in [-1, 1]:
                    candidate = row_num + delta
                    if candidate in s_rows:
                        return s_rows[candidate]

        return _order_match(row.name, all_imgs)

    def _order_match(idx, images):
        if not images:
            return None
        return images[idx % len(images)][2]

    df['_image_path'] = df.apply(find_image, axis=1)

    # ─── 图片扩散：同一 sheet 内，将图片扩散到相邻无图片的产品 ───
    if '_sheet' in df.columns and '_row' in df.columns:
        for sheet_name in df['_sheet'].unique():
            sheet_df = df[df['_sheet'] == sheet_name]
            # 找到有图片的行
            img_rows = sheet_df[sheet_df['_image_path'].notna() & (sheet_df['_image_path'] != '')]
            for _, img_row in img_rows.iterrows():
                img_path = img_row['_image_path']
                img_row_num = img_row['_row']
                # 上下各扩散 5 行
                for delta in range(-5, 6):
                    if delta == 0:
                        continue
                    target = sheet_df[sheet_df['_row'] == img_row_num + delta]
                    if target.empty:
                        continue
                    target_idx = target.index[0]
                    if not df.at[target_idx, '_image_path']:
                        df.at[target_idx, '_image_path'] = img_path

    # 对仍未匹配的, 尝试 SKU 文件夹回退
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    default_image_dirs = [
        os.path.join(base_dir, 'data', '新能源电动车', 'images'),
        os.path.join(base_dir, 'data', 'images'),
        os.path.join(os.path.dirname(file_path), 'images'),
    ]
    for idx2, row in df.iterrows():
        if row.get('_image_path'):
            continue
        sku = row.get('model', '')
        img = match_sku_folder(sku, default_image_dirs)
        if img:
            df.at[idx2, '_image_path'] = img

    return df
