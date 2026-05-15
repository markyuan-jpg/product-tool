# -*- coding: utf-8 -*-
"""
Excel通用解析器 V3.1 - 修复版
修复了以下问题:
1. param_price: 支持 Model: 和 型号: 双标记
2. BAOSHIMA: 新增single_spec格式
3. SONLINK: 产品区域截断
4. 车型价格表: 横向检测加数据行验证
5. 横向参数名: 多级路径合并
6. 参数列表边界: 空行检测去重
"""
import pandas as pd
import numpy as np
import os
import re
from typing import Optional, Dict, List, Tuple, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..utils.price import clean_price_value


def extract_images_from_worksheet(ws, output_dir: str, base_name: str, file_path: str = None) -> Dict[int, str]:
    """Extract images from worksheet and save to files.
    
    Returns: dict mapping row_number -> image_file_path
    """
    images_by_row = {}
    
    if not hasattr(ws, '_images') or not ws._images:
        return images_by_row
    
    output_dir = os.path.join(output_dir, 'images', base_name)
    os.makedirs(output_dir, exist_ok=True)
    
    # Use the provided file_path
    if not file_path:
        return images_by_row
    
    # Read images from zip archive
    import zipfile
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            media_files = [n for n in z.namelist() if n.startswith('xl/media/')]
            
            for idx, img in enumerate(ws._images):
                try:
                    # Get anchor row
                    anchor = img.anchor
                    row = 1
                    if hasattr(anchor, '_from') and hasattr(anchor._from, 'row'):
                        row = anchor._from.row + 1
                    elif hasattr(anchor, 'row'):
                        row = anchor.row + 1
                    
                    # Get image path
                    img_path_str = img.path
                    if isinstance(img_path_str, str):
                        # Find matching file in archive
                        media_name = os.path.basename(img_path_str)
                        for media_file in media_files:
                            if media_name in media_file:
                                img_data = z.read(media_file)
                                
                                ext = 'jpg' if 'jpeg' in media_file or 'jpg' in media_file else 'png'
                                img_name = f'img_{idx + 1}_{row}.{ext}'
                                full_img_path = os.path.join(output_dir, img_name)
                                
                                if not os.path.exists(full_img_path):
                                    with open(full_img_path, 'wb') as f:
                                        f.write(img_data)
                                images_by_row[row] = full_img_path
                                print(f'Saved image: {full_img_path}')
                                break
                except Exception as e:
                    print(f'Image extract error: {e}')
                    continue
    except Exception as e:
        print(f'Zip access error: {e}')
    
    return images_by_row


def is_numeric(value) -> bool:
    """Check if value is numeric (supporting currency prefixes)"""
    if value is None:
        return False
    try:
        v = str(value).replace(',', '').replace('$', '').replace('¥', '').replace('€', '').replace('£', '').strip()
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def count_numeric(values: List) -> int:
    """Count numeric values in list"""
    return sum(1 for v in values if is_numeric(v))


def is_model_code(value) -> bool:
    """Check if value looks like a model code (alphanumeric, no Chinese chars)"""
    if value is None:
        return False
    v = str(value).strip()
    if not v:
        return False
    # Model codes: alphanumeric with optional hyphens/underscores, no Chinese
    if re.search(r'[\u4e00-\u9fff]', v):
        return False
    if re.match(r'^[A-Za-z0-9][A-Za-z0-9\-_\.\/]*$', v) and len(v) >= 2:
        return True
    return False


def count_model_code(values: List) -> int:
    """Count model codes in list"""
    return sum(1 for v in values if is_model_code(v))


def is_spec_keyword(value) -> bool:
    """Check if value contains spec-related keywords"""
    if value is None:
        return False
    v = str(value).strip().lower()
    spec_kw = ['mm', 'cm', 'kg', 'g', 'w', 'v', 'a', 'hz',
               '电压', '功率', '尺寸', '重量', '材质', '颜色',
               '电池', '电机', '速度', '电流', '容量']
    return any(kw in v for kw in spec_kw)


def count_spec_keywords(values: List) -> int:
    """Count spec keyword occurrences in list"""
    return sum(1 for v in values if is_spec_keyword(v))


def get_column_stats(values: List) -> Dict[str, float]:
    values = [v for v in values if v is not None and str(v).strip()]
    if not values:
        return {}
    n = len(values)
    scores = {}
    
    num_count = count_numeric(values)
    numeric_ratio = num_count / n if n > 0 else 0
    num_values = [float(str(v).replace(',', '').replace('$', '').replace('¥', '')) 
                for v in values if is_numeric(v)]
    if num_values:
        max_price = max(num_values)
        min_price = min(num_values)
        if 0.01 <= min_price <= max_price <= 100000 and numeric_ratio > 0.5:
            scores['price'] = numeric_ratio * 0.9
        elif numeric_ratio > 0.8 and max_price > 100000:
            scores['price'] = numeric_ratio * 0.8
    
    model_count = count_model_code(values)
    model_ratio = model_count / n if n > 0 else 0
    if model_ratio > 0.3:
        scores['model'] = model_ratio * 0.85
    
    spec_count = count_spec_keywords(values)
    spec_ratio = spec_count / n if n > 0 else 0
    if spec_ratio > 0.1:
        scores['spec'] = spec_ratio * 0.8
    
    name_keywords = ['产品', '名称', '品名', 'product', 'name', 'description', 
                   '规格', '型号', 'item', '货号']
    name_count = sum(1 for v in values for kw in name_keywords 
                    if kw in str(v).lower())
    name_ratio = name_count / n if n > 0 else 0
    avg_len = sum(len(str(v)) for v in values) / n if n > 0 else 0
    if name_ratio > 0.05 or avg_len > 10:
        scores['name'] = 0.5
    
    return scores


# ==================== 第二部分: 格式分类器 ====================

def count_row_non_null(ws, row_idx, start_col=1, end_col=None):
    """统计某行非空值数量"""
    if end_col is None:
        end_col = ws.max_column + 1
    count = 0
    for c in range(start_col, end_col):
        if ws.cell(row_idx, c).value:
            count += 1
    return count


def is_single_spec_sheet(ws) -> bool:
    """判断是否是单产品规格表"""
    if ws.max_row < 10:
        return False
    if ws.max_column > 6:
        return False
    first_col = [ws.cell(r, 1).value for r in range(2, min(12, ws.max_row))]
    model_count = sum(1 for v in first_col if is_model_code(v))
    return model_count < 2


def classify_format(ws) -> str:
    """判断Excel布局类型"""
    # 1. 参数列表: Model: 或 型号: 标记
    for row_idx in range(1, min(25, ws.max_row + 1)):
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(row_idx, col_idx).value
            if not cell_value:
                continue
            cell_str = str(cell_value).strip().lower()
            if cell_str in ['model:', '型号:', 'item:']:
                return 'param_list'
    
    # 2. PI发票
    for row_idx in range(1, min(25, ws.max_row + 1)):
        cell_value = ws.cell(row_idx, 1).value
        if cell_value:
            cell_lower = str(cell_value).lower()
            if 'proforma invoice' in cell_lower or 'description of goods' in cell_lower:
                return 'invoice'
    
    # 3. 纯价格表: 列少(<8) + 有价格列 (提前检测)
    has_model = False
    has_price = False
    for row_idx in range(1, min(10, ws.max_row + 1)):
        for col_idx in range(1, min(15, ws.max_column + 1)):
            val = ws.cell(row_idx, col_idx).value
            if val:
                val_str = str(val).lower()
                if 'model' in val_str or '型号' in val_str:
                    has_model = True
                if 'price' in val_str or '价格' in val_str or '价' in val_str:
                    has_price = True
    if has_model and has_price and ws.max_column < 10:
        return 'price'
    
    # 4. 横向表: 需要表头行和数据行都满足多列非空
    for row_idx in range(1, min(15, ws.max_row + 1)):
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(row_idx, col_idx).value
            if not cell_value:
                continue
            cell_str = str(cell_value).strip().lower()
            cell_text = str(cell_value).strip()
            if 'model' in cell_str or '型号' in cell_text:
                # 检查表头和数据行是否都满足条件
                non_null_header = count_row_non_null(ws, row_idx, col_idx, col_idx + 8)
                non_null_data = count_row_non_null(ws, row_idx + 1, col_idx, col_idx + 8)
                if non_null_header >= 3 and non_null_data >= 3:
                    # 额外检查: 如果第一行包含"价格"关键词,则判定为price
                    header_has_price = False
                    for c in range(col_idx, min(col_idx + 8, ws.max_column + 1)):
                        h_val = ws.cell(row_idx, c).value
                        if h_val and ('price' in str(h_val).lower() or '价格' in str(h_val)):
                            header_has_price = True
                            break
                    if header_has_price:
                        return 'price'
                    return 'horizontal'
    
    # 5. 单产品规格表
    if is_single_spec_sheet(ws):
        return 'single_spec'
    
    # 6. 默认
    return 'vertical'


# ==================== 第三部分: 模糊列名匹配 ====================

def match_column_fuzzy(header_value) -> str:
    if not header_value:
        return None
    text = str(header_value).lower().strip()
    
    model_keywords = ['型号', 'model', '产品型号', '规格型号', 'item#', '产品编码', 
                   '货号', 'code', 'no.', '编号', 'item no', 'part no']
    for kw in model_keywords:
        if kw in text:
            return 'model'
    
    name_keywords = ['产品名称', '品名', '名称', 'product name', 'description',
                   '商品名称', '品名', 'name', '商品']
    for kw in name_keywords:
        if kw in text:
            return 'name'
    
    spec_keywords = ['规格', 'spec', '参数', '产品描述', 'description', '产品说明',
                   'specification', 'detail', 'details', 'specs']
    for kw in spec_keywords:
        if kw in text:
            return 'spec'
    
    price_keywords = ['价格', 'price', '售价', '报价', '单价', 'unit price',
                     'rmb', 'cny', 'fob', 'exw', '优惠价', '批发价', 'system price']
    for kw in price_keywords:
        if kw in text:
            return 'price'
    
    remark_keywords = ['备注', 'remark', 'note', '注释', '说明']
    for kw in remark_keywords:
        if kw in text:
            return 'remark'
    
    return None


def detect_columns(data_rows: List[Dict], header_labels: List[str] = None) -> Dict[str, int]:
    if not data_rows:
        return {}
    
    col_map = {}
    # 优先使用外部传入的表头文本（如parse_vertical从深层找到的真正表头）
    if header_labels:
        for col_idx, label in enumerate(header_labels):
            matched_type = match_column_fuzzy(str(label))
            if matched_type:
                col_map[matched_type] = col_idx
    else:
        for col_idx, header in enumerate(data_rows[0].keys()):
            matched_type = match_column_fuzzy(header)
            if matched_type:
                col_map[matched_type] = col_idx
    
    for col_idx in range(len(list(data_rows[0].keys()))):
        values = [row.get(list(row.keys())[col_idx]) for row in data_rows[:20]]
        scores = get_column_stats(values)
        
        if 'model' not in col_map and scores.get('model', 0) > 0.5:
            col_map['model'] = col_idx
        if 'price' not in col_map and scores.get('price', 0) > 0.5:
            col_map['price'] = col_idx
        if 'spec' not in col_map and scores.get('spec', 0) > 0.5:
            col_map['spec'] = col_idx
        if 'name' not in col_map and scores.get('name', 0) > 0.4:
            col_map['name'] = col_idx
        if 'remark' not in col_map and scores.get('remark', 0) > 0.4:
            col_map['remark'] = col_idx
    
    return col_map


# ==================== 第四部分: 各类型解析器 ====================

def parse_horizontal(ws) -> pd.DataFrame:
    """横向参数表解析"""
    model_row = None
    for r in range(1, min(15, ws.max_row + 1)):
        for c in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(r, c).value
            if val:
                val_str = str(val).lower().strip()
                if val_str == 'model' or 'model' in val_str or '型号' in str(val):
                    non_null = 0
                    for cc in range(c, min(c + 10, ws.max_column + 1)):
                        if ws.cell(r, cc).value:
                            non_null += 1
                    if non_null >= 2:
                        model_row = r
                        break
        if model_row:
            break
    
    if not model_row:
        return pd.DataFrame()
    
    models = []
    for c in range(2, ws.max_column + 1):
        val = ws.cell(model_row, c).value
        if val and str(val).strip():
            models.append(str(val).strip())
    
    if not models:
        return pd.DataFrame()
    
    result = []
    for col_idx, model in enumerate(models, start=1):
        specs = []
        for r in range(model_row + 1, min(model_row + 50, ws.max_row + 1)):
            name = ws.cell(r, 1).value
            val = ws.cell(r, col_idx).value
            if name and val and str(name).strip():
                # 合并多级路径
                path = str(name).strip()
                # 检查左侧列
                if col_idx > 1:
                    left = ws.cell(r, col_idx - 1).value
                    if left and str(left).strip() and str(left).strip() != str(name).strip():
                        path = f"{left} / {path}"
                specs.append(f"{path}: {val}")
        
        result.append({
            'model': model,
            'name_zh': None,
            'spec_zh': '\n'.join(specs) if specs else '',
            '_row': model_row,
            '_sheet': ws.title,
        })
    
    return pd.DataFrame(result)


def parse_param_list(ws) -> pd.DataFrame:
    """参数列表解析 - 支持 Model: 和 型号: 双标记"""
    rows = []
    current_model = None
    seen_params = set()  # 去重
    
    for r in range(1, min(ws.max_row + 1, 2000)):
        col2 = ws.cell(r, 2).value
        col3 = ws.cell(r, 3).value
        
        # 新产品标记: Model: 或 型:
        if col2 and str(col2).strip().lower() in ['model:', '型号:', 'item:']:
            current_model = col3
            seen_params.clear()
            continue
        
        if current_model and col2 and col3:
            val_str = str(col3).strip()
            if val_str and val_str.lower() not in ['nan', 'none']:
                # 参数去重(取前50字符作为key)
                param_key = str(col2).strip()[:50]
                if param_key not in seen_params:
                    seen_params.add(param_key)
                    rows.append({
                        'model': str(current_model).strip(),
                        'name_zh': str(col2).strip(),
                        'spec_zh': val_str,
                        '_row': r,
                        '_sheet': ws.title,
                    })
    
    if not rows:
        return pd.DataFrame()
    
    result = aggregate_specs(rows)
    first = {}
    for row in rows:
        m = row['model']
        if m not in first:
            first[m] = (row['_row'], row['_sheet'])
    result['_row'] = result['model'].map(lambda m: first.get(m, (None, None))[0])
    result['_sheet'] = result['model'].map(lambda m: first.get(m, (None, None))[1])
    return result


def parse_invoice(ws) -> pd.DataFrame:
    """PI发票解析 - 产品区域截断,多sheet支持"""
    result = []
    
    # Find header row
    header_row = 1
    for r in range(1, min(25, ws.max_row + 1)):
        val = ws.cell(r, 1).value
        if val:
            val_str = str(val).lower()
            if 'description' in val_str or 'goods' in val_str or 'item' in val_str:
                header_row = r
                break
        if r == 3 and ws.cell(r, 1).value:
            header_row = r
            break
    
    for r in range(header_row + 1, min(ws.max_row + 1, 100)):
        col1 = ws.cell(r, 1).value
        col2 = ws.cell(r, 2).value
        
        if not col1 or not str(col1).strip():
            if len(result) >= 2:
                break
            continue
        
        model = str(col1).strip()
        
        # 跳过: TOTAL/AMOUNT行, 条款行(数字+句点), 银行/条款关键词
        model_lower = model.lower()
        if re.match(r'^\d+\.\s', model):
            break
        if 'total' in model_lower or 'amount' in model_lower:
            break
        
        skip_keywords = ['bank', 'notice', 'term', 'delivery', 'port', 'transship', 
                       'handling', 'seller', 'signed', 'payment', 'contact', 'fax',
                       'to:', 'description', 'xuzhou', 'add:', 'e-mail']
        if any(kw in model_lower for kw in skip_keywords):
            continue
        
        # 清理型号
        if ' ' in model and not model[0].isdigit():
            parts = model.split(None, 1)
            if len(parts) > 1:
                name = parts[1][:30]
                model = parts[0][:20]
        elif len(model) > 20:
            model = model[:20]
        
        if model and len(model) > 1:
            result.append({
                'model': model,
                'name_zh': model[:30],
                'spec_zh': str(col2).strip()[:2000] if col2 else '',
                '_row': r,
                '_sheet': ws.title,
            })
    
    if not result:
        return None
    
    return pd.DataFrame(result)


def parse_price_list(ws) -> pd.DataFrame:
    """价格表解析 - 简化版"""
    # 直接找Model列(第2列)和价格列(第4列)
    result = []
    consecutive_empty = 0
    last_model = None
    
    for r in range(2, min(ws.max_row + 1, 50)):
        model = ws.cell(r, 2).value
        price = ws.cell(r, 4).value
        has_spec = any(ws.cell(r, c).value for c in range(6, min(ws.max_column + 1, 15)))
        
        # 纯空行跳过
        if not model and not price and not has_spec:
            consecutive_empty += 1
            if consecutive_empty >= 2 and len(result) >= 2:
                break
            continue
        
        consecutive_empty = 0
        
        # 处理model列：空则继承上一个model（合并单元格场景）
        if model:
            model_str = str(model).strip()
            if model_str and model_str != 'None':
                last_model = model_str
        if not last_model:
            continue
        
        model_str = last_model
        
        # 价格处理
        price_val = None
        if price:
            try:
                price_val = clean_price_value(price)
            except Exception:
                pass
        
        # 如果有价格或没有价格但有型号,都添加
        if price_val or price_val == 0:
            result.append({
                'model': model_str,
                'name_zh': None,
                'spec_zh': '',
                'price_rmb': price_val,
                '_row': r,
                '_sheet': ws.title,
            })
        elif model_str and len(model_str) < 20:
            result.append({
                'model': model_str,
                'name_zh': None,
                'spec_zh': '',
                'price_rmb': None,
                '_row': r,
                '_sheet': ws.title,
            })
    
    return pd.DataFrame(result)


def parse_vertical(ws, images_by_row: Dict[int, str] = None) -> pd.DataFrame:
    """标准纵向解析"""
    all_rows = []
    
    # 如果第1行不是表头（不含model/price等关键词），扫描前30行找真正的表头
    row1_text = ' '.join(str(ws.cell(1, c).value or '') for c in range(1, min(10, ws.max_column + 1))).lower()
    has_header_kw = any(kw in row1_text for kw in ['model', '型号', 'price', '价格', 'spec', '规格', 'description'])
    
    data_start = 2
    data_end = min(ws.max_row + 1, 2000)
    headers = []
    
    if not has_header_kw:
        # 搜索包含 model+price 关键词的表头行（合同/报价单类文件）
        for r in range(2, min(ws.max_row + 1, 35)):
            row_text = ' '.join(str(ws.cell(r, c).value or '') for c in range(1, min(10, ws.max_column + 1))).lower()
            has_m = 'model' in row_text or '型号' in row_text
            has_p = 'price' in row_text or '金额' in row_text or '价格' in row_text or '数量' in row_text
            if has_m or has_p:
                # 找到表头行，取它下面的行作为数据
                data_start = r + 1
                # 用它做表头列名
                for c in range(1, ws.max_column + 1):
                    headers.append(str(ws.cell(r, c).value or f'col_{c}'))
                # 找表尾（total/amount 行）
                for end_r in range(data_start, min(ws.max_row + 1, 200)):
                    end_val = str(ws.cell(end_r, 1).value or '').lower().strip()
                    if 'total' in end_val or 'subtotal' in end_val:
                        data_end = end_r
                        break
                break
    
    if not headers:
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            headers.append(str(val) if val else f'col_{c}')
    
    for r in range(data_start, data_end):
        row_data = {}
        has_data = False
        for c in range(1, len(headers) + 1):
            val = ws.cell(r, c).value
            row_data[f'col_{c}'] = val
            if val:
                has_data = True
        if has_data:
            all_rows.append(row_data)
    
    # 收集底部备注（data_end 之后的行）
    footer_remark = ''
    if data_end < ws.max_row:
        footer_parts = []
        for fr in range(data_end, min(ws.max_row + 1, data_end + 30)):
            fv = str(ws.cell(fr, 1).value or '').strip()
            if fv:
                footer_parts.append(fv)
        if footer_parts:
            footer_remark = '\n'.join(footer_parts)
    
    if not all_rows:
        return pd.DataFrame()
    
    # 如果找到了深层表头，传入表头文本帮助列检测
    col_map = detect_columns(all_rows, header_labels=headers if not has_header_kw else None)
    
    def _col_key(idx):
        return f'col_{idx + 1}' if isinstance(idx, int) else idx
    
    result = []
    model_col = _col_key(col_map.get('model', 'col_2'))
    name_col = _col_key(col_map.get('name', 'col_3'))
    spec_col = _col_key(col_map.get('spec', 'col_4'))
    price_col = _col_key(col_map.get('price')) if col_map.get('price') is not None else None
    remark_col = _col_key(col_map.get('remark')) if col_map.get('remark') is not None else None
    
    for row in all_rows:
        model = row.get(model_col)
        name = row.get(name_col)
        spec = row.get(spec_col)
        price = row.get(price_col) if price_col else None
        
        price_val = None
        if price:
            try:
                price_val = clean_price_value(price)
            except Exception:
                pass
        
        spec_parts = []
        if spec:
            spec_parts.append(str(spec))
        # 提取备注
        remark_val = ''
        if remark_col:
            remark_val = str(row.get(remark_col, '') or '').strip()
        
        for col_name, col_value in row.items():
            if col_name not in [model_col, name_col, spec_col, price_col, remark_col]:
                if col_value and str(col_value).strip():
                    if str(col_value).strip() not in [str(spec) if spec else '', str(name) if name else '', str(model) if model else '']:
                        spec_parts.append(str(col_value).strip())
        
        result.append({
            'model': str(model).strip() if model else '',
            'name_zh': str(name).strip() if name else '',
            'spec_zh': '\n'.join(spec_parts) if spec_parts else '',
            'price_rmb': price_val,
            'remark': remark_val,
            'image_path': images_by_row.get(r, '') if images_by_row else '',
            '_row': r,
            '_sheet': ws.title,
        })
    
    # 后处理：过滤空行 + 型号截断
    filtered = []
    for row in result:
        model_raw = row.get('model', '').strip()
        price = row.get('price_rmb')
        has_model = bool(model_raw) and len(model_raw) >= 2
        has_price = price is not None and (isinstance(price, (int, float)) and price > 0)
        if not has_model and not has_price:
            continue  # 无型号无价格 → 非产品行
        if has_model:
            # 型号截断：只取换行前的第一段，最长30字符
            row['model'] = model_raw.split('\n')[0].strip()[:30]
        # 底部备注填充到没有备注的产品
        if footer_remark and not row.get('remark'):
            row['remark'] = footer_remark
        filtered.append(row)
    
    # 分数过滤：单行评分 < -1 的噪音行删除
    def _score_row(model, price, spec_zh):
        """内联评分，避免跨模块导入依赖"""
        import re as _re
        m = str(model).strip() if model else ''
        p = price if isinstance(price, (int, float)) else None
        has_price = p is not None and p > 0
        is_empty = not m or len(m) < 2
        is_real = (2 <= len(m) <= 30 and bool(_re.search(r'[A-Za-z]', m))
                   and bool(_re.search(r'\d', m)) and ':' not in m and '\uff1a' not in m)
        if is_real and has_price: return 7.0
        if is_real: return 2.0
        if is_empty and not has_price: return -3.0
        if len(m) > 40: return -2.0
        return 1.0
    
    clean = []
    for row in filtered:
        if _score_row(row.get('model',''), row.get('price_rmb'), row.get('spec_zh','')) >= -1.5:
            clean.append(row)
    return pd.DataFrame(clean)


def _parse_single_spec_old(ws) -> pd.DataFrame:
    """单产品规格表 - 合并所有参数到一个产品 (旧版, 接收ws)"""
    specs = []
    model_name = None
    
    for r in range(2, ws.max_row + 1):
        col1 = ws.cell(r, 1).value
        col2 = ws.cell(r, 2).value
        col3 = ws.cell(r, 3).value
        col4 = ws.cell(r, 4).value
        col5 = ws.cell(r, 5).value
        
        # 合并所有非空列作为参数
        vals = [col2, col3, col4, col5]
        for v in vals:
            if v and str(v).strip():
                specs.append(str(v).strip())
        
        # 尝试提取型号
        if not model_name:
            for v in vals:
                if v and is_model_code(v):
                    model_name = str(v).strip()
                    break
    
    if not model_name:
        model_name = f"PRODUCT_{ws.max_row - 1}"
    
    return pd.DataFrame([{
        'model': model_name,
        'name_zh': None,
        'spec_zh': '\n'.join(specs) if specs else '',
        '_row': 2,
        '_sheet': ws.title,
    }])


# ==================== 第五部分: 参数聚合 ====================

def aggregate_specs(rows: List[Dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    
    model_dict = {}
    
    for row in rows:
        model = row.get('model', '')
        if not model:
            continue
        
        model = str(model).strip()
        if model not in model_dict:
            model_dict[model] = {
                'model': model,
                'name_zh': row.get('name_zh'),
                'specs': [],
                'price_rmb': row.get('price_rmb')
            }
        
        spec = row.get('spec_zh')
        if spec:
            model_dict[model]['specs'].append(spec)
    
    result = []
    for model, data in model_dict.items():
        specs = data['specs']
        unique_specs = []
        seen = set()
        for s in specs:
            s_key = s[:50] if s else ''
            if s_key not in seen:
                seen.add(s_key)
                unique_specs.append(s)
        
        result.append({
            'model': model,
            'name_zh': data.get('name_zh'),
            'spec_zh': '; '.join(unique_specs[:50]),
            'price_rmb': data.get('price_rmb')
        })
    
    return pd.DataFrame(result)


# ==================== 导入新解析器 ====================

try:
    from src.parsers import (
        parse_param_price,
        parse_invoice,
        parse_price_table,
        parse_single_spec,
        format_spec_spec
    )
    HAS_NEW_PARSERS = True
except ImportError:
    HAS_NEW_PARSERS = False


# ==================== 第六部分: 主入口 ====================

def parse_excel_v3(file_path: str) -> Optional[pd.DataFrame]:
    if not os.path.exists(file_path):
        return None
    
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_dir = os.path.dirname(file_path)
    
    # Extract images first
    images_by_row = {}
    if HAS_NEW_PARSERS:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True, read_only=False)
            ws = wb.active
            
            # Pass file_path to extraction
            images_by_row = extract_images_from_worksheet(ws, output_dir, base_name, file_path)
            wb.close()
        except Exception as e:
            print(f'Image extraction error: {e}')
    
    # 尝试使用新解析器
    if HAS_NEW_PARSERS:
        try:
            # 检测文件类型并使用对应解析器
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            
            layout = classify_format(ws)
            _product_codes = 0
            if layout == 'single_spec':
                import re as _re
                # 扫描 A~D 列（1~4），查找短字母+数字组合的产品代码
                for _c in range(1, 5):
                    for _r in range(2, min(ws.max_row + 1, 40)):
                        _v = str(ws.cell(_r, _c).value or '').strip()
                        if 2 <= len(_v) <= 40 and _re.search(r'[A-Za-z]', _v) and _re.search(r'\d', _v):
                            _product_codes += 1
                            if _product_codes >= 3:
                                break
                    if _product_codes >= 3:
                        break
            wb.close()
            
            # 使用新的专用解析器
            if layout == 'param_list':
                df = parse_param_price(file_path)
                if df is not None and len(df) > 0:
                    df['_source_file'] = os.path.basename(file_path)
                    # 格式化spec
                    if 'spec_zh' in df.columns:
                        df['spec_zh'] = df['spec_zh'].apply(
                            lambda x: format_spec_spec(str(x)) if pd.notna(x) and x else ''
                        )
                    return df
            
            elif layout == 'invoice':
                df = parse_invoice(file_path)
                if df is not None and len(df) > 0:
                    df['_source_file'] = os.path.basename(file_path)
                    if 'spec_zh' in df.columns:
                        df['spec_zh'] = df['spec_zh'].apply(
                            lambda x: format_spec_spec(str(x)) if pd.notna(x) and x else ''
                        )
                    return df
            
            elif layout == 'price':
                df = parse_price_table(file_path)
                if df is not None and len(df) > 0:
                    df['_source_file'] = os.path.basename(file_path)
                    # Add image_path column (empty for now - images extracted above)
                    if 'image_path' not in df.columns:
                        df['image_path'] = ''
                    if 'spec_zh' in df.columns:
                        df['spec_zh'] = df['spec_zh'].apply(
                            lambda x: format_spec_spec(str(x)) if pd.notna(x) and x else ''
                        )
                    return df
            
            elif layout == 'single_spec':
                # 已在 wb.close() 之前计算 _product_codes，直接使用
                if _product_codes >= 3:
                    df = None  # 回退到 fallback 解析
                else:
                    df = parse_single_spec(file_path)
                if df is not None and len(df) > 0:
                    df['_source_file'] = os.path.basename(file_path)
                    if 'spec_zh' in df.columns:
                        df['spec_zh'] = df['spec_zh'].apply(
                            lambda x: format_spec_spec(str(x)) if pd.notna(x) and x else ''
                        )
                    return df
        
        except Exception as e:
            print(f"New parser error, falling back: {e}")
    
    # Fallback: 使用原有解析器
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return None
    
    all_dfs = []
    for sheet in wb:
        ws = sheet
        
        layout = classify_format(ws)
        
        if layout == 'horizontal':
            df = parse_horizontal(ws)
        elif layout == 'param_list':
            df = parse_param_list(ws)
        elif layout == 'invoice':
            df = parse_invoice(ws)
        elif layout == 'price':
            df = parse_price_list(ws)
        elif layout == 'single_spec':
            df = _parse_single_spec_old(ws)
        else:
            df = parse_vertical(ws, images_by_row)
        
        if df is not None and len(df) > 0:
            df['_source_file'] = os.path.basename(file_path)
            df['_sheet'] = ws.title
            all_dfs.append(df)
    
    wb.close()
    
    if not all_dfs:
        return pd.DataFrame()
    
    return pd.concat(all_dfs, ignore_index=True)


def parse_excel(file_path: str) -> Optional[pd.DataFrame]:
    df = parse_excel_v3(file_path)
    if df is None or df.empty:
        return df

    df['_source_file'] = os.path.basename(file_path)

    if '_row' in df.columns:
        try:
            from src.core.image import match_images_to_products
            df = match_images_to_products(df, file_path)
        except Exception as e:
            import logging
            logging.warning(f"Image matching failed: {e}")

    if 'image_path' not in df.columns:
        df['image_path'] = ''
    if '_image_path' in df.columns:
        df['image_path'] = df['_image_path']

    return df