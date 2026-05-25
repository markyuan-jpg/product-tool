# -*- coding: utf-8 -*-
"""
single_spec_parser.py - 单产品规格+复合价格拆分
解决 BAOSHIMA 的复合价格拆分问题
"""
import re
import pandas as pd
from typing import Optional, List, Dict
from openpyxl import load_workbook
from ..utils.price import clean_price_value


from ..price_config import get_industry_config, classify_price, validate_price, detect_industry, resolve_priority


# 复合价格模式: "EV: CNY 4980 / Battery: CNY 2530 / Charger: CNY 230"
COMPOSITE_PRICE_PATTERN = r'(EV|电池|充电器|Controller|控制器)?[:\s]*([A-Z]{3})?\s*([\d,]+)'


def parse_composite_price(text: str, industry_config: Dict = None) -> List[Dict]:
    """
    拆分复合价格
    输入: "EV: CNY 4980 / Battery: CNY 2530 / Charger: CNY 230"
    输出: [
        {'item': 'EV', 'price': 4980.0},
        {'item': 'Battery', 'price': 2530.0},
        {'item': 'Charger', 'price': 230.0}
    ]
    """
    if not text:
        return []
    
    items = []
    parts = str(text).replace('\n', '/').split('/')
    
    for part in parts:
        part = part.strip()
        if not part:
            continue
        
        # 匹配 item: price 格式
        # 例如: "EV: CNY 4980" 或 "Battery: CNY 2530"
        match = re.match(r'([A-Za-z]+)\s*[:\s]*([A-Z]{3})?\s*([\d,]+)', part)
        if match:
            item_name = match.group(1)
            price_str = match.group(3)
            price = clean_price_value(price_str)
            if price:
                item_entry = {'item': item_name, 'price': price}
                # 使用配置分类
                if industry_config:
                    ptype, label, _ = classify_price(item_name, industry_config)
                    item_entry['type'] = ptype
                    item_entry['label'] = label
                items.append(item_entry)
        else:
            # 尝试提取数字
            nums = re.findall(r'[\d,]+', part)
            if nums:
                price = clean_price_value(nums[0])
                if price:
                    part_lower = part.lower()
                    # 使用配置推断item名称
                    classified = False
                    if industry_config:
                        # Try classify with all secondary keywords
                        for entry in industry_config.get('secondary_keywords', []):
                            for kw in entry.get('keywords', []):
                                if kw.lower() in part_lower:
                                    items.append({'item': entry['type'], 'price': price,
                                                  'type': entry['type'], 'label': entry['label']})
                                    classified = True
                                    break
                            if classified:
                                break
                    if not classified:
                        if 'ev' in part_lower:
                            items.append({'item': 'EV', 'price': price})
                        elif 'battery' in part_lower or '电池' in part:
                            items.append({'item': 'Battery', 'price': price})
                        elif 'charger' in part_lower or '充电' in part:
                            items.append({'item': 'Charger', 'price': price})
                        else:
                            items.append({'item': 'Other', 'price': price})
    
    return items


def detect_model_from_spec(spec_text: str) -> Optional[str]:
    """从规格文本中提取型号"""
    if not spec_text:
        return None
    
    # 常见型号模式: XF-1, XP, G5000, YL1200
    patterns = [
        r'^([A-Z]{1,3}[\-\d]+)',  # XF-1, XP, G5000
        r'^([A-Z]\d{3,})',      # YL1200
    ]
    
    for pattern in patterns:
        match = re.search(pattern, str(spec_text))
        if match:
            return match.group(1)
    
    return None


def parse_single_spec(file_path: str, industry: str = None) -> pd.DataFrame:
    """
    解析 BAOSHIMA/单产品规格表
    
    特点:
    - 单个产品规格表(多个section)
    - 复合价格结构
    
    格式支持:
    - BAOSHIMA XF-1: 规格在C4:C5, section headers在C3, 型号在C5
    - 通用格式: 规格在C2:C3
    
    参数:
        file_path: Excel 文件路径
        industry: 行业标识，None 则自动检测
        
    返回: DataFrame [model, spec_zh, price_rmb]
    """
    # 检测行业
    ic = get_industry_config(industry, file_path)
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    
    current_model = None
    current_sections = {}
    current_section = 'General'
    composite_price_text = None
    
    # 检测文件类型: BAOSHIMA XF-1 格式 vs 通用格式
    is_baoshima = False
    for r in range(1, min(ws.max_row + 1, 6)):
        c2 = str(ws.cell(r, 2).value or '').strip()
        c3 = str(ws.cell(r, 3).value or '').strip()
        if c2 in ('Appearance',) and c3 in ('Specifications',):
            is_baoshima = True
            break
        if c2 in ('Model',) and c3:
            is_baoshima = True
            break
    
    for r in range(2, min(ws.max_row + 1, 50)):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        c5 = ws.cell(r, 5).value
        
        c2_str = str(c2).strip() if c2 else ''
        c3_str = str(c3).strip() if c3 else ''
        c4_str = str(c4).strip() if c4 else ''
        c5_str = str(c5).strip() if c5 else ''
        
        if is_baoshima:
            # BAOSHIMA 格式: section在C3, 参数在C4:C5, 型号/价格在C5
            section_keywords = ['power specifications', 'electrical specifications', 
                               'bodywork specification', 'packaging specifications',
                               '动力规格', '电气规格', '车身规格', '包装规格']
            if c3_str.lower() in section_keywords:
                current_section = c3_str
                continue
            
            # 检测型号: "Model   XF-1" in C5
            if not current_model:
                model_from_c5 = detect_model_from_spec(c5_str)
                if model_from_c5:
                    current_model = model_from_c5
                # Also check C5 directly for "Model" keyword
                if not current_model and 'model' in c5_str.lower():
                    parts = c5_str.replace('Model', '').replace('model', '').strip().split()
                    if parts:
                        current_model = parts[0]
            
            # 检测价格: "EXW PRICE" in C3, price in C5
            if c3_str and ('price' in c3_str.lower() or '价格' in c3_str or 'exw' in c3_str.lower()):
                if c5_str:
                    composite_price_text = c5_str
                continue
            
            # 收集参数行 (C4=参数名, C5=参数值)
            if c4_str and c5_str:
                if current_section not in current_sections:
                    current_sections[current_section] = []
                current_sections[current_section].append(f"{c4_str}: {c5_str}")
        else:
            # 通用格式: section在C2, 参数在C2:C3
            if c2_str in ['Power Spec', 'Electrical Spec', 'Bodywork Spec', 'Packaging Spec',
                        '动力规格', '电气规格', '车身规格', '包装规格']:
                current_section = c2_str
                continue
            
            # 检测价格
            if 'price' in c2_str.lower() or '价格' in c2_str:
                if c3_str or c5_str:
                    price_line = c3_str or c5_str
                    if '/' in price_line or 'CNY' in price_line:
                        composite_price_text = price_line
                    else:
                        price = clean_price_value(c3) or clean_price_value(c5)
                        if price and not current_model:
                            for rr in range(1, min(r, 10)):
                                m = ws.cell(rr, 1).value
                                if m and detect_model_from_spec(str(m)):
                                    current_model = detect_model_from_spec(str(m))
            
            # 提取型号
            if not current_model:
                model = ws.cell(r, 1).value
                if model:
                    detected = detect_model_from_spec(str(model))
                    if detected:
                        current_model = detected
            
            # 收集参数行
            if c2_str and c3_str and c2_str not in ['Power Spec', 'Electrical Spec', 
                                                  'Bodywork Spec', 'Packaging Spec',
                                                  '动力规格', '电气规格', '车身规格', '包装规格']:
                if current_section not in current_sections:
                    current_sections[current_section] = []
                current_sections[current_section].append(f"{c2_str}: {c3_str}")
    
    if not current_model:
        import os
        base = os.path.basename(file_path)
        # Remove common prefixes from filename to get shorter model
        name = base.replace('BAOSHIMA_Quotation_', '').replace('_EXW_SKD_260314', '').split('.')[0]
        current_model = name[:20]
    
    spec_parts = []
    for section, specs in current_sections.items():
        spec_parts.append('\n'.join(specs))
    spec_zh = '\n'.join(spec_parts)
    
    prices = []
    prices_dict = {}
    if composite_price_text:
        from ..utils.price import classify_composite_price
        classified = classify_composite_price(composite_price_text, ic)
        prices_dict = classified['prices']
        if classified['primary_price']:
            # Use primary price + append secondary to spec_zh
            total_price = classified['primary_price']
            if classified['spec_lines']:
                spec_zh += '\n\n[Optional Accessories]\n' + '\n'.join(classified['spec_lines'])
        else:
            # Fallback to old behavior
            prices = parse_composite_price(composite_price_text)
            total_price = sum(p.get('price', 0) for p in prices) if prices else None
    
    result = []
    if composite_price_text:
        price_items = ' + '.join([f"{p.get('item','')}={p.get('price',0)}" for p in prices]) if prices else ''
        result.append({
            'model': current_model,
            'name_zh': current_model,
            'spec_zh': spec_zh,
            'spec_detail': price_items,
            'price_rmb': total_price,
            'prices': prices_dict,
            '_row': 2,
            '_sheet': ws.title,
        })
    else:
        result.append({
            'model': current_model,
            'name_zh': current_model,
            'spec_zh': spec_zh,
            'price_item': 'EV',
            'price_rmb': None,
            '_row': 2,
            '_sheet': ws.title,
        })
    
    wb.close()
    
    df = pd.DataFrame(result)
    return df


def parse(file_path: str, industry: str = None) -> pd.DataFrame:
    """入口"""
    return parse_single_spec(file_path, industry)