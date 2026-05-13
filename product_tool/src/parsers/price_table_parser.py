# -*- coding: utf-8 -*-
"""
price_table_parser.py - 价格表全规格列提取
解决 车型价格表 只提取价格+漏规格列问题
"""
import re
import pandas as pd
from typing import Optional, List
from openpyxl import load_workbook
from ..utils.price import clean_price_value


# 标准规格列名映射
SPEC_COLUMNS = {
    2: 'model',           # 型号
    3: 'picture',        # 图片
    4: 'price',         # 出厂价格
    5: 'note',         # 备注
    6: 'motor_type',    # 电机类型
    7: 'motor_power',   # 电机功率
    8: 'gear_type',    # 齿轮类型
    9: 'controller',  # 控制器类型
    10: 'current',     # 电流
    11: 'speed',       # 限速
    12: 'warranty'     # 电机质保
}


def parse_price_table(file_path: str) -> pd.DataFrame:
    """
    解析 车型价格表
    
    提取:
    - Col 2: 型号
    - Col 4: 价格
    - Col 6-12: 所有规格列
    
    返回: DataFrame [model, spec_zh, price_rmb, motor_type, motor_power, ...]
    """
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    
    result = []
    last_model = None
    
    for r in range(2, ws.max_row + 1):
        col_model = ws.cell(r, 2).value
        col_price = ws.cell(r, 4).value
        
        # 跳过纯空行（无价格也无备注）
        if not col_price and not any(ws.cell(r, c).value for c in range(6, min(ws.max_column + 1, 15))):
            continue
        
        # 跳过非价格行（如备注行、电池说明行）
        if col_price and isinstance(col_price, str):
            price_str = str(col_price).strip()
            # 跳过以字母开头的（如"72V18A电池480元" - 不是价格）
            if price_str and not price_str[0].isdigit():
                continue
            # 跳过混合中文的非数字文本（如"备注：不含电池"）
            if re.search(r'[\u4e00-\u9fff]', price_str) and not re.match(r'^[\d,.\s\-~]+$', price_str):
                first_num = re.search(r'[\d,]+(?:\.\d+)?', price_str)
                if not first_num or first_num.group() != price_str.split('\n')[0].split('(')[0].strip().replace(',', ''):
                    continue
        
        # 处理model列：空则继承上一个model（合并单元格场景）
        if col_model:
            raw = str(col_model)
            # 清理特殊字符：移除前导的非字母数字/CJK字符
            raw = re.sub(r'^[^a-zA-Z0-9_\u4e00-\u9fff]+', '', raw.strip())
            # 移除内嵌换行符
            raw = raw.replace('\n', '').replace('\r', '').replace('\u3000', '').strip()
            if raw and raw != 'None' and raw.lower() not in ['model', '型号', '序号', 'no.']:
                last_model = raw
        if not last_model:
            continue
        
        model = last_model
        
        # 价格
        price = clean_price_value(col_price)
        
        # 收集规格列
        spec_parts = []
        for col_idx in range(6, min(ws.max_column + 1, 15)):
            val = ws.cell(r, col_idx).value
            if val and str(val).strip():
                col_name = SPEC_COLUMNS.get(col_idx, f'col_{col_idx}')
                spec_parts.append(f"{col_name}: {val}")
        
        # 构建spec_zh
        spec_zh = '; '.join(spec_parts) if spec_parts else ''
        
        record = {
            'model': model,
            'spec_zh': spec_zh,
            'price_rmb': price,
            '_row': r,
            '_sheet': ws.title,
        }
        
        # 添加各个规格字段
        for col_idx in range(6, min(ws.max_column + 1, 15)):
            val = ws.cell(r, col_idx).value
            if val:
                col_name = SPEC_COLUMNS.get(col_idx, f'spec_{col_idx}')
                record[col_name] = val
        
        result.append(record)
    
    wb.close()
    
    if not result:
        return pd.DataFrame()
    
    df = pd.DataFrame(result)
    
    try:
        from src.core.image import match_images_to_products
        df = match_images_to_products(df, file_path)
    except Exception:
        pass
    if '_image_path' not in df.columns:
        df['_image_path'] = ''
    
    # 过滤超长model
    df = df[df['model'].str.len() < 30]
    
    return df


def parse(file_path: str) -> pd.DataFrame:
    """入口"""
    return parse_price_table(file_path)