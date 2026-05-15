# -*- coding: utf-8 -*-
"""
invoice_parser.py - PI发票全字段精确提取
解决 SONLINK PI 的 QTY/单价/总价漏提取问题
"""
import re
import pandas as pd
from typing import Optional, List, Dict
from openpyxl import load_workbook
from ..utils.price import clean_price_value


# 跳过关键词
SKIP_KEYWORDS = [
    'bank', 'notice', 'term', 'delivery', 'port', 'transship',
    'handling', 'seller', 'signed', 'payment', 'contact', 'fax',
    'to:', 'description', 'xuzhou', 'add:', 'e-mail', 'email',
    'attn', 'reference', 'ref:', 'date:', 'validity',
    'pi:', 'invoice', 'proforma'
]


def _detect_invoice_currency(ws, header_row: int, col_price: int) -> str:
    """检测发票币种：检查价格列头是否含 USD/FOB 等出口术语"""
    header = str(ws.cell(header_row, col_price).value or '').lower()
    if any(kw in header for kw in ['usd', '$', 'fob', 'cif', 'cfr', 'dap', 'ddp']):
        return 'USD'
    return 'RMB'


def parse_invoice_sheet(ws) -> pd.DataFrame:
    """解析单个sheet"""
    result = []
    header_row = 1
    
    for r in range(1, min(25, ws.max_row + 1)):
        val = ws.cell(r, 1).value
        if val:
            val_str = str(val).lower()
            if 'description' in val_str or 'item' in val_str:
                header_row = r
                break
        # r==3 fallback: 仅当未找到更好表头时使用，但不阻断后续匹配
        if r == 3 and ws.cell(r, 1).value and header_row == 1:
            header_row = r
    
    col_desc = 1   # A
    col_spec = 3   # C  
    col_qty = 4    # D
    col_price = 5   # E
    col_total = 6   # F
    currency = _detect_invoice_currency(ws, header_row, col_price)
    
    if ws.max_column >= 6:
        # 检查第一行的列名
        for c in range(1, 10):
            val = ws.cell(header_row, c).value
            if val:
                val_lower = str(val).lower()
                if 'quantity' in val_lower or 'qty' in val_lower:
                    col_qty = c
                elif 'unit' in val_lower and 'price' in val_lower:
                    col_price = c
                elif 'total' in val_lower and 'amount' in val_lower:
                    col_total = c
    
    # 解析数据行
    product_count = 0
    for r in range(header_row + 1, min(ws.max_row + 1, 200)):
        col1 = ws.cell(r, col_desc).value
        col2 = ws.cell(r, col_spec).value if col_spec <= ws.max_column else None
        col3 = ws.cell(r, col_qty).value if col_qty <= ws.max_column else None
        col4 = ws.cell(r, col_price).value if col_price <= ws.max_column else None
        col5 = ws.cell(r, col_total).value if col_total <= ws.max_column else None
        
        # 空行检查
        if not col1 or not str(col1).strip():
            if product_count >= 2:
                break
            continue
        
        model_raw = str(col1).strip()
        
        # 跳过规则
        model_lower = model_raw.lower()
        
        # 条款行(数字+.+空格)
        if re.match(r'^\d+\.\s', model_raw):
            continue
        
        # TOTAL行
        if 'total' in model_lower or 'amount' in model_lower:
            continue
        
        # 跳过关键词
        if any(kw in model_lower for kw in SKIP_KEYWORDS):
            continue
        
        # 清理型号
        if ' ' in model_raw and not model_raw[0].isdigit():
            parts = model_raw.split(None, 1)
            name = parts[1][:50] if len(parts) > 1 else ''
            model = parts[0][:20]
        elif len(model_raw) > 25:
            model = model_raw[:20]
            name = model_raw[:50]
        else:
            model = model_raw
            name = model_raw[:50]
        
        if not model or len(model) < 2:
            continue
        
        # 提取数量
        qty = None
        if col3:
            try:
                qty = int(float(str(col3).replace(',', '')))
            except (ValueError, TypeError):
                pass
        
        # 提取单价(FOB价)
        unit_price = clean_price_value(col4)
        if not unit_price and col5 and qty:
            # 如果单价为空但有总价和数量,计算单价
            total = clean_price_value(col5)
            if total and qty and qty > 0:
                unit_price = total / qty
        
        # 总金额
        total_amount = clean_price_value(col5)
        
        # 规格描述
        spec_zh = str(col2).strip()[:500] if col2 else ''
        
        result.append({
            'model': model,
            'name_zh': name,
            'spec_zh': spec_zh,
            'price_rmb': unit_price,
            'currency': currency,
            'quantity': qty,
            'unit_price': unit_price,
            'total_amount': total_amount,
            '_row': r,
        })
        product_count += 1
    
    return pd.DataFrame(result)


def parse_invoice(file_path: str) -> pd.DataFrame:
    """主入口 - 解析SONLINK PI
    
    支持多sheet
    返回: DataFrame [model, spec_zh, quantity, unit_price, total_amount]
    """
    wb = load_workbook(file_path, data_only=True, read_only=True)
    all_dfs = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 检查是否是PI格式
        is_invoice = False
        for r in range(1, min(20, ws.max_row + 1)):
            val = ws.cell(r, 1).value
            if val:
                val_lower = str(val).lower()
                if 'proforma' in val_lower or 'invoice' in val_lower or 'description' in val_lower:
                    is_invoice = True
                    break
        
        if not is_invoice:
            # 也检查是否至少有 Description 列
            if ws.max_column > 5:
                is_invoice = True
        
        if is_invoice:
            df = parse_invoice_sheet(ws)
            if df is not None and len(df) > 0:
                df['_sheet'] = sheet_name
                df['_source_file'] = file_path
                all_dfs.append(df)
    
    wb.close()
    
    if not all_dfs:
        return pd.DataFrame()
    
    df = pd.concat(all_dfs, ignore_index=True)
    try:
        from src.core.image import match_images_to_products
        df = match_images_to_products(df, file_path)
    except Exception:
        pass
    if '_image_path' not in df.columns:
        df['_image_path'] = ''
    return df


def parse(file_path: str) -> pd.DataFrame:
    """兼容入口"""
    return parse_invoice(file_path)