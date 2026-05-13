# -*- coding: utf-8 -*-
"""
param_price_parser.py - 参数列表+价格精确提取 V4.1
修复: 字符串价格格式提取 + 行号记录 + 图片匹配
"""
import re
import os
import pandas as pd
from openpyxl import load_workbook
from ..utils.price import clean_price_value
from ..price_config import get_industry_config, match_keyword, validate_price, detect_industry


def is_price_marker(text: str, industry_config: dict = None) -> bool:
    """检测是否是价格标记，使用配置中的关键词列表"""
    if not text:
        return False
    text_lower = str(text).lower().strip()
    if not text_lower:
        return False

    # 直接精确匹配格式 "关键词:"
    direct_pattern = re.compile(r'^(价格|售价|报价|单价|裸车价格|整车价格|CKD散件价格|出厂价格|price|exw)\s*:', re.I)
    if direct_pattern.search(text_lower):
        return True

    # 使用配置中的 main_price_keywords + secondary_price_keywords 联合匹配
    if industry_config:
        pd_config = industry_config.get('price_detection', {})
        all_price_kw = list(pd_config.get('main_price_keywords', []))
        all_price_kw.extend(pd_config.get('secondary_price_keywords', []))
        if all_price_kw:
            return match_keyword(text_lower, all_price_kw)

    # 若没有配置或配置为空，使用旧的硬编码关键词（降级兼容）
    markers = ['price', '价格', 'rmb', 'cny', 'fob', 'exw']
    if any(m in text_lower for m in markers):
        return len(text_lower) < 30
    return False


def extract_price_from_string(value, industry_config: dict = None) -> float:
    """
    从字符串提取价格数字
    支持格式: RMB1325.00, 裸车价2480.00元, 价格:1610.00元(不含税), CKD散装价格1270.00元
    使用配置中的 value_range 验证价格有效性。
    """
    if value is None:
        return None
    
    val_str = str(value).strip()
    if not val_str:
        return None
    
    # 计算法1: 直接清洗(原值是纯数字)
    try:
        if isinstance(value, (int, float)):
            if validate_price(float(value), 'CNY', industry_config):
                return float(value)
    except Exception:
        pass
    
    # 计算法2: 正则提取数字(支持各种字符串格式)
    # 匹配: 1325.00, 1325, RMB1325.00, 2480.00元, 1610.00元(不含税)
    match = re.search(r'(\d+[\d,]*\.?\d*)', val_str)
    if match:
        try:
            num_str = match.group(1).replace(',', '')
            price = float(num_str)
            if validate_price(price, 'CNY', industry_config):
                return price
        except (ValueError, TypeError):
            pass
    
    return None


from ..dedup_engine import dedup_dataframe

def parse_param_price(file_path: str, industry: str = None) -> pd.DataFrame:
    """解析 param_price.xlsx
    
    结构:
    - Model:/型号: (Column 2)
    - 参数行 (Column 2:参数名, Column 3:参数值)
    - 价格行 (Column 2-5, 字符串格式)
    
    参数:
        file_path: Excel 文件路径
        industry: 行业标识，None 则自动检测
    """
    ic = get_industry_config(industry, file_path)
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    
    current_model = None
    current_specs = []
    result = []
    pending_price = None
    current_row = None
    
    for r in range(1, min(ws.max_row + 1, 3000)):
        col2 = ws.cell(r, 2).value
        col3 = ws.cell(r, 3).value
        col4 = ws.cell(r, 4).value
        col5 = ws.cell(r, 5).value
        
        col2_str = str(col2).strip() if col2 else ''
        
        # Model:/型号: 标记
        if col2_str.lower() in ['model:', '型号:', 'item:']:
            if current_model and current_specs:
                result.append({
                    'model': current_model,
                    'spec_zh': '; '.join(current_specs[:50]),
                    'price_rmb': pending_price,
                    '_row': current_row,
                    '_source': file_path
                })
            current_model = str(col3).strip() if col3 else ''
            current_specs = []
            pending_price = None
            current_row = r
            continue
        
        # 价格行: 扫描 Column 2-5
        if current_model:
            for ci, cell_val in [(2, col2), (3, col3), (4, col4), (5, col5)]:
                if is_price_marker(str(cell_val) if cell_val else '', ic):
                    price = extract_price_from_string(cell_val, ic)
                    if price:
                        pending_price = price
                        break
                    # 递归搜索col2, 也查找同行其他列
                    if ci == 2:
                        for ci2, cv2 in [(3, col3), (4, col4), (5, col5)]:
                            price = extract_price_from_string(cv2, ic)
                            if price:
                                pending_price = price
                                break
                elif col2_str in ['', ' ']:
                    # Column 2为空时,检查 Column 4是否是价格
                    price = extract_price_from_string(col4, ic)
                    if price:
                        pending_price = price
            
            # 参数收集
            if col2_str and col3 and not is_price_marker(col2_str, ic):
                current_specs.append(f"{col2_str}: {col3}")
    
    # 处理最后一个
    if current_model and current_specs:
        result.append({
            'model': current_model,
            'spec_zh': '; '.join(current_specs[:50]),
            'price_rmb': pending_price,
            '_row': current_row,
            '_source': file_path
        })
    
    wb.close()
    
    if not result:
        return pd.DataFrame()
    
    df = pd.DataFrame(result)
    df = df[df['model'].str.len() < 30]
    
    # 提取并关联图片
    df = match_images_to_products(df, file_path)
    
    return df


def parse(file_path: str, industry: str = None) -> pd.DataFrame:
    return parse_param_price(file_path, industry)


def parse_table(file_path: str) -> pd.DataFrame:
    """解析表格型布局 - 每个型号占一列，多行配置
    
    结构 (新能源电动车价格表):
    - Column 2: 车型/Model 表头
    - Column 4: 出厂价格
    - Columns 6-12: 电机类型/功率/齿轮/控制器/限速/质保
    - 空model行: 继承上一行型号（多配置）
    - "--" model行: 继承上一行型号（多配置）
    """
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    
    result = []
    last_model = None
    last_specs = []
    
    for r in range(1, min(ws.max_row + 1, 3000)):
        col2 = ws.cell(r, 2).value
        col4 = ws.cell(r, 4).value
        col6 = ws.cell(r, 6).value
        col7 = ws.cell(r, 7).value
        col8 = ws.cell(r, 8).value
        col9 = ws.cell(r, 9).value
        col10 = ws.cell(r, 10).value
        col11 = ws.cell(r, 11).value
        col12 = ws.cell(r, 12).value
        
        col2_str = str(col2).strip() if col2 else ''
        col2_lower = col2_str.lower()
        
        if col2_lower in ['model:', '型号:', 'item:', '车型 model', '车型\nmodel']:
            continue
        
        if col2_lower == '' and not col4:
            continue
        
        is_header = any(x in col2_lower for x in ['电机类型', 'motor type', 'serial', 'no.', '序号'])
        if is_header:
            continue
        
        model = None
        excel_row = r  # 记录Excel行号，用于图片匹配
        if col2_str and col2_str not in ['--', '-', '/']:
            model = col2_str
            last_model = model
        elif col2_str in ['--', '-', '/', '']:
            model = last_model
        
        price = extract_price_from_string(col4)
        
        if model and len(model) < 30 and price and price >= 100:
            specs = []
            if col6: specs.append(f"电机类型: {col6}")
            if col7: specs.append(f"电机功率: {col7}")
            if col8: specs.append(f"齿轮类型: {col8}")
            if col9: specs.append(f"控制器类型: {col9}")
            if col10: specs.append(f"电流: {col10}")
            if col11: specs.append(f"限速: {col11}")
            if col12: specs.append(f"质保: {col12}")
            
            result.append({
                'model': model,
                'spec_zh': '; '.join(specs[:50]),
                'price_rmb': price,
                '_source': file_path,
                '_row': excel_row
            })
    
    wb.close()
    
    if not result:
        return pd.DataFrame()
    
    df = pd.DataFrame(result)
    
    df = match_images_to_products(df, file_path)
    
    return df


def match_images_to_products(df: pd.DataFrame, source_file: str) -> pd.DataFrame:
    """根据行号匹配图片到产品 (统一接口, 为兼容保留)"""
    if df.empty:
        return df
    try:
        from src.core.image import match_images_to_products as unified_match
        return unified_match(df, source_file)
    except Exception:
        return df