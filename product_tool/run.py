"""
run.py - CLI入口
通用Excel解析器命令行工具
"""
import os
import sys
import argparse
import glob
import pandas as pd

# 添加src到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from src.core.excel_parser_v3 import parse_excel_v3
from src.parsers import (
    parse_param_price
)
from src.parsers.param_price_parser import parse_table
from src.parsers.invoice_parser import parse_invoice
from src.parsers.price_table_parser import parse_price_table
from src.parsers.single_spec_parser import parse_single_spec
from src.output.quotation_excel import create_quotation
from src.utils.translator import translate_text, bilingual_text
from src.utils.categorizer import load_categories, categorize_data
from src.core.pdf_parser import extract_products_from_pdf_v2
from src.core.doc_parser import extract_products_from_docx

try:
    from src.output.pi_generator import generate_pi
    from src.packing.generator import generate_packing_list, generate_commercial_invoice
    from src.config import DEFAULT_SELLER_INFO
    _HAS_PI_DEPS = True
except ImportError:
    _HAS_PI_DEPS = False


def parse_args():
    parser = argparse.ArgumentParser(
        description='Product Tool - Excel Parser & Quotation Generator',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python run.py --input ./data --output ./output
  python run.py --input ./data --output ./output --lang bilingual
  python run.py --input ./data/myfile.xlsx --output ./result.xlsx --supplier "SONLINK"
  python run.py --input ./data --output ./output --output-group-by sort
  python run.py --input ./data --output ./output --category categories.json --no-category
        """
    )
    
    parser.add_argument(
        '--input', '-i',
        required=True,
        help='输入文件或文件夹路径'
    )
    
    parser.add_argument(
        '--output', '-o',
        required=True,
        help='输出文件或文件夹路径'
    )
    
    parser.add_argument(
        '--lang', '-l',
        choices=['chinese', 'english', 'bilingual'],
        default='bilingual',
        help='输出语言 (default: bilingual)'
    )
    
    parser.add_argument(
        '--output-group-by',
        choices=['sort', 'sheet', 'file', 'category'],
        default='sort',
        help='输出分组方式 (default: sort)'
    )
    
    parser.add_argument(
        '--category',
        help='分类配置文件路径 (JSON)'
    )
    
    parser.add_argument(
        '--no-category',
        action='store_true',
        help='禁用自动分类'
    )
    
    parser.add_argument(
        '--supplier',
        default='',
        help='供应商名称'
    )
    
    parser.add_argument(
        '--quotation',
        action='store_true',
        help='输出为报价单格式'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='显示详细信息'
    )
    
    parser.add_argument(
        '--save-to-db',
        action='store_true',
        help='解析后保存到产品库'
    )
    
    parser.add_argument(
        '--user-id',
        default='local',
        help='产品库用户ID (default: local)'
    )
    
    parser.add_argument(
        '--merge',
        action='store_true',
        help='合并去重后保存到产品库'
    )
    
    parser.add_argument('--pi', action='store_true', help='Generate Proforma Invoice PDF')
    parser.add_argument('--pdf', action='store_true', help='Generate PDF quotation')
    
    # Packing / Invoice group
    packing_grp = parser.add_argument_group('Packing & Invoice')
    packing_grp.add_argument('--packing', action='store_true', help='Generate packing list + commercial invoice')
    packing_grp.add_argument('--buyer-name', default='', help='Buyer name for packing/invoice')
    packing_grp.add_argument('--buyer-address', default='', help='Buyer address')
    packing_grp.add_argument('--port-loading', default='Qingdao', help='Port of loading (default: Qingdao)')
    packing_grp.add_argument('--port-discharge', default='', help='Port of discharge (required for packing)')
    packing_grp.add_argument('--vessel', default='', help='Vessel name / voyage')
    packing_grp.add_argument('--bl-no', default='', help='Bill of lading number')
    packing_grp.add_argument('--trade-terms', default='FOB', help='Trade terms (default: FOB)')
    packing_grp.add_argument('--no-interactive', action='store_true', help='Skip interactive input, fail on missing fields')
    
    return parser.parse_args()


def detect_parser_type(file_path: str, wb=None):
    """
    检测文件类型,返回最适合的解析器。
    如果传入了 wb，则复用（read_only 亦可）；否则自行加载（read_only=True）并关闭。
    返回 (parser_type: str, wb_or_None)
    """
    close_wb = False
    if wb is None:
        from openpyxl import load_workbook
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            close_wb = True
        except:
            return 'default', None
    
    # 检查sheet名和内容
    has_model_marker = False
    has_invoice = False
    has_price_table = False
    has_table_layout = False
    has_single = True
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Model:/型号: 标记
        for r in range(1, min(25, ws.max_row + 1)):
            for c in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(r, c).value
                if val:
                    val_str = str(val).strip().lower()
                    if val_str in ['model:', '型号:', 'item:']:
                        has_model_marker = True
                    if 'proforma invoice' in val_str or 'description of goods' in val_str:
                        has_invoice = True
                    if '车型' in val_str or 'model\nmodel' in val_str.replace(' ', ''):
                        has_table_layout = True
        
        # 价格表: 列少+有价格+有型号（不限行数，合同类文件可能超过30行）
        if ws.max_column < 12:
            has_price_kw = False
            has_model_kw = False
            for r in range(1, min(25, ws.max_row + 1)):
                for c in range(1, min(15, ws.max_column + 1)):
                    val = ws.cell(r, c).value
                    if val:
                        val_str = str(val).lower()
                        if 'price' in val_str:
                            has_price_kw = True
                        if 'model' in val_str:
                            has_model_kw = True
            if has_price_kw and has_model_kw:
                has_price_table = True
        
        # 单产品规格
        if ws.max_row > 10:
            if ws.max_column < 6:
                first_col = [ws.cell(r, 1).value for r in range(2, min(12, ws.max_row))]
                if not any(first_col):
                    has_single = False
            else:
                has_single = False  # 列数>=6 → 多产品表（如合同/报价单）
    
    if close_wb:
        wb.close()
    
    # 返回最适合的解析器
    if has_table_layout:
        return 'table', None
    if has_model_marker:
        return 'param_price', None
    elif has_invoice:
        return 'invoice', None
    elif has_price_table:
        return 'price_table', None
    elif has_single:
        return 'single_spec', None
    else:
        return 'default', None


def parse_file(file_path: str, parser_type: str = None, verbose: bool = False, wb=None) -> pd.DataFrame:
    """解析单个文件
    Args:
        wb: 预加载的 openpyxl Workbook（Excel 解析时复用）
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    # PDF 解析
    if ext == '.pdf':
        df = extract_products_from_pdf_v2(file_path)
        if df is not None and len(df) > 0:
            df['_source_file'] = os.path.basename(file_path)
        return df
    
    # DOCX 解析
    if ext == '.docx':
        df = extract_products_from_docx(file_path)
        if df is not None and len(df) > 0:
            df['_source_file'] = os.path.basename(file_path)
        return df
    
    # Excel 解析
    if parser_type is None:
        parser_type, _ = detect_parser_type(file_path, wb=wb)
    
    if parser_type == 'param_price':
        df = parse_param_price(file_path)
    elif parser_type == 'invoice':
        df = parse_invoice(file_path)
    elif parser_type == 'price_table':
        df = parse_price_table(file_path)
        # 如果 price_table 解析结果不合理（型号不含字母+数字组合），回退到 parse_excel_v3
        if df is not None and len(df) > 0 and 'model' in df.columns:
            import re
            _real_model = lambda m: bool(re.search(r'[A-Za-z]', m) and re.search(r'\d', m))
            real_models = sum(1 for m in df['model'].astype(str).str.strip() if _real_model(m))
            if real_models < 2:
                df = parse_excel_v3(file_path, wb=wb)
    elif parser_type == 'single_spec':
        df = parse_single_spec(file_path)
    elif parser_type == 'table':
        df = parse_table(file_path)
    else:
        df = parse_excel_v3(file_path, wb=wb)
    
    # 格式化spec - 保留原始格式
    if df is not None and len(df) > 0 and 'spec_zh' in df.columns:
        df['spec_zh'] = df['spec_zh'].fillna('')
    
    # Excel 嵌入式图片匹配（某些专用解析器可能没做）
    if df is not None and len(df) > 0 and ext in ('.xlsx', '.xls'):
        has_images = '_image_path' in df.columns and df['_image_path'].notna().any()
        if not has_images:
            try:
                from src.core.image import match_images_to_products
                df = match_images_to_products(df, file_path)
            except Exception:
                pass
    
    return df


def parse_folder(folder_path: str, verbose: bool = False) -> pd.DataFrame:
    """解析文件夹中的所有Excel和PDF文件"""
    files = []
    for ext in ['*.xlsx', '*.xls', '*.pdf']:
        files.extend(glob.glob(os.path.join(folder_path, ext)))
        files.extend(glob.glob(os.path.join(folder_path, '*', ext)))
    
    all_dfs = []
    
    for f in files:
        try:
            df = parse_file(f, verbose=verbose)
            if df is not None and len(df) > 0:
                df['_source_file'] = os.path.basename(f)
                all_dfs.append(df)
                if verbose:
                    print(f"  OK {os.path.basename(f)}: {len(df)} products")
        except Exception as e:
            if verbose:
                print(f"  FAIL {os.path.basename(f)}: {e}")
    
    if not all_dfs:
        return pd.DataFrame()
    
    return pd.concat(all_dfs, ignore_index=True)


def group_output(df: pd.DataFrame, group_by: str) -> dict:
    """分组输出"""
    if group_by == 'file' or group_by == 'sheet':
        return {k: v for k, v in df.groupby('_source_file')}
    elif group_by == 'category' and 'category' in df.columns:
        return {k: v for k, v in df.groupby('category')}
    else:
        # sort - 按model排序
        df_sorted = df.sort_values('model')
        return {'all': df_sorted}


def associate_external_images(df: pd.DataFrame, input_dir: str, verbose: bool = False) -> pd.DataFrame:
    """为缺少图片的产品关联外部图片 - 按来源文件分组
    
    规则: 
    - param_price.xlsx → images/param_price/
    - 车型价格表.xlsx → images/车型价格表.../
    - e-motorcycle.pdf → images/e-motorcycle/
    - 找不到对应目录时 fallback 到 images/ 全部图片池
    """
    import glob as g
    
    if df is None or len(df) == 0:
        return df
    
    if '_image_path' not in df.columns:
        df['_image_path'] = ''
        df['_image_path'] = df['_image_path'].astype(object)
    else:
        df['_image_path'] = df['_image_path'].astype(object)
    
    # 构建源文件 → 图片目录映射
    images_base = os.path.join(input_dir, 'images')
    source_to_img_dir = {}
    if os.path.isdir(images_base):
        for item in os.listdir(images_base):
            sub = os.path.join(images_base, item)
            if os.path.isdir(sub):
                key = item.lower().replace('_', '').replace('-', '').replace(' ', '')
                source_to_img_dir[key] = sub
    
    # 收集全部图片作为 fallback 池
    all_fallback_images = []
    if os.path.isdir(images_base):
        for ext in ['*.jpg', '*.jpeg', '*.png', '*.JPG', '*.JPEG', '*.PNG']:
            all_fallback_images.extend(g.glob(os.path.join(images_base, '**', ext), recursive=True))
    # 也包含 temp_images 中提取的图片
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_images')
    if os.path.isdir(temp_dir):
        for ext in ['*.jpg', '*.jpeg', '*.png', '*.JPG', '*.JPEG', '*.PNG']:
            all_fallback_images.extend(g.glob(os.path.join(temp_dir, ext)))
    
    def get_img_dir_for_source(source_name: str) -> str:
        if not source_name:
            return ''
        key = os.path.splitext(source_name)[0].lower().replace('_', '').replace('-', '').replace(' ', '')
        if key in source_to_img_dir:
            return source_to_img_dir[key]
        for dk, dv in source_to_img_dir.items():
            if key in dk or dk in key:
                return dv
        return ''
    
    fallback_idx = 0
    
    # 按来源分组处理
    for source_file in df['_source_file'].unique() if '_source_file' in df.columns else ['unknown']:
        mask = (df['_source_file'] == source_file) if '_source_file' in df.columns else slice(None)
        empty_idx = df[mask][df['_image_path'].isna() | (df['_image_path'] == '')].index
        
        if len(empty_idx) == 0:
            continue
        
        img_dir = get_img_dir_for_source(source_file)
        
        # 收集该组对应的图片
        group_images = []
        if img_dir and os.path.isdir(img_dir):
            for ext in ['*.jpg', '*.jpeg', '*.png', '*.JPG', '*.JPEG', '*.PNG']:
                group_images.extend(g.glob(os.path.join(img_dir, ext)))
        
        if group_images:
            # 尝试按型号匹配文件名
            unmatched_products = []
            for idx in empty_idx:
                model = str(df.loc[idx, 'model']).strip().lower()
                if not model:
                    unmatched_products.append(idx)
                    continue
                matched = False
                for img_path in group_images:
                    img_name = os.path.splitext(os.path.basename(img_path))[0].lower()
                    if model in img_name or img_name in model:
                        df.loc[idx, '_image_path'] = img_path
                        matched = True
                        break
                if not matched:
                    unmatched_products.append(idx)
            # 型号没匹配到的也不循环分配，留空
        else:
            # 没有对应目录 → 从 fallback 池按型号匹配
            for idx in empty_idx:
                model = str(df.loc[idx, 'model']).strip().lower()
                if not model:
                    continue
                for fi in range(fallback_idx, len(all_fallback_images)):
                    img_name = os.path.splitext(os.path.basename(all_fallback_images[fi]))[0].lower()
                    if model in img_name or img_name in model:
                        df.loc[idx, '_image_path'] = all_fallback_images[fi]
                        fallback_idx = fi + 1
                        break
    
    return df


def main():
    args = parse_args()
    
    print(f"Product Tool - Starting...")
    print(f"  Input: {args.input}")
    print(f"  Output: {args.output}")
    
    # 确定输入
    if os.path.isfile(args.input):
        files = [args.input]
    elif os.path.isdir(args.input):
        files = glob.glob(os.path.join(args.input, '*.xlsx'))
        files.extend(glob.glob(os.path.join(args.input, '*.xls')))
        files.extend(glob.glob(os.path.join(args.input, '*.pdf')))  # 新增 PDF
    else:
        print(f"Error: Input path not found: {args.input}")
        sys.exit(1)
    
    if not files:
        print(f"Error: No Excel files found")
        sys.exit(1)
    
    # 解析
    if len(files) == 1:
        df = parse_file(files[0], verbose=args.verbose)
        if args.verbose:
            print(f"  Parsed: {len(df)} products")
    else:
        df = parse_folder(args.input, verbose=args.verbose)
        print(f"  Total: {len(df)} products from {len(files)} files")
    
    if df is None or len(df) == 0:
        print(f"Error: No products parsed")
        sys.exit(1)
    
    # 保存到产品库
    if args.save_to_db or args.merge:
        print(f"\nSaving to product library...")
        
        # 导入产品库
        from src.product_manage import import_from_df, init_db
        
        init_db()
        
        # 合并去重
        if args.merge:
            from src.dedup_engine import dedup_dataframe
            df = dedup_dataframe(df)
            print(f"  After dedup: {len(df)} products")
        
        # 关联外部图片 (在 dedup 之后)
        if os.path.isdir(args.input):
            print(f"  Associating external images...")
            df = associate_external_images(df, args.input, verbose=args.verbose)
            img_count = (df['_image_path'].notna() & (df['_image_path'] != '')).sum() if '_image_path' in df.columns else 0
            print(f"  Associated images: {img_count}/{len(df)}")
            if args.verbose and img_count > 0:
                first = df[df['_image_path'].notna() & (df['_image_path'] != '')]
                if len(first) > 0:
                    print(f"    Sample: {first.iloc[0]['model']}: {first.iloc[0]['_image_path']}")
        
        result = import_from_df(df, user_id=args.user_id)
        print(f"  Import result: {result}")
        
        print(f"  Saved to product library (user: {args.user_id})")
    
    # 分类
    if not args.no_category and args.category:
        try:
            categories = load_categories(args.category)
            df = categorize_data(df.to_dict('records'), categories=categories)
            df = pd.DataFrame(df)
            print(f"  Categorized: {df['category'].value_counts().to_dict()}")
        except Exception as e:
            if args.verbose:
                print(f"  Category error: {e}")
    
    # 翻译 (当 --lang english 或 bilingual)
    if args.lang in ('english', 'bilingual'):
        print(f"  Translating to English...")
        if 'name_zh' in df.columns:
            df['name_en'] = df['name_zh'].apply(
                lambda x: translate_text(str(x)) if pd.notna(x) and str(x).strip() else ''
            )
        if 'spec_zh' in df.columns:
            translate_spec = lambda x: translate_text(str(x)) if pd.notna(x) and str(x).strip() else ''
            if args.lang == 'bilingual':
                df['spec_en'] = df['spec_zh'].apply(translate_spec)
            else:
                df['spec_zh'] = df['spec_zh'].apply(translate_spec)
    
    # 确保输出目录
    os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)
    
    # 输出格式
    if args.quotation:
        output_path = args.output
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
        
        # 转换为报价单格式
        data = df.to_dict('records')
        create_quotation(data, output_path, supplier=args.supplier, lang=args.lang)
        print(f"  Saved: {output_path}")
    else:
        # 标准Excel输出（带样式）
        output_path = args.output
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
        
        from src.output.excel_writer import write_excel_styled
        data = df.to_dict('records')
        write_excel_styled(output_path, data, sheet_name='Products')
        print(f"  Saved: {output_path}")
    
    # --pi: Generate Proforma Invoice PDF
    if args.pi:
        if not _HAS_PI_DEPS:
            print("  PI: missing dependencies (weasyprint/fpdf2). Skipping.")
        else:
            try:
                if '_source_file' in df.columns:
                    df = df.drop(columns=['_source_file'])
                buyer_info = {
                    'company': args.supplier or 'Buyer',
                    'address': '',
                    'contact': '',
                }
                pi_path = args.output.replace('.xlsx', '_PI.pdf') if args.output.endswith('.xlsx') else args.output + '_PI.pdf'
                result = generate_pi(df, buyer_info, DEFAULT_SELLER_INFO, pi_path, use_rmb=True)
                if result:
                    print(f"  PI saved: {result}")
            except Exception as e:
                print(f"  PI error: {e}")
    
    # --pdf: Generate PDF quotation
    if args.pdf:
        if not _HAS_PI_DEPS:
            print("  PDF: missing dependencies (weasyprint). Skipping.")
        else:
            try:
                from src.output.pdf_generator import create_quote_pdf
                
                # Convert DataFrame to list of dicts
                records = df.to_dict('records')
                # Clean up internal columns
                cleaned = []
                for r in records:
                    cleaned.append({k: v for k, v in r.items() if not k.startswith('_')})
                
                pdf_path = args.output.replace('.xlsx', '_Quote.pdf') if args.output.endswith('.xlsx') else args.output + '_Quote.pdf'
                result = create_quote_pdf(cleaned, pdf_path)
                if result:
                    print(f"  PDF saved: {result}")
            except Exception as e:
                print(f"  PDF error: {e}")
    
    # --packing: Generate packing list + commercial invoice
    if args.packing:
        if not _HAS_PI_DEPS:
            print("  Packing/Invoice: missing dependencies. Skipping.")
        else:
            try:
                # Collect buyer & shipping info
                buyer_name = args.buyer_name
                buyer_address = args.buyer_address
                port_loading = args.port_loading or 'Qingdao'
                port_discharge = args.port_discharge
                vessel = args.vessel
                bl_no = args.bl_no
                trade_terms = args.trade_terms or 'FOB'
                is_interactive = not args.no_interactive and sys.stdin.isatty()

                # Interactive input for missing required fields
                if is_interactive:
                    if not buyer_name:
                        buyer_name = input("  Buyer name: ").strip()
                    if not port_discharge:
                        port_discharge = input("  Port of discharge [Hamburg]: ").strip() or 'Hamburg'
                    if not buyer_address:
                        addr = input("  Buyer address (optional): ").strip()
                        if addr:
                            buyer_address = addr
                    if not vessel:
                        v = input("  Vessel name (optional): ").strip()
                        if v:
                            vessel = v

                # Validate required fields
                required_missing = []
                if not buyer_name:
                    required_missing.append('--buyer-name')
                if not port_discharge:
                    required_missing.append('--port-discharge')
                if required_missing:
                    print(f"  Error: Required fields missing: {', '.join(required_missing)}")
                    if not args.no_interactive:
                        print("  Use --no-interactive for non-interactive mode.")
                    return

                # Save to packing cache for reuse
                try:
                    cache_path = os.path.join(os.path.dirname(args.output) or '.', 'packing_cache.json')
                    cache = {}
                    if os.path.exists(cache_path):
                        with open(cache_path, 'r', encoding='utf-8') as f:
                            cache = json.load(f)
                    cache[buyer_name] = {'address': buyer_address, 'port_of_discharge': port_discharge, 'vessel': vessel}
                    with open(cache_path, 'w', encoding='utf-8') as f:
                        json.dump(cache, f, ensure_ascii=False, indent=2)
                except Exception:
                    pass

                # Load company config
                from src.config import DEFAULT_SELLER_INFO
                company_config = DEFAULT_SELLER_INFO.copy() if DEFAULT_SELLER_INFO else {}
                try:
                    from src.company import load_company_config
                    cfg = load_company_config()
                    if cfg:
                        company_config.update(cfg)
                except Exception:
                    pass

                pi_items = df.to_dict('records')
                packing_path = args.output.replace('.xlsx', '_PackingList.xlsx') if args.output.endswith('.xlsx') else args.output + '_PackingList.xlsx'
                inv_path = args.output.replace('.xlsx', '_Invoice.xlsx') if args.output.endswith('.xlsx') else args.output + '_Invoice.xlsx'
                
                from datetime import datetime
                inv_no = datetime.now().strftime('INV%Y%m%d%H%M%S')
                inv_date = datetime.now().strftime('%Y-%m-%d')
                
                p_result = generate_packing_list(
                    pi_items=pi_items,
                    invoice_no=inv_no,
                    invoice_date=inv_date,
                    buyer_name=buyer_name,
                    buyer_address=buyer_address,
                    port_loading=port_loading,
                    port_discharge=port_discharge,
                    vessel=vessel,
                    bl_no=bl_no,
                    trade_terms=trade_terms,
                    company_config=company_config,
                    output_path=packing_path,
                )
                if p_result:
                    print(f"  Packing list saved: {p_result}")
                
                c_result = generate_commercial_invoice(
                    pi_items=pi_items,
                    invoice_no=inv_no,
                    invoice_date=inv_date,
                    buyer_name=buyer_name,
                    buyer_address=buyer_address,
                    port_loading=port_loading,
                    port_discharge=port_discharge,
                    vessel=vessel,
                    bl_no=bl_no,
                    trade_terms=trade_terms,
                    company_config=company_config,
                    output_path=inv_path,
                )
                if c_result:
                    print(f"  Commercial invoice saved: {c_result}")
            except Exception as e:
                import traceback
                print(f"  Packing/Invoice error: {e}")
                traceback.print_exc()
    
    # 统计
    print(f"\nDone! {len(df)} products")
    
    if 'price_rmb' in df.columns:
        has_price = df['price_rmb'].notna().sum()
        print(f"  With price: {has_price}/{len(df)} ({has_price*100//len(df)}%)")


if __name__ == '__main__':
    main()